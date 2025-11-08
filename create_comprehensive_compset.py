import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Compset Analysis"

# Define styles
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=10)
hotel_header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
hotel_header_font = Font(bold=True, color="FFFFFF", size=11)
category_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
category_font = Font(bold=True, size=10)
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
wrap_alignment = Alignment(wrap_text=True, vertical='top')

# Comprehensive hotel data
hotels_data = {
    "Grand Millennium Dubai": {
        "Address": "Sheikh Zayed Road, Barsha Heights (Tecom), Dubai",
        "Distance from Hotel (KM)": "0",
        "Property Type": "City",
        "Star Rating": "5",
        "Opening Year": "N/A",
        "Last Renovation": "N/A",
        "Brand": "Millennium Hotels and Resorts",
        "Owning Company/Management": "Millennium Hotels and Resorts (MHR) / City Developments Limited (CDL)",
        "Licensed for Alcohol (Y/N)": "Y",
        "Licensed for Shisha (Y/N)": "Y",
        "Total Number of Keys": "339",
        "Type of Hotel": "Mixed (Rooms, Suites, Apartments)",
        "Room Types Detail": "Superior Room (33m²), Deluxe Room (33m²), Club Room (33m²), Business Suite (60m²), Executive Suite (66m²), Studio Suite/Apartment (54m²), One Bedroom Suite/Apartment (92m²), Two Bedroom Suite/Apartment (111m²), Royal Suite (120m²), Penthouse (335m²)",
        "Percent Suites": "Multiple suite categories available",
        "Restaurants": "3 restaurants: The Atrium (international casual dining), Asian Restaurant (18th floor with live cooking stations), Belgian/European restaurant",
        "Bars": "2 bars: Crystal Bar (stylish lounge bar with 3am license), Poolside Bar, Belgian Beer Café",
        "Lounge/Coffee Shop": "Hotel lobby lounge, coffee facilities",
        "Total Number of Function Rooms": "7",
        "Total Function Space (sqm)": "N/A",
        "Number of Ballroom": "1",
        "Number of Meeting rooms": "7",
        "Largest Room/Ballroom (sqm)": "N/A",
        "Maximum capacity (theater style)": "N/A",
        "Spa Name": "Jasmine Spa",
        "Treatment rooms (#)": "9",
        "Spa total area (sqm)": "N/A",
        "Health Club / Fitness (Y/N)": "Y - Bodytecture Fitness Center",
        "Kids Club": "Y - On-site play area (dedicated team)",
        "Kid Amenities": "Play area, babysitting services",
        "Swimming Pool": "Y - Temperature-controlled rooftop pool",
        "Direct Beach": "N - Shuttle to Zero Gravity beach (Mon, Tue, Thu)",
        "Retail shops": "Y - Gift shops/newsstands",
        "Additional Facilities": "Sauna, steam room, outdoor & indoor Jacuzzi, 24-hour room service, free parking"
    },
    "Zabeel House The Greens": {
        "Address": "The Onyx Tower 3, The Greens, Sheikh Zayed Road, Dubai",
        "Distance from Hotel (KM)": "~3-5",
        "Property Type": "City",
        "Star Rating": "4",
        "Opening Year": "2018 (Q4)",
        "Last Renovation": "N/A",
        "Brand": "Jumeirah Group",
        "Owning Company/Management": "The Onyx for Development (Ishraqah) / Managed by Jumeirah Group",
        "Licensed for Alcohol (Y/N)": "Y",
        "Licensed for Shisha (Y/N)": "Y",
        "Total Number of Keys": "210",
        "Type of Hotel": "Mixed (Rooms, Studios, Apartments)",
        "Room Types Detail": "Popular Room (28m²), Popular Balcony Room (28m²), Studio Room (28m²), The Apartment/Suite (28m²), Connected Rooms (56m²)",
        "Percent Suites": "Mix of rooms and studio apartments",
        "Restaurants": "2 restaurants: Social Company (artisanal co-working cafe/bar, 6:30 AM-midnight), Lah Lah (Pan-Asian restaurant with rooftop shisha terrace)",
        "Bars": "Integrated within restaurants - Social Company bar, Lah Lah bar/lounge",
        "Lounge/Coffee Shop": "Social Company (co-working cafe)",
        "Total Number of Function Rooms": "4",
        "Total Function Space (sqm)": "~156 sqm (1,679 sq ft largest)",
        "Number of Ballroom": "0",
        "Number of Meeting rooms": "4",
        "Largest Room/Ballroom (sqm)": "156 sqm (1,679 sq ft)",
        "Maximum capacity (theater style)": "200 (outdoor space)",
        "Spa Name": "Native Club by Soul Senses",
        "Treatment rooms (#)": "N/A",
        "Spa total area (sqm)": "N/A",
        "Health Club / Fitness (Y/N)": "Y - Fully equipped gym with indoor/outdoor functional training zone",
        "Kids Club": "Y - Dedicated kids area, Play House",
        "Kid Amenities": "Kids club, play house, game room, kids pool area",
        "Swimming Pool": "Y - Outdoor pool on 4th floor (Native Club)",
        "Direct Beach": "N - Access to Jumeirah Zabeel Saray beach (discounted vouchers)",
        "Retail shops": "Y - Grocery/convenience store",
        "Additional Facilities": "2 outdoor padel courts, outdoor gym, Native Club facilities (gym 6am-9:30pm, spa 11am-9pm)"
    },
    "voco Bonnington Dubai": {
        "Address": "Cluster J, Jumeirah Lakes Towers, Dubai",
        "Distance from Hotel (KM)": "~2-3",
        "Property Type": "City",
        "Star Rating": "5",
        "Opening Year": "N/A (Recently rebranded to voco by IHG)",
        "Last Renovation": "2022 (rebrand)",
        "Brand": "voco (IHG Hotels & Resorts)",
        "Owning Company/Management": "Managed by IHG Hotels & Resorts",
        "Licensed for Alcohol (Y/N)": "Y",
        "Licensed for Shisha (Y/N)": "Y",
        "Total Number of Keys": "208",
        "Type of Hotel": "Rooms",
        "Room Types Detail": "Classic Guest Room, Executive Room (4 distinct categories total with blackout blinds, rainfall shower, bathtub)",
        "Percent Suites": "N/A",
        "Restaurants": "5 F&B outlets: The Cavendish Restaurant (British-European with Middle Eastern), McGettigan's Irish Pub (award-winning), The Cheeky Camel (restaurant/bar), The Authors' Lounge (cafe with teas/coffees/pastries), Coffee Box",
        "Bars": "McGettigan's Irish Pub, The Cheeky Camel, Pool Bar/Leisure Deck",
        "Lounge/Coffee Shop": "The Authors' Lounge, Coffee Box",
        "Total Number of Function Rooms": "6",
        "Total Function Space (sqm)": "N/A",
        "Number of Ballroom": "0",
        "Number of Meeting rooms": "6",
        "Largest Room/Ballroom (sqm)": "Kinsale room (capacity 120)",
        "Maximum capacity (theater style)": "120",
        "Spa Name": "Beauty salon/spa on 11th floor Leisure Deck",
        "Treatment rooms (#)": "N/A",
        "Spa total area (sqm)": "N/A",
        "Health Club / Fitness (Y/N)": "Y - Modern gym on 11th floor",
        "Kids Club": "N",
        "Kid Amenities": "N/A",
        "Swimming Pool": "Y - 11th floor infinity pool with Dubai skyline views",
        "Direct Beach": "N - Complimentary shuttle to Riva Beach Club",
        "Retail shops": "N/A",
        "Additional Facilities": "Sauna, steam rooms, beauty salon with spa treatments (facials, body treatments, manicures, pedicures, massage)"
    },
    "Radisson Blu Media City": {
        "Address": "Dubai Media City, Dubai",
        "Distance from Hotel (KM)": "~2-3",
        "Property Type": "City",
        "Star Rating": "4",
        "Opening Year": "N/A",
        "Last Renovation": "N/A",
        "Brand": "Radisson Hotel Group",
        "Owning Company/Management": "Radisson Hotel Group",
        "Licensed for Alcohol (Y/N)": "Y",
        "Licensed for Shisha (Y/N)": "N/A",
        "Total Number of Keys": "246",
        "Type of Hotel": "Mixed (Rooms, Suites)",
        "Room Types Detail": "Standard Room, Executive Room (with mini-bar, Executive Lounge access, complimentary breakfast), One Bedroom Suite with Lounge Access, Connected Rooms (for families)",
        "Percent Suites": "Limited suite inventory",
        "Restaurants": "2 restaurants: Chef's House (breakfast buffet), Certo Italian Restaurant, Tamanya Goes Thai",
        "Bars": "2 bars: ICON Sports Bar, additional bar",
        "Lounge/Coffee Shop": "2 coffee shops/cafes: Jones the Grocer Express",
        "Total Number of Function Rooms": "12-13",
        "Total Function Space (sqm)": "582 sqm (6,456 sq ft)",
        "Number of Ballroom": "1",
        "Number of Meeting rooms": "12",
        "Largest Room/Ballroom (sqm)": "Al Nada ballroom (divisible into 2 spaces)",
        "Maximum capacity (theater style)": "150",
        "Spa Name": "Dreamworks SPA / Senso Wellness Centre",
        "Treatment rooms (#)": "5 Asian-themed treatment rooms",
        "Spa total area (sqm)": "N/A",
        "Health Club / Fitness (Y/N)": "Y - State-of-the-art gym with fitness classes",
        "Kids Club": "N",
        "Kid Amenities": "Children welcome, no specific kids club",
        "Swimming Pool": "Y - 2 outdoor pools (free to use)",
        "Direct Beach": "N - Public beach 20km away with shuttle bus (Jumeirah Beach Park)",
        "Retail shops": "N/A",
        "Additional Facilities": "Sauna, steam bath, massage room, Senso Terrace (poolside events for up to 60 guests), 24-hour room service"
    },
    "Media One Hotel": {
        "Address": "Dubai Media City, Dubai",
        "Distance from Hotel (KM)": "~2-3",
        "Property Type": "City/Lifestyle",
        "Star Rating": "4",
        "Opening Year": "Operating over a decade",
        "Last Renovation": "N/A",
        "Brand": "Independent",
        "Owning Company/Management": "Independent mixed-use property",
        "Licensed for Alcohol (Y/N)": "Y",
        "Licensed for Shisha (Y/N)": "Y",
        "Total Number of Keys": "260-264",
        "Type of Hotel": "Mixed (Rooms, Studios, Suites)",
        "Room Types Detail": "Hip Room/Standard (30m²), Hip Urban/City Deluxe (30m²), Business Room (30m²), Executive Room (35m²), Studio (50m²), Suite (70m²)",
        "Percent Suites": "Mix with studio and suite options",
        "Restaurants": "4 on-site restaurants serving international cuisine, all-day-dining restaurant, additional themed restaurants",
        "Bars": "2 bars + outdoor lounge: Pool bar (8th floor), Q43 (43rd floor with billiards), Friends-themed lounge, nightclub",
        "Lounge/Coffee Shop": "Coffee shop with co-working desks, lobby lounge",
        "Total Number of Function Rooms": "7 + 42nd floor space",
        "Total Function Space (sqm)": "~702 sqm (7,553 sq ft)",
        "Number of Ballroom": "0",
        "Number of Meeting rooms": "7 (7-Squared on 7th floor) + 42nd floor multi-purpose space",
        "Maximum capacity (theater style)": "200+ (42nd floor)",
        "Spa Name": "Spa services (2 minutes from hotel - may be off-site)",
        "Treatment rooms (#)": "N/A",
        "Spa total area (sqm)": "N/A",
        "Health Club / Fitness (Y/N)": "Y - 24/7 gymnasium on 8th floor",
        "Kids Club": "N - Babysitting services available",
        "Kid Amenities": "Kids menu, babysitting services",
        "Swimming Pool": "Y - Large outdoor rooftop pool on 8th floor",
        "Direct Beach": "N - Free shuttle to Peaches & Cream beach club",
        "Retail shops": "Y - Number of retail outlets",
        "Additional Facilities": "Sauna, steam rooms, P7 Arena (underground entertainment space), co-working spaces, 9 F&B venues, 230,000 sq ft office space"
    },
    "Pullman JLT": {
        "Address": "Sheikh Zayed Road, Jumeirah Lakes Towers (near Sobha Realty Metro), Dubai",
        "Distance from Hotel (KM)": "~2-3",
        "Property Type": "City",
        "Star Rating": "5",
        "Opening Year": "2015",
        "Last Renovation": "N/A",
        "Brand": "Pullman (Accor Hotels)",
        "Owning Company/Management": "Managed by Accor Hotels",
        "Licensed for Alcohol (Y/N)": "Y",
        "Licensed for Shisha (Y/N)": "Y",
        "Total Number of Keys": "354 (including 76 suites/apartments)",
        "Type of Hotel": "Mixed (Rooms, Suites, Apartments)",
        "Room Types Detail": "Superior Room (35m²), Deluxe Room (40m²), Executive Room (35m²), Deluxe Executive Room (40m²), Studio Suite, 1-Bedroom Suite, 2-Bedroom Suite, 3-Bedroom Suite",
        "Percent Suites": "21% (76 out of 354)",
        "Restaurants": "5 restaurants/bars: Seasons Restaurant (all-day international dining), Manzoni Bistro & Bar (Italian, award-winning), La Vue (outdoor lounge), Terrace Bar & Lounge",
        "Bars": "Manzoni Bar, La Vue, Terrace Bar & Lounge, Rooftop bar",
        "Lounge/Coffee Shop": "La Vue (outdoor lounge), lounges in hotel",
        "Total Number of Function Rooms": "11",
        "Total Function Space (sqm)": "550 sqm",
        "Number of Ballroom": "0",
        "Number of Meeting rooms": "11",
        "Largest Room/Ballroom (sqm)": "N/A",
        "Maximum capacity (theater style)": "N/A",
        "Spa Name": "Fit and Spa Lounge (34th floor)",
        "Treatment rooms (#)": "5",
        "Spa total area (sqm)": "N/A",
        "Health Club / Fitness (Y/N)": "Y - Pullman Fit with Technogym equipment",
        "Kids Club": "Y - Kids' play area",
        "Kid Amenities": "Kids' play area with activities",
        "Swimming Pool": "Y - Rooftop pool",
        "Direct Beach": "N - Near Marina Beach, Dubai Marina less than 0.5 miles",
        "Retail shops": "N/A",
        "Additional Facilities": "Separate relaxation areas for men/women, steam & sauna facilities, Nespresso machines (Executive+), Bose docking stations (Executive+), Pool Hours: 8AM-7PM, Gym: 8AM-8PM, Spa: 12PM-9PM"
    },
    "Millennium Place Barsha Heights": {
        "Address": "Barsha Heights (Tecom), near Mall of the Emirates, Dubai",
        "Distance from Hotel (KM)": "~0.5-1",
        "Property Type": "City",
        "Star Rating": "4",
        "Opening Year": "Recent (brand new)",
        "Last Renovation": "N/A",
        "Brand": "Millennium Hotels and Resorts",
        "Owning Company/Management": "Millennium Hotels and Resorts",
        "Licensed for Alcohol (Y/N)": "N - No alcohol served",
        "Licensed for Shisha (Y/N)": "Y",
        "Total Number of Keys": "468 rooms + 447 apartments",
        "Type of Hotel": "Mixed (Hotel Rooms + Serviced Apartments)",
        "Room Types Detail": "Superior Room, Deluxe Room, Deluxe Sky Room, Premium Room, Premium Sky Room, Deluxe Suite, Two-Bedroom Suite, Studio Apartment, 1-Bedroom Apartment, 2-Bedroom Apartment, Royal Penthouse",
        "Percent Suites": "Mix of hotel rooms and apartments",
        "Restaurants": "5 dining options: M One Restaurant (international buffet with themed nights), M Two Restaurant (apartment side), Eva's Fusion, additional outlets",
        "Bars": "0 bars (no alcohol)",
        "Lounge/Coffee Shop": "Twenty9 Lounge (level 29 rooftop with shisha), Level Social Lobby Lounge, Splash Pool Bar & Café",
        "Total Number of Function Rooms": "7",
        "Total Function Space (sqm)": "146+ sqm (The M Ballroom)",
        "Number of Ballroom": "1 (The M Ballroom - divisible)",
        "Number of Meeting rooms": "7 (including Inspiration, Motivation rooms)",
        "Largest Room/Ballroom (sqm)": "146 sqm (The M + 43 sqm pre-function)",
        "Maximum capacity (theater style)": "170",
        "Spa Name": "Spa treatment rooms",
        "Treatment rooms (#)": "N/A",
        "Spa total area (sqm)": "N/A",
        "Health Club / Fitness (Y/N)": "Y - Fitness Centre on 29th floor",
        "Kids Club": "Y - Kids Club with indoor play area",
        "Kid Amenities": "Kids Club, indoor play area, children's pool",
        "Swimming Pool": "Y - 4 pools (3 outdoor + children's pool)",
        "Direct Beach": "N - Shuttle to Kite Beach (AED 10/person), minutes from Jumeirah Beach",
        "Retail shops": "N/A",
        "Additional Facilities": "Spa therapy, fitness classes, wellness centre, cycling facilities, 4K Laser Projector in ballroom, highspeed WiFi, USB charging, 43-inch IPTV"
    },
    "TRYP by Wyndham Dubai": {
        "Address": "Barsha Heights (Tecom), Dubai",
        "Distance from Hotel (KM)": "~0.5-1",
        "Property Type": "City/Lifestyle",
        "Star Rating": "4",
        "Opening Year": "2017",
        "Last Renovation": "N/A",
        "Brand": "TRYP (Wyndham Hotels & Resorts)",
        "Owning Company/Management": "Wyndham Hotels & Resorts (many apartments privately owned)",
        "Licensed for Alcohol (Y/N)": "Y",
        "Licensed for Shisha (Y/N)": "Y",
        "Total Number of Keys": "650",
        "Type of Hotel": "Mixed (Rooms, Suites)",
        "Room Types Detail": "TRYP Room, TRYP King Room, Triple Room, TRYP One Bedroom Suite (41 sqm/441 sqft), TRYP Premium Suite (loft-type), Accessible Room (20m²)",
        "Percent Suites": "Limited suite inventory",
        "Restaurants": "3 trendy restaurants: Local Restaurant (rustic eatery/bar, all-day with live sports/shisha/live music), Lola Taberna Espanola (Spanish/paella), Haze Lounge (Greek/Levant)",
        "Bars": "Local Restaurant bar (with happy hour, Ladies Night Wed), Pool bar",
        "Lounge/Coffee Shop": "Haze Lounge, Coffee shop, club lounge",
        "Total Number of Function Rooms": "1-2",
        "Total Function Space (sqm)": "357 sqm (3,843 sq ft largest)",
        "Number of Ballroom": "0",
        "Number of Meeting rooms": "1-2 (Local Space, NEST)",
        "Largest Room/Ballroom (sqm)": "357 sqm (3,843 sq ft)",
        "Maximum capacity (theater style)": "80 (conference), 30 (banquet)",
        "Spa Name": "Rayya Wellness Spa",
        "Treatment rooms (#)": "N/A",
        "Spa total area (sqm)": "N/A",
        "Health Club / Fitness (Y/N)": "Y - Abz Gym (cardio, free weights, elliptical)",
        "Kids Club": "N - Babysitting/childcare available",
        "Kid Amenities": "Kids pool, babysitting services, day-care centre",
        "Swimming Pool": "Y - Terrace pool (8AM-8PM) with sun deck & pool bar + kids pool",
        "Direct Beach": "N - Exclusive access to Soluna Beach Club at The Palm (fee applies, closed Jun-Aug 2025 for renovation)",
        "Retail shops": "N/A",
        "Additional Facilities": "Sauna, spa treatments (hot stone, Thai massage), tech-savvy co-working space, free shuttle to beach/mall"
    },
    "Mercure Dubai Barsha Heights": {
        "Address": "Sheikh Zayed Road, Barsha Heights (Tecom), Dubai",
        "Distance from Hotel (KM)": "~0.5-1",
        "Property Type": "City",
        "Star Rating": "4",
        "Opening Year": "N/A",
        "Last Renovation": "N/A",
        "Brand": "Mercure (Accor Hotels)",
        "Owning Company/Management": "Managed by Accor Hotels",
        "Licensed for Alcohol (Y/N)": "Y",
        "Licensed for Shisha (Y/N)": "Y",
        "Total Number of Keys": "408 suites (some sources say up to 1,015)",
        "Type of Hotel": "All Suites & Apartments",
        "Room Types Detail": "1-Bedroom Deluxe Suite (60m²), 2-Bedroom Prestige Suite (105m²) - all units are suites with living areas, kitchenettes, balconies",
        "Percent Suites": "100%",
        "Restaurants": "4 dining spots: Day & Night Restaurant (international buffet with live cooking, shisha), Eva's (mention unclear), in-room dining",
        "Bars": "The Exit Sports Bar (London-style, 8 screens with outdoor terrace)",
        "Lounge/Coffee Shop": "Café Social (lobby lounge with 50% discount after 8pm), Mixolounge/VIP Lounge (27th floor terrace), Corner 8 (pool bar on 8th floor)",
        "Total Number of Function Rooms": "Meeting rooms on 40th & 41st levels",
        "Total Function Space (sqm)": "N/A",
        "Number of Ballroom": "0",
        "Number of Meeting rooms": "Multiple on 40th & 41st levels",
        "Largest Room/Ballroom (sqm)": "N/A",
        "Maximum capacity (theater style)": "100 (cocktail), 50 (conference)",
        "Spa Name": "Spa with various treatments",
        "Treatment rooms (#)": "N/A",
        "Spa total area (sqm)": "N/A",
        "Health Club / Fitness (Y/N)": "Y - Fitness Centre (Excellent GymFactor rating)",
        "Kids Club": "Y - Kids' Club with indoor play area",
        "Kid Amenities": "Kids' Club, indoor play area, children's pool",
        "Swimming Pool": "Y - Rooftop pool terrace on 8th floor with hot tub & children's pool",
        "Direct Beach": "N - Free shuttle to Jumeirah Open beach (daily at 9am/10am/11am, return 1pm/2pm/3pm)",
        "Retail shops": "N/A",
        "Additional Facilities": "Hammam, sauna, Turkish steam bath, indoor pool, squash courts, indoor football court, kickboxing, salons, pool deck level 8, 41-story tower"
    }
}

# Create field labels based on the original Compset Tool template
field_labels = [
    ("GENERAL INFORMATION", "header"),
    ("Address", "data"),
    ("Distance from Hotel (KM)", "data"),
    ("Property Type", "data"),
    ("Star Rating", "data"),
    ("Opening Year", "data"),
    ("Last Renovation", "data"),
    ("Brand", "data"),
    ("Owning Company/Management", "data"),
    ("Licensed for Alcohol (Y/N)", "data"),
    ("Licensed for Shisha (Y/N)", "data"),
    ("", "blank"),
    ("ROOMS INFORMATION", "header"),
    ("Total Number of Keys", "data"),
    ("Type of Hotel", "data"),
    ("Room Types Detail", "data"),
    ("Percent Suites", "data"),
    ("", "blank"),
    ("FOOD & BEVERAGE", "header"),
    ("Restaurants", "data"),
    ("Bars", "data"),
    ("Lounge/Coffee Shop", "data"),
    ("", "blank"),
    ("MEETINGS & EVENTS", "header"),
    ("Total Number of Function Rooms", "data"),
    ("Total Function Space (sqm)", "data"),
    ("Number of Ballroom", "data"),
    ("Number of Meeting rooms", "data"),
    ("Largest Room/Ballroom (sqm)", "data"),
    ("Maximum capacity (theater style)", "data"),
    ("", "blank"),
    ("FITNESS & SPA", "header"),
    ("Spa Name", "data"),
    ("Treatment rooms (#)", "data"),
    ("Spa total area (sqm)", "data"),
    ("Health Club / Fitness (Y/N)", "data"),
    ("", "blank"),
    ("OTHER FACILITIES", "header"),
    ("Kids Club", "data"),
    ("Kid Amenities", "data"),
    ("Swimming Pool", "data"),
    ("Direct Beach", "data"),
    ("Retail shops", "data"),
    ("Additional Facilities", "data"),
]

# Set column A width
ws.column_dimensions['A'].width = 30

# Write field labels in column A
for idx, (label, label_type) in enumerate(field_labels, start=1):
    cell = ws.cell(row=idx, column=1)
    cell.value = label
    cell.border = border
    cell.alignment = wrap_alignment

    if label_type == "header":
        cell.fill = header_fill
        cell.font = header_font
    elif label_type == "blank":
        cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

# Write hotel data in columns B onwards
col_offset = 2
for hotel_name, hotel_data in hotels_data.items():
    col = col_offset

    # Set column width
    ws.column_dimensions[get_column_letter(col)].width = 40

    # Write hotel name as header
    cell = ws.cell(row=1, column=col)
    cell.value = hotel_name
    cell.fill = hotel_header_fill
    cell.font = hotel_header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border

    # Map data to rows
    row_idx = 2
    for label, label_type in field_labels[1:]:  # Skip first header
        cell = ws.cell(row=row_idx, column=col)
        cell.border = border
        cell.alignment = wrap_alignment

        if label_type == "data" and label in hotel_data:
            cell.value = hotel_data[label]
        elif label_type == "header":
            cell.fill = category_fill

        row_idx += 1

    col_offset += 1

# Freeze panes (freeze first column and first row)
ws.freeze_panes = 'B2'

# Save workbook
output_file = r"C:\Users\reservations\Desktop\Compset Tool\Comprehensive_Compset_Analysis.xlsx"
wb.save(output_file)
print(f"Comprehensive compset analysis Excel file created: {output_file}")
