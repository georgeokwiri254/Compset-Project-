#!/usr/bin/env python3
"""
Analyze the Excel sheets structure and identify what needs to be filled
"""

import pandas as pd
from openpyxl import load_workbook
import json

excel_path = '/home/gee_devops254/Downloads/Compset Tool/NEW - STR Competitor Set Analysis.xlsx'
wb = load_workbook(excel_path)

print("="*80)
print("EXCEL FILE STRUCTURE ANALYSIS")
print("="*80)

for sheet_name in wb.sheetnames:
    print(f"\n{'='*80}")
    print(f"SHEET: {sheet_name}")
    print(f"{'='*80}")
    ws = wb[sheet_name]

    # Print dimensions
    print(f"Dimensions: {ws.dimensions}")
    print(f"Max row: {ws.max_row}, Max column: {ws.max_column}")

    # Print first 15 rows
    print(f"\nFirst 15 rows (showing first 10 columns):")
    print("-" * 80)
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=15, max_col=10, values_only=True), 1):
        print(f"Row {i:2d}: {row}")

    # Check for empty cells in key areas
    empty_count = 0
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        for cell in row:
            if cell is None or str(cell).strip() == '':
                empty_count += 1

    print(f"\n✓ Total empty cells (excluding row 1): {empty_count}")

print("\n" + "="*80)
print("ANALYSIS COMPLETE")
print("="*80)
