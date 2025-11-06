import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference, RadarChart, PieChart, LineChart, ScatterChart
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
from openpyxl.styles import Font, PatternFill, Alignment
import openpyxl.styles
import json

# Load the Excel file
excel_file = r'C:\Users\reservations\Desktop\Compset Tool\Grand_Millennium_Dubai_CompSet_Analysis.xlsx'
wb = load_workbook(excel_file)

# Create or get the Visual Analysis sheet
if 'Compset Scoring Analysis' in wb.sheetnames:
    del wb['Compset Scoring Analysis']

visual_sheet = wb.create_sheet('Compset Scoring Analysis', 0)

# Extract data from Hotels sheet
ws = wb['Hotels']

# Prepare data for charts
hotel_names = []
distances = []
scores = []
recommendations = []
star_ratings = []
total_rooms_list = []
meeting_spaces = []
tripadvisor_ratings = []
booking_ratings = []

for row in range(2, ws.max_row + 1):
    name = ws.cell(row=row, column=2).value
    if name:
        hotel_names.append(name if len(name) <= 30 else name[:27] + "...")
        distances.append(ws.cell(row=row, column=5).value or 0)
        star_ratings.append(ws.cell(row=row, column=3).value or 0)

        score = ws.cell(row=row, column=34).value
        if isinstance(score, (int, float)):
            scores.append(score)
        else:
            scores.append(0)

        recommendation = ws.cell(row=row, column=35).value
        recommendations.append(recommendation if recommendation else "TBD")

        rooms = ws.cell(row=row, column=7).value
        if isinstance(rooms, (int, float)):
            total_rooms_list.append(rooms)
        else:
            total_rooms_list.append(0)

        meeting = ws.cell(row=row, column=12).value
        if isinstance(meeting, (int, float)):
            meeting_spaces.append(meeting)
        else:
            meeting_spaces.append(0)

        trip_rating = ws.cell(row=row, column=20).value
        if isinstance(trip_rating, (int, float)):
            tripadvisor_ratings.append(trip_rating)
        else:
            tripadvisor_ratings.append(0)

        book_rating = ws.cell(row=row, column=22).value
        if isinstance(book_rating, (int, float)):
            booking_ratings.append(book_rating)
        else:
            booking_ratings.append(0)

# Get top scored hotels (with actual scores)
hotel_score_pairs = [(name, score, dist, rooms, meeting, trip, book) for name, score, dist, rooms, meeting, trip, book in
                      zip(hotel_names, scores, distances, total_rooms_list, meeting_spaces, tripadvisor_ratings, booking_ratings)
                      if score > 0]
hotel_score_pairs.sort(key=lambda x: x[1], reverse=True)

# Take top 15 scored hotels for detailed analysis
top_hotels = hotel_score_pairs[:15]

# Add header and summary
visual_sheet['A1'] = 'Grand Millennium Dubai - Competitive Set Analysis Dashboard'
visual_sheet['A1'].font = openpyxl.styles.Font(bold=True, size=16)
visual_sheet.merge_cells('A1:H1')

visual_sheet['A3'] = 'Analysis Summary'
visual_sheet['A3'].font = openpyxl.styles.Font(bold=True, size=12)

visual_sheet['A4'] = f'Total Hotels Analyzed: {len(hotel_names)}'
visual_sheet['A5'] = f'Hotels with Complete Data: {len([s for s in scores if s > 0])}'
visual_sheet['A6'] = f'Primary Compset Candidates: {recommendations.count("Primary Compset")}'
visual_sheet['A7'] = f'Secondary Compset Candidates: {recommendations.count("Secondary Compset")}'
visual_sheet['A8'] = f'Extended Reference Hotels: {recommendations.count("Extended Reference")}'

# Create data table for top hotels
start_row = 10
visual_sheet['A10'] = 'Hotel Name'
visual_sheet['B10'] = 'Overall Score'
visual_sheet['C10'] = 'Distance (km)'
visual_sheet['D10'] = 'Total Rooms'
visual_sheet['E10'] = 'Meeting Space (sqm)'
visual_sheet['F10'] = 'TripAdvisor'
visual_sheet['G10'] = 'Booking.com'
visual_sheet['H10'] = 'Recommendation'

# Apply header formatting
for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
    cell = visual_sheet[f'{col}10']
    cell.font = openpyxl.styles.Font(bold=True)
    cell.fill = openpyxl.styles.PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    cell.font = openpyxl.styles.Font(bold=True, color='FFFFFF')

# Populate top hotels data
for idx, (name, score, dist, rooms, meeting, trip, book) in enumerate(top_hotels, start=11):
    visual_sheet[f'A{idx}'] = name
    visual_sheet[f'B{idx}'] = score
    visual_sheet[f'C{idx}'] = dist
    visual_sheet[f'D{idx}'] = rooms
    visual_sheet[f'E{idx}'] = meeting
    visual_sheet[f'F{idx}'] = trip
    visual_sheet[f'G{idx}'] = book

    # Determine recommendation
    if score >= 90:
        visual_sheet[f'H{idx}'] = "Primary Compset"
        visual_sheet[f'H{idx}'].fill = openpyxl.styles.PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    elif score >= 75:
        visual_sheet[f'H{idx}'] = "Secondary Compset"
        visual_sheet[f'H{idx}'].fill = openpyxl.styles.PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    elif score >= 60:
        visual_sheet[f'H{idx}'] = "Extended Reference"
        visual_sheet[f'H{idx}'].fill = openpyxl.styles.PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    else:
        visual_sheet[f'H{idx}'] = "Not Recommended"

# Adjust column widths
visual_sheet.column_dimensions['A'].width = 35
visual_sheet.column_dimensions['B'].width = 14
visual_sheet.column_dimensions['C'].width = 14
visual_sheet.column_dimensions['D'].width = 14
visual_sheet.column_dimensions['E'].width = 18
visual_sheet.column_dimensions['F'].width = 14
visual_sheet.column_dimensions['G'].width = 14
visual_sheet.column_dimensions['H'].width = 20

# Create Bar Chart: Top 10 Hotels by Overall Score
chart1 = BarChart()
chart1.type = "col"
chart1.style = 10
chart1.title = "Top 10 Hotels by Overall Compset Score"
chart1.y_axis.title = 'Overall Score'
chart1.x_axis.title = 'Hotel'

# Data for chart (top 10 hotels)
data = Reference(visual_sheet, min_col=2, min_row=10, max_row=min(20, start_row + len(top_hotels)))
cats = Reference(visual_sheet, min_col=1, min_row=11, max_row=min(20, start_row + len(top_hotels)))

chart1.add_data(data, titles_from_data=True)
chart1.set_categories(cats)
chart1.height = 12
chart1.width = 24

visual_sheet.add_chart(chart1, "J3")

# Create Scatter Chart: Distance vs Score
chart2 = ScatterChart()
chart2.title = "Distance vs. Compset Score Analysis"
chart2.style = 13
chart2.x_axis.title = 'Distance from Grand Millennium (km)'
chart2.y_axis.title = 'Compset Score'

# Add data for scored hotels only
xvalues = Reference(visual_sheet, min_col=3, min_row=11, max_row=min(25, start_row + len(top_hotels)))
yvalues = Reference(visual_sheet, min_col=2, min_row=11, max_row=min(25, start_row + len(top_hotels)))

series = openpyxl.chart.Series(values=yvalues, xvalues=xvalues, title="Hotels")
chart2.series.append(series)

chart2.height = 12
chart2.width = 24

visual_sheet.add_chart(chart2, "J22")

# Create summary statistics section
visual_sheet['J41'] = 'Competitive Set Recommendations'
visual_sheet['J41'].font = openpyxl.styles.Font(bold=True, size=14)

visual_sheet['J43'] = 'PRIMARY COMPSET (Score >= 90):'
visual_sheet['J43'].font = openpyxl.styles.Font(bold=True)
visual_sheet['J43'].fill = openpyxl.styles.PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

row_idx = 44
primary_hotels = [h for h in top_hotels if h[1] >= 90]
if primary_hotels:
    for name, score, _, _, _, _, _ in primary_hotels:
        visual_sheet[f'J{row_idx}'] = f"• {name} (Score: {score:.1f})"
        row_idx += 1
else:
    visual_sheet[f'J{row_idx}'] = "• None - Consider adjusting criteria"
    row_idx += 1

row_idx += 1
visual_sheet[f'J{row_idx}'] = 'SECONDARY COMPSET (Score 75-89):'
visual_sheet[f'J{row_idx}'].font = openpyxl.styles.Font(bold=True)
visual_sheet[f'J{row_idx}'].fill = openpyxl.styles.PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
row_idx += 1

secondary_hotels = [h for h in top_hotels if 75 <= h[1] < 90]
if secondary_hotels:
    for name, score, _, _, _, _, _ in secondary_hotels[:5]:  # Top 5 secondary
        visual_sheet[f'J{row_idx}'] = f"• {name} (Score: {score:.1f})"
        row_idx += 1
else:
    visual_sheet[f'J{row_idx}'] = "• None identified yet"
    row_idx += 1

# Key Insights Section
row_idx += 2
visual_sheet[f'J{row_idx}'] = 'KEY INSIGHTS:'
visual_sheet[f'J{row_idx}'].font = openpyxl.styles.Font(bold=True, size=12)
row_idx += 1

avg_score = sum([s for s in scores if s > 0]) / len([s for s in scores if s > 0]) if scores else 0
visual_sheet[f'J{row_idx}'] = f'• Average Compset Score: {avg_score:.1f}/100'
row_idx += 1

avg_distance = sum([d for d in distances if d > 0]) / len([d for d in distances if d > 0]) if distances else 0
visual_sheet[f'J{row_idx}'] = f'• Average Distance: {avg_distance:.2f} km'
row_idx += 1

hotels_within_2km = len([d for d in distances if 0 < d <= 2])
visual_sheet[f'J{row_idx}'] = f'• Hotels within 2 km (Immediate Competition): {hotels_within_2km}'
row_idx += 1

upscale_count = sum(1 for s in star_ratings if s == 4)
visual_sheet[f'J{row_idx}'] = f'• 4-Star Upscale Hotels: {upscale_count}'
row_idx += 1

midscale_count = sum(1 for s in star_ratings if s == 3)
visual_sheet[f'J{row_idx}'] = f'• 3-Star Midscale Hotels: {midscale_count}'

# Save the workbook
wb.save(excel_file)

print(f"\nCharts and analysis dashboard created successfully!")
print(f"Sheet 'Compset Scoring Analysis' added to: {excel_file}")
print(f"\nTop 10 Compset Candidates:")
for idx, (name, score, dist, _, _, _, _) in enumerate(top_hotels[:10], 1):
    print(f"{idx}. {name}")
    print(f"   Score: {score:.1f} | Distance: {dist:.2f} km")
print(f"\nAverage Compset Score: {avg_score:.1f}/100")
