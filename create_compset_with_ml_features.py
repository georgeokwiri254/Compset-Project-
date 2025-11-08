import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Create workbook
wb = openpyxl.Workbook()

# Sheet 1: Main Compset Analysis with Normalization
ws_main = wb.active
ws_main.title = "Compset Analysis"

# Sheet 2: ML Features & Scoring
ws_ml = wb.create_sheet("ML Features & Scoring")

# Define styles
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=10)
hotel_header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
hotel_header_font = Font(bold=True, color="FFFFFF", size=11)
category_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
category_font = Font(bold=True, size=10)
normalize_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
normalize_font = Font(bold=True, size=10, color="000000")
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
wrap_alignment = Alignment(wrap_text=True, vertical='top')

# Comprehensive hotel data
hotels_data = {
    "Grand Millennium Dubai": {
        "Address": "Sheikh Zayed Road, Barsha Heights (Tecom), Dubai",
        "Distance from Hotel (KM)": "0",
        "Property Type": "City",
        "Star Rating": "5",
        "Opening Year": "N/A",
        "Last Renovation": "N/A",
        "Brand": "Millennium Hotels and Resorts",
        "Owning Company/Management": "Millennium Hotels and Resorts (MHR) / City Developments Limited (CDL)",
        "Licensed for Alcohol (Y/N)": "Y",
        "Licensed for Shisha (Y/N)": "Y",
        "Total Number of Keys": "339",
        "Type of Hotel": "Mixed (Rooms, Suites, Apartments)",
        "Room Types Detail": "Superior Room (33m²), Deluxe Room (33m²), Club Room (33m²), Business Suite (60m²), Executive Suite (66m²), Studio Suite/Apartment (54m²), One Bedroom Suite/Apartment (92m²), Two Bedroom Suite/Apartment (111m²), Royal Suite (120m²), Penthouse (335m²)",
        "Number of Room Types": "10",
        "Percent Suites": "Multiple suite categories available",
        "Restaurants Count": "3",
        "Restaurants": "3 restaurants: The Atrium (international casual dining), Asian Restaurant (18th floor with live cooking stations), Belgian/European restaurant",
        "Bars Count": "2",
        "Bars": "2 bars: Crystal Bar (stylish lounge bar with 3am license), Poolside Bar, Belgian Beer Café",
        "Lounge/Coffee Shop": "Hotel lobby lounge, coffee facilities",
        "Total Number of Function Rooms": "7",
        "Total Function Space (sqm)": "N/A",
        "Number of Ballroom": "1",
        "Number of Meeting rooms": "7",
        "Largest Room/Ballroom (sqm)": "N/A",
        "Maximum capacity (theater style)": "N/A",
        "Spa Name": "Jasmine Spa",
        "Treatment rooms (#)": "9",
        "Spa total area (sqm)": "N/A",
        "Health Club / Fitness (Y/N)": "Y",
        "Health Club Details": "Y - Bodytecture Fitness Center",
        "Kids Club": "Y",
        "Kids Club Details": "Y - On-site play area (dedicated team)",
        "Kid Amenities": "Play area, babysitting services",
        "Swimming Pool": "Y",
        "Swimming Pool Count": "1",
        "Pool Details": "Y - Temperature-controlled rooftop pool",
        "Direct Beach": "N",
        "Beach Access Details": "N - Shuttle to Zero Gravity beach (Mon, Tue, Thu)",
        "Retail shops": "Y",
        "Retail Details": "Y - Gift shops/newsstands",
        "Additional Facilities": "Sauna, steam room, outdoor & indoor Jacuzzi, 24-hour room service, free parking"
    },
    "Zabeel House The Greens": {
        "Address": "The Onyx Tower 3, The Greens, Sheikh Zayed Road, Dubai",
        "Distance from Hotel (KM)": "3-5",
        "Property Type": "City",
        "Star Rating": "4",
        "Opening Year": "2018",
        "Last Renovation": "N/A",
        "Brand": "Jumeirah Group",
        "Owning Company/Management": "The Onyx for Development (Ishraqah) / Managed by Jumeirah Group",
        "Licensed for Alcohol (Y/N)": "Y",
        "Licensed for Shisha (Y/N)": "Y",
        "Total Number of Keys": "210",
        "Type of Hotel": "Mixed (Rooms, Studios, Apartments)",
        "Room Types Detail": "Popular Room (28m²), Popular Balcony Room (28m²), Studio Room (28m²), The Apartment/Suite (28m²), Connected Rooms (56m²)",
        "Number of Room Types": "5",
        "Percent Suites": "Mix of rooms and studio apartments",
        "Restaurants Count": "2",
        "Restaurants": "2 restaurants: Social Company (artisanal co-working cafe/bar, 6:30 AM-midnight), Lah Lah (Pan-Asian restaurant with rooftop shisha terrace)",
        "Bars Count": "2",
        "Bars": "Integrated within restaurants - Social Company bar, Lah Lah bar/lounge",
        "Lounge/Coffee Shop": "Social Company (co-working cafe)",
        "Total Number of Function Rooms": "4",
        "Total Function Space (sqm)": "156",
        "Number of Ballroom": "0",
        "Number of Meeting rooms": "4",
        "Largest Room/Ballroom (sqm)": "156",
        "Maximum capacity (theater style)": "200",
        "Spa Name": "Native Club by Soul Senses",
        "Treatment rooms (#)": "N/A",
        "Spa total area (sqm)": "N/A",
        "Health Club / Fitness (Y/N)": "Y",
        "Health Club Details": "Y - Fully equipped gym with indoor/outdoor functional training zone",
        "Kids Club": "Y",
        "Kids Club Details": "Y - Dedicated kids area, Play House",
        "Kid Amenities": "Kids club, play house, game room, kids pool area",
        "Swimming Pool": "Y",
        "Swimming Pool Count": "1",
        "Pool Details": "Y - Outdoor pool on 4th floor (Native Club)",
        "Direct Beach": "N",
        "Beach Access Details": "N - Access to Jumeirah Zabeel Saray beach (discounted vouchers)",
        "Retail shops": "Y",
        "Retail Details": "Y - Grocery/convenience store",
        "Additional Facilities": "2 outdoor padel courts, outdoor gym, Native Club facilities (gym 6am-9:30pm, spa 11am-9pm)"
    },
    "voco Bonnington Dubai": {
        "Address": "Cluster J, Jumeirah Lakes Towers, Dubai",
        "Distance from Hotel (KM)": "2-3",
        "Property Type": "City",
        "Star Rating": "5",
        "Opening Year": "N/A",
        "Last Renovation": "2022",
        "Brand": "voco (IHG Hotels & Resorts)",
        "Owning Company/Management": "Managed by IHG Hotels & Resorts",
        "Licensed for Alcohol (Y/N)": "Y",
        "Licensed for Shisha (Y/N)": "Y",
        "Total Number of Keys": "208",
        "Type of Hotel": "Rooms",
        "Room Types Detail": "Classic Guest Room, Executive Room (4 distinct categories total with blackout blinds, rainfall shower, bathtub)",
        "Number of Room Types": "4",
        "Percent Suites": "N/A",
        "Restaurants Count": "5",
        "Restaurants": "5 F&B outlets: The Cavendish Restaurant (British-European with Middle Eastern), McGettigan's Irish Pub (award-winning), The Cheeky Camel (restaurant/bar), The Authors' Lounge (cafe with teas/coffees/pastries), Coffee Box",
        "Bars Count": "3",
        "Bars": "McGettigan's Irish Pub, The Cheeky Camel, Pool Bar/Leisure Deck",
        "Lounge/Coffee Shop": "The Authors' Lounge, Coffee Box",
        "Total Number of Function Rooms": "6",
        "Total Function Space (sqm)": "N/A",
        "Number of Ballroom": "0",
        "Number of Meeting rooms": "6",
        "Largest Room/Ballroom (sqm)": "N/A",
        "Maximum capacity (theater style)": "120",
        "Spa Name": "Beauty salon/spa on 11th floor Leisure Deck",
        "Treatment rooms (#)": "N/A",
        "Spa total area (sqm)": "N/A",
        "Health Club / Fitness (Y/N)": "Y",
        "Health Club Details": "Y - Modern gym on 11th floor",
        "Kids Club": "N",
        "Kids Club Details": "N",
        "Kid Amenities": "N/A",
        "Swimming Pool": "Y",
        "Swimming Pool Count": "1",
        "Pool Details": "Y - 11th floor infinity pool with Dubai skyline views",
        "Direct Beach": "N",
        "Beach Access Details": "N - Complimentary shuttle to Riva Beach Club",
        "Retail shops": "N",
        "Retail Details": "N/A",
        "Additional Facilities": "Sauna, steam rooms, beauty salon with spa treatments (facials, body treatments, manicures, pedicures, massage)"
    },
    "Radisson Blu Media City": {
        "Address": "Dubai Media City, Dubai",
        "Distance from Hotel (KM)": "2-3",
        "Property Type": "City",
        "Star Rating": "4",
        "Opening Year": "N/A",
        "Last Renovation": "N/A",
        "Brand": "Radisson Hotel Group",
        "Owning Company/Management": "Radisson Hotel Group",
        "Licensed for Alcohol (Y/N)": "Y",
        "Licensed for Shisha (Y/N)": "N/A",
        "Total Number of Keys": "246",
        "Type of Hotel": "Mixed (Rooms, Suites)",
        "Room Types Detail": "Standard Room, Executive Room (with mini-bar, Executive Lounge access, complimentary breakfast), One Bedroom Suite with Lounge Access, Connected Rooms (for families)",
        "Number of Room Types": "4",
        "Percent Suites": "Limited suite inventory",
        "Restaurants Count": "2",
        "Restaurants": "2 restaurants: Chef's House (breakfast buffet), Certo Italian Restaurant, Tamanya Goes Thai",
        "Bars Count": "2",
        "Bars": "2 bars: ICON Sports Bar, additional bar",
        "Lounge/Coffee Shop": "2 coffee shops/cafes: Jones the Grocer Express",
        "Total Number of Function Rooms": "12",
        "Total Function Space (sqm)": "582",
        "Number of Ballroom": "1",
        "Number of Meeting rooms": "12",
        "Largest Room/Ballroom (sqm)": "N/A",
        "Maximum capacity (theater style)": "150",
        "Spa Name": "Dreamworks SPA / Senso Wellness Centre",
        "Treatment rooms (#)": "5",
        "Spa total area (sqm)": "N/A",
        "Health Club / Fitness (Y/N)": "Y",
        "Health Club Details": "Y - State-of-the-art gym with fitness classes",
        "Kids Club": "N",
        "Kids Club Details": "N",
        "Kid Amenities": "Children welcome, no specific kids club",
        "Swimming Pool": "Y",
        "Swimming Pool Count": "2",
        "Pool Details": "Y - 2 outdoor pools (free to use)",
        "Direct Beach": "N",
        "Beach Access Details": "N - Public beach 20km away with shuttle bus (Jumeirah Beach Park)",
        "Retail shops": "N",
        "Retail Details": "N/A",
        "Additional Facilities": "Sauna, steam bath, massage room, Senso Terrace (poolside events for up to 60 guests), 24-hour room service"
    },
    "Media One Hotel": {
        "Address": "Dubai Media City, Dubai",
        "Distance from Hotel (KM)": "2-3",
        "Property Type": "City/Lifestyle",
        "Star Rating": "4",
        "Opening Year": "N/A",
        "Last Renovation": "N/A",
        "Brand": "Independent",
        "Owning Company/Management": "Independent mixed-use property",
        "Licensed for Alcohol (Y/N)": "Y",
        "Licensed for Shisha (Y/N)": "Y",
        "Total Number of Keys": "264",
        "Type of Hotel": "Mixed (Rooms, Studios, Suites)",
        "Room Types Detail": "Hip Room/Standard (30m²), Hip Urban/City Deluxe (30m²), Business Room (30m²), Executive Room (35m²), Studio (50m²), Suite (70m²)",
        "Number of Room Types": "6",
        "Percent Suites": "Mix with studio and suite options",
        "Restaurants Count": "4",
        "Restaurants": "4 on-site restaurants serving international cuisine, all-day-dining restaurant, additional themed restaurants",
        "Bars Count": "3",
        "Bars": "2 bars + outdoor lounge: Pool bar (8th floor), Q43 (43rd floor with billiards), Friends-themed lounge, nightclub",
        "Lounge/Coffee Shop": "Coffee shop with co-working desks, lobby lounge",
        "Total Number of Function Rooms": "8",
        "Total Function Space (sqm)": "702",
        "Number of Ballroom": "0",
        "Number of Meeting rooms": "8",
        "Largest Room/Ballroom (sqm)": "N/A",
        "Maximum capacity (theater style)": "200",
        "Spa Name": "Spa services (off-site nearby)",
        "Treatment rooms (#)": "N/A",
        "Spa total area (sqm)": "N/A",
        "Health Club / Fitness (Y/N)": "Y",
        "Health Club Details": "Y - 24/7 gymnasium on 8th floor",
        "Kids Club": "N",
        "Kids Club Details": "N - Babysitting services available",
        "Kid Amenities": "Kids menu, babysitting services",
        "Swimming Pool": "Y",
        "Swimming Pool Count": "1",
        "Pool Details": "Y - Large outdoor rooftop pool on 8th floor",
        "Direct Beach": "N",
        "Beach Access Details": "N - Free shuttle to Peaches & Cream beach club",
        "Retail shops": "Y",
        "Retail Details": "Y - Number of retail outlets",
        "Additional Facilities": "Sauna, steam rooms, P7 Arena (underground entertainment space), co-working spaces, 9 F&B venues, 230,000 sq ft office space"
    },
    "Pullman JLT": {
        "Address": "Sheikh Zayed Road, Jumeirah Lakes Towers (near Sobha Realty Metro), Dubai",
        "Distance from Hotel (KM)": "2-3",
        "Property Type": "City",
        "Star Rating": "5",
        "Opening Year": "2015",
        "Last Renovation": "N/A",
        "Brand": "Pullman (Accor Hotels)",
        "Owning Company/Management": "Managed by Accor Hotels",
        "Licensed for Alcohol (Y/N)": "Y",
        "Licensed for Shisha (Y/N)": "Y",
        "Total Number of Keys": "354",
        "Type of Hotel": "Mixed (Rooms, Suites, Apartments)",
        "Room Types Detail": "Superior Room (35m²), Deluxe Room (40m²), Executive Room (35m²), Deluxe Executive Room (40m²), Studio Suite, 1-Bedroom Suite, 2-Bedroom Suite, 3-Bedroom Suite",
        "Number of Room Types": "8",
        "Percent Suites": "21% (76 out of 354)",
        "Restaurants Count": "5",
        "Restaurants": "5 restaurants/bars: Seasons Restaurant (all-day international dining), Manzoni Bistro & Bar (Italian, award-winning), La Vue (outdoor lounge), Terrace Bar & Lounge",
        "Bars Count": "3",
        "Bars": "Manzoni Bar, La Vue, Terrace Bar & Lounge, Rooftop bar",
        "Lounge/Coffee Shop": "La Vue (outdoor lounge), lounges in hotel",
        "Total Number of Function Rooms": "11",
        "Total Function Space (sqm)": "550",
        "Number of Ballroom": "0",
        "Number of Meeting rooms": "11",
        "Largest Room/Ballroom (sqm)": "N/A",
        "Maximum capacity (theater style)": "N/A",
        "Spa Name": "Fit and Spa Lounge (34th floor)",
        "Treatment rooms (#)": "5",
        "Spa total area (sqm)": "N/A",
        "Health Club / Fitness (Y/N)": "Y",
        "Health Club Details": "Y - Pullman Fit with Technogym equipment",
        "Kids Club": "Y",
        "Kids Club Details": "Y - Kids' play area",
        "Kid Amenities": "Kids' play area with activities",
        "Swimming Pool": "Y",
        "Swimming Pool Count": "1",
        "Pool Details": "Y - Rooftop pool",
        "Direct Beach": "N",
        "Beach Access Details": "N - Near Marina Beach, Dubai Marina less than 0.5 miles",
        "Retail shops": "N",
        "Retail Details": "N/A",
        "Additional Facilities": "Separate relaxation areas for men/women, steam & sauna facilities, Nespresso machines (Executive+), Bose docking stations (Executive+), Pool Hours: 8AM-7PM, Gym: 8AM-8PM, Spa: 12PM-9PM"
    },
    "Millennium Place Barsha Heights": {
        "Address": "Barsha Heights (Tecom), near Mall of the Emirates, Dubai",
        "Distance from Hotel (KM)": "0.5-1",
        "Property Type": "City",
        "Star Rating": "4",
        "Opening Year": "Recent",
        "Last Renovation": "N/A",
        "Brand": "Millennium Hotels and Resorts",
        "Owning Company/Management": "Millennium Hotels and Resorts",
        "Licensed for Alcohol (Y/N)": "N",
        "Licensed for Shisha (Y/N)": "Y",
        "Total Number of Keys": "468",
        "Type of Hotel": "Mixed (Hotel Rooms + Serviced Apartments)",
        "Room Types Detail": "Superior Room, Deluxe Room, Deluxe Sky Room, Premium Room, Premium Sky Room, Deluxe Suite, Two-Bedroom Suite, Studio Apartment, 1-Bedroom Apartment, 2-Bedroom Apartment, Royal Penthouse",
        "Number of Room Types": "11",
        "Percent Suites": "Mix of hotel rooms and apartments",
        "Restaurants Count": "5",
        "Restaurants": "5 dining options: M One Restaurant (international buffet with themed nights), M Two Restaurant (apartment side), Eva's Fusion, additional outlets",
        "Bars Count": "0",
        "Bars": "0 bars (no alcohol)",
        "Lounge/Coffee Shop": "Twenty9 Lounge (level 29 rooftop with shisha), Level Social Lobby Lounge, Splash Pool Bar & Café",
        "Total Number of Function Rooms": "7",
        "Total Function Space (sqm)": "146",
        "Number of Ballroom": "1",
        "Number of Meeting rooms": "7",
        "Largest Room/Ballroom (sqm)": "146",
        "Maximum capacity (theater style)": "170",
        "Spa Name": "Spa treatment rooms",
        "Treatment rooms (#)": "N/A",
        "Spa total area (sqm)": "N/A",
        "Health Club / Fitness (Y/N)": "Y",
        "Health Club Details": "Y - Fitness Centre on 29th floor",
        "Kids Club": "Y",
        "Kids Club Details": "Y - Kids Club with indoor play area",
        "Kid Amenities": "Kids Club, indoor play area, children's pool",
        "Swimming Pool": "Y",
        "Swimming Pool Count": "4",
        "Pool Details": "Y - 4 pools (3 outdoor + children's pool)",
        "Direct Beach": "N",
        "Beach Access Details": "N - Shuttle to Kite Beach (AED 10/person), minutes from Jumeirah Beach",
        "Retail shops": "N",
        "Retail Details": "N/A",
        "Additional Facilities": "Spa therapy, fitness classes, wellness centre, cycling facilities, 4K Laser Projector in ballroom, highspeed WiFi, USB charging, 43-inch IPTV"
    },
    "TRYP by Wyndham Dubai": {
        "Address": "Barsha Heights (Tecom), Dubai",
        "Distance from Hotel (KM)": "0.5-1",
        "Property Type": "City/Lifestyle",
        "Star Rating": "4",
        "Opening Year": "2017",
        "Last Renovation": "N/A",
        "Brand": "TRYP (Wyndham Hotels & Resorts)",
        "Owning Company/Management": "Wyndham Hotels & Resorts (many apartments privately owned)",
        "Licensed for Alcohol (Y/N)": "Y",
        "Licensed for Shisha (Y/N)": "Y",
        "Total Number of Keys": "650",
        "Type of Hotel": "Mixed (Rooms, Suites)",
        "Room Types Detail": "TRYP Room, TRYP King Room, Triple Room, TRYP One Bedroom Suite (41 sqm/441 sqft), TRYP Premium Suite (loft-type), Accessible Room (20m²)",
        "Number of Room Types": "6",
        "Percent Suites": "Limited suite inventory",
        "Restaurants Count": "3",
        "Restaurants": "3 trendy restaurants: Local Restaurant (rustic eatery/bar, all-day with live sports/shisha/live music), Lola Taberna Espanola (Spanish/paella), Haze Lounge (Greek/Levant)",
        "Bars Count": "2",
        "Bars": "Local Restaurant bar (with happy hour, Ladies Night Wed), Pool bar",
        "Lounge/Coffee Shop": "Haze Lounge, Coffee shop, club lounge",
        "Total Number of Function Rooms": "2",
        "Total Function Space (sqm)": "357",
        "Number of Ballroom": "0",
        "Number of Meeting rooms": "2",
        "Largest Room/Ballroom (sqm)": "357",
        "Maximum capacity (theater style)": "80",
        "Spa Name": "Rayya Wellness Spa",
        "Treatment rooms (#)": "N/A",
        "Spa total area (sqm)": "N/A",
        "Health Club / Fitness (Y/N)": "Y",
        "Health Club Details": "Y - Abz Gym (cardio, free weights, elliptical)",
        "Kids Club": "N",
        "Kids Club Details": "N - Babysitting/childcare available",
        "Kid Amenities": "Kids pool, babysitting services, day-care centre",
        "Swimming Pool": "Y",
        "Swimming Pool Count": "2",
        "Pool Details": "Y - Terrace pool (8AM-8PM) with sun deck & pool bar + kids pool",
        "Direct Beach": "N",
        "Beach Access Details": "N - Exclusive access to Soluna Beach Club at The Palm (fee applies, closed Jun-Aug 2025 for renovation)",
        "Retail shops": "N",
        "Retail Details": "N/A",
        "Additional Facilities": "Sauna, spa treatments (hot stone, Thai massage), tech-savvy co-working space, free shuttle to beach/mall"
    },
    "Mercure Dubai Barsha Heights": {
        "Address": "Sheikh Zayed Road, Barsha Heights (Tecom), Dubai",
        "Distance from Hotel (KM)": "0.5-1",
        "Property Type": "City",
        "Star Rating": "4",
        "Opening Year": "N/A",
        "Last Renovation": "N/A",
        "Brand": "Mercure (Accor Hotels)",
        "Owning Company/Management": "Managed by Accor Hotels",
        "Licensed for Alcohol (Y/N)": "Y",
        "Licensed for Shisha (Y/N)": "Y",
        "Total Number of Keys": "408",
        "Type of Hotel": "All Suites & Apartments",
        "Room Types Detail": "1-Bedroom Deluxe Suite (60m²), 2-Bedroom Prestige Suite (105m²) - all units are suites with living areas, kitchenettes, balconies",
        "Number of Room Types": "2",
        "Percent Suites": "100%",
        "Restaurants Count": "4",
        "Restaurants": "4 dining spots: Day & Night Restaurant (international buffet with live cooking, shisha), Eva's (mention unclear), in-room dining",
        "Bars Count": "1",
        "Bars": "The Exit Sports Bar (London-style, 8 screens with outdoor terrace)",
        "Lounge/Coffee Shop": "Café Social (lobby lounge with 50% discount after 8pm), Mixolounge/VIP Lounge (27th floor terrace), Corner 8 (pool bar on 8th floor)",
        "Total Number of Function Rooms": "Multiple",
        "Total Function Space (sqm)": "N/A",
        "Number of Ballroom": "0",
        "Number of Meeting rooms": "Multiple",
        "Largest Room/Ballroom (sqm)": "N/A",
        "Maximum capacity (theater style)": "100",
        "Spa Name": "Spa with various treatments",
        "Treatment rooms (#)": "N/A",
        "Spa total area (sqm)": "N/A",
        "Health Club / Fitness (Y/N)": "Y",
        "Health Club Details": "Y - Fitness Centre (Excellent GymFactor rating)",
        "Kids Club": "Y",
        "Kids Club Details": "Y - Kids' Club with indoor play area",
        "Kid Amenities": "Kids' Club, indoor play area, children's pool",
        "Swimming Pool": "Y",
        "Swimming Pool Count": "2",
        "Pool Details": "Y - Rooftop pool terrace on 8th floor with hot tub & children's pool",
        "Direct Beach": "N",
        "Beach Access Details": "N - Free shuttle to Jumeirah Open beach (daily at 9am/10am/11am, return 1pm/2pm/3pm)",
        "Retail shops": "N",
        "Retail Details": "N/A",
        "Additional Facilities": "Hammam, sauna, Turkish steam bath, indoor pool, squash courts, indoor football court, kickboxing, salons, pool deck level 8, 41-story tower"
    }
}

# Define normalized categories for features
normalized_categories = {
    "Property Type": {"City": "CITY", "City/Lifestyle": "CITY_LIFESTYLE"},
    "Star Rating": {"Numeric": "1-5 scale"},
    "Type of Hotel": {
        "Mixed (Rooms, Suites, Apartments)": "MIXED_RSA",
        "Mixed (Rooms, Studios, Apartments)": "MIXED_RSA",
        "Rooms": "ROOMS_ONLY",
        "Mixed (Rooms, Suites)": "MIXED_RS",
        "Mixed (Rooms, Studios, Suites)": "MIXED_RSS",
        "Mixed (Hotel Rooms + Serviced Apartments)": "MIXED_RSA",
        "All Suites & Apartments": "SUITES_ONLY"
    },
    "Distance": {"Numeric": "0-10 km scale"},
    "Keys": {"Numeric": "Room count"},
    "Licensed for Alcohol": {"Y": 1, "N": 0},
    "Licensed for Shisha": {"Y": 1, "N": 0},
    "Kids Club": {"Y": 1, "N": 0},
    "Swimming Pool": {"Y": 1, "N": 0},
    "Direct Beach": {"Y": 1, "N": 0},
    "Retail shops": {"Y": 1, "N": 0},
    "Health Club": {"Y": 1, "N": 0}
}

# Main Sheet: Compset Analysis with Normalization Column
field_labels = [
    ("GENERAL INFORMATION", "header"),
    ("Address", "data"),
    ("Distance from Hotel (KM)", "data"),
    ("Property Type", "data"),
    ("Star Rating", "data"),
    ("Opening Year", "data"),
    ("Last Renovation", "data"),
    ("Brand", "data"),
    ("Owning Company/Management", "data"),
    ("Licensed for Alcohol (Y/N)", "data"),
    ("Licensed for Shisha (Y/N)", "data"),
    ("", "blank"),
    ("ROOMS INFORMATION", "header"),
    ("Total Number of Keys", "data"),
    ("Type of Hotel", "data"),
    ("Room Types Detail", "data"),
    ("Number of Room Types", "data"),
    ("Percent Suites", "data"),
    ("", "blank"),
    ("FOOD & BEVERAGE", "header"),
    ("Restaurants Count", "data"),
    ("Restaurants", "data"),
    ("Bars Count", "data"),
    ("Bars", "data"),
    ("Lounge/Coffee Shop", "data"),
    ("", "blank"),
    ("MEETINGS & EVENTS", "header"),
    ("Total Number of Function Rooms", "data"),
    ("Total Function Space (sqm)", "data"),
    ("Number of Ballroom", "data"),
    ("Number of Meeting rooms", "data"),
    ("Largest Room/Ballroom (sqm)", "data"),
    ("Maximum capacity (theater style)", "data"),
    ("", "blank"),
    ("FITNESS & SPA", "header"),
    ("Spa Name", "data"),
    ("Treatment rooms (#)", "data"),
    ("Spa total area (sqm)", "data"),
    ("Health Club / Fitness (Y/N)", "data"),
    ("", "blank"),
    ("OTHER FACILITIES", "header"),
    ("Kids Club", "data"),
    ("Kid Amenities", "data"),
    ("Swimming Pool", "data"),
    ("Swimming Pool Count", "data"),
    ("Direct Beach", "data"),
    ("Beach Access Details", "data"),
    ("Retail shops", "data"),
    ("Additional Facilities", "data"),
]

# Column A: Field Labels
ws_main.column_dimensions['A'].width = 32
for idx, (label, label_type) in enumerate(field_labels, start=1):
    cell = ws_main.cell(row=idx, column=1)
    cell.value = label
    cell.border = border
    cell.alignment = wrap_alignment
    if label_type == "header":
        cell.fill = header_fill
        cell.font = header_font

# Column B: Normalization/Categorization
ws_main.column_dimensions['B'].width = 35
ws_main.cell(row=1, column=2).value = "NORMALIZATION/CATEGORIZATION"
ws_main.cell(row=1, column=2).fill = normalize_fill
ws_main.cell(row=1, column=2).font = normalize_font
ws_main.cell(row=1, column=2).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
ws_main.cell(row=1, column=2).border = border

normalization_mapping = {
    "Address": "TEXT - Location string",
    "Distance from Hotel (KM)": "NUMERIC - Distance scale [0-10]",
    "Property Type": "CATEGORICAL - [CITY, CITY_LIFESTYLE]",
    "Star Rating": "NUMERIC - Rating scale [1-5]",
    "Opening Year": "NUMERIC - Year",
    "Last Renovation": "NUMERIC - Year",
    "Brand": "CATEGORICAL - Brand name",
    "Owning Company/Management": "CATEGORICAL - Management company",
    "Licensed for Alcohol (Y/N)": "BINARY - [0=N, 1=Y]",
    "Licensed for Shisha (Y/N)": "BINARY - [0=N, 1=Y]",
    "Total Number of Keys": "NUMERIC - Room count [0-1000]",
    "Type of Hotel": "CATEGORICAL - [MIXED_RSA, ROOMS_ONLY, MIXED_RS, MIXED_RSS, SUITES_ONLY] **KEY MATCHING FEATURE**",
    "Room Types Detail": "TEXT - Descriptive",
    "Number of Room Types": "NUMERIC - Count [0-15]",
    "Percent Suites": "TEXT - Descriptive",
    "Restaurants Count": "NUMERIC - Count [0-10]",
    "Restaurants": "TEXT - Descriptive",
    "Bars Count": "NUMERIC - Count [0-10]",
    "Bars": "TEXT - Descriptive",
    "Lounge/Coffee Shop": "TEXT - Descriptive",
    "Total Number of Function Rooms": "NUMERIC - Count [0-20]",
    "Total Function Space (sqm)": "NUMERIC - Area [0-2000]",
    "Number of Ballroom": "NUMERIC - Count [0-5]",
    "Number of Meeting rooms": "NUMERIC - Count [0-20]",
    "Largest Room/Ballroom (sqm)": "NUMERIC - Area [0-1000]",
    "Maximum capacity (theater style)": "NUMERIC - Capacity [0-500]",
    "Spa Name": "TEXT - Spa name",
    "Treatment rooms (#)": "NUMERIC - Count [0-20]",
    "Spa total area (sqm)": "NUMERIC - Area [0-2000]",
    "Health Club / Fitness (Y/N)": "BINARY - [0=N, 1=Y]",
    "Kids Club": "BINARY - [0=N, 1=Y]",
    "Kid Amenities": "TEXT - Descriptive",
    "Swimming Pool": "BINARY - [0=N, 1=Y]",
    "Swimming Pool Count": "NUMERIC - Count [0-10]",
    "Direct Beach": "BINARY - [0=N, 1=Y]",
    "Beach Access Details": "TEXT - Descriptive",
    "Retail shops": "BINARY - [0=N, 1=Y]",
    "Additional Facilities": "TEXT - Descriptive"
}

row_idx = 2
for label, label_type in field_labels[1:]:
    cell = ws_main.cell(row=row_idx, column=2)
    cell.border = border
    cell.alignment = wrap_alignment
    if label in normalization_mapping:
        cell.value = normalization_mapping[label]
        cell.fill = normalize_fill
    row_idx += 1

# Columns C onwards: Hotel Data
col_offset = 3
for hotel_name, hotel_data in hotels_data.items():
    ws_main.column_dimensions[get_column_letter(col_offset)].width = 38

    # Hotel name header
    cell = ws_main.cell(row=1, column=col_offset)
    cell.value = hotel_name
    cell.fill = hotel_header_fill
    cell.font = hotel_header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border

    # Data rows
    row_idx = 2
    for label, label_type in field_labels[1:]:
        cell = ws_main.cell(row=row_idx, column=col_offset)
        cell.border = border
        cell.alignment = wrap_alignment

        if label_type == "data" and label in hotel_data:
            cell.value = hotel_data[label]
        elif label_type == "header":
            cell.fill = category_fill

        row_idx += 1

    col_offset += 1

# Freeze panes
ws_main.freeze_panes = 'C2'

# ===== ML FEATURES & SCORING SHEET =====
ws_ml.column_dimensions['A'].width = 40

# Headers for ML sheet
ml_headers = [
    "Feature Name",
    "Grand Millennium Dubai",
    "Zabeel House The Greens",
    "voco Bonnington Dubai",
    "Radisson Blu Media City",
    "Media One Hotel",
    "Pullman JLT",
    "Millennium Place Barsha Heights",
    "TRYP by Wyndham Dubai",
    "Mercure Dubai Barsha Heights"
]

for col_idx, header in enumerate(ml_headers, start=1):
    cell = ws_ml.cell(row=1, column=col_idx)
    cell.value = header
    cell.fill = hotel_header_fill
    cell.font = hotel_header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border
    ws_ml.column_dimensions[get_column_letter(col_idx)].width = 22

# Define scoring features
scoring_features = [
    ("SECTION: PROPERTY BASICS", "header", {}),
    ("Star Rating (0-10 points)", "score", {"5": 10, "4": 7}),
    ("Distance from Hotel (10 points max, closer = higher)", "score", {}),
    ("Property Type Match", "score", {"City": 10, "City/Lifestyle": 8}),
    ("", "blank", {}),
    ("SECTION: ROOM MIX (CRITICAL)", "header", {}),
    ("Type of Hotel - MIXED RSA Match (20 points)", "score", {
        "Mixed (Rooms, Suites, Apartments)": 20,
        "Mixed (Rooms, Studios, Apartments)": 20,
        "Mixed (Hotel Rooms + Serviced Apartments)": 20,
        "Mixed (Rooms, Suites)": 12,
        "Mixed (Rooms, Studios, Suites)": 15,
        "Rooms": 0,
        "All Suites & Apartments": 5
    }),
    ("Number of Room Types (0-10 points)", "score", {}),
    ("Total Room Count Match (0-10 points)", "score", {}),
    ("", "blank", {}),
    ("SECTION: F&B FACILITIES", "header", {}),
    ("Number of Restaurants (0-5 points)", "score", {}),
    ("Number of Bars (0-5 points)", "score", {}),
    ("Licensed for Alcohol (5 points)", "score", {"Y": 5, "N": 0}),
    ("Licensed for Shisha (3 points)", "score", {"Y": 3, "N": 0}),
    ("", "blank", {}),
    ("SECTION: MEETINGS & EVENTS", "header", {}),
    ("Has Ballroom (10 points)", "score", {"1": 10, "0": 0}),
    ("Number of Meeting Rooms (0-5 points)", "score", {}),
    ("", "blank", {}),
    ("SECTION: WELLNESS & RECREATION", "header", {}),
    ("Health Club/Fitness (5 points)", "score", {"Y": 5, "N": 0}),
    ("Spa Treatment Rooms (0-5 points)", "score", {}),
    ("Kids Club (5 points)", "score", {"Y": 5, "N": 0}),
    ("Swimming Pool (3 points)", "score", {"Y": 3, "N": 0}),
    ("Retail Shops (2 points)", "score", {"Y": 2, "N": 0}),
    ("", "blank", {}),
    ("TOTAL COMPSET MATCH SCORE (out of 100)", "total", {}),
    ("MATCH PERCENTAGE", "percentage", {}),
    ("RANK (1=Best Match)", "rank", {})
]

# ONE-HOT ENCODING SECTION
one_hot_headers = [
    ("", "blank", {}),
    ("ONE-HOT ENCODING FEATURES", "header", {}),
    ("Type: MIXED_RSA (Rooms+Suites+Apartments)", "onehot", {}),
    ("Type: ROOMS_ONLY", "onehot", {}),
    ("Type: MIXED_RS (Rooms+Suites)", "onehot", {}),
    ("Type: MIXED_RSS (Rooms+Studios+Suites)", "onehot", {}),
    ("Type: SUITES_ONLY", "onehot", {}),
    ("Star: 5-Star", "onehot", {}),
    ("Star: 4-Star", "onehot", {}),
    ("Property: City", "onehot", {}),
    ("Property: City/Lifestyle", "onehot", {}),
    ("Has_Alcohol_License", "onehot", {}),
    ("Has_Shisha_License", "onehot", {}),
    ("Has_Ballroom", "onehot", {}),
    ("Has_Kids_Club", "onehot", {}),
    ("Has_Spa", "onehot", {}),
    ("Has_Retail", "onehot", {}),
]

# Combine all features
all_features = scoring_features + one_hot_headers

# Write features and calculate scores
row_idx = 2
hotel_names = list(hotels_data.keys())

for feature_name, feature_type, scoring_map in all_features:
    cell = ws_ml.cell(row=row_idx, column=1)
    cell.value = feature_name
    cell.border = border
    cell.alignment = wrap_alignment

    if feature_type == "header":
        cell.fill = header_fill
        cell.font = header_font
    elif feature_type == "blank":
        pass
    elif feature_type in ["score", "onehot"]:
        cell.fill = category_fill
        cell.font = category_font

        # Calculate values for each hotel
        for col_idx, hotel_name in enumerate(hotel_names, start=2):
            cell = ws_ml.cell(row=row_idx, column=col_idx)
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')

            hotel = hotels_data[hotel_name]

            # Scoring logic
            if feature_type == "score":
                if "Star Rating" in feature_name:
                    star = hotel.get("Star Rating", "N/A")
                    cell.value = scoring_map.get(star, 0)

                elif "Distance from Hotel" in feature_name:
                    dist_str = hotel.get("Distance from Hotel (KM)", "0")
                    try:
                        if "-" in dist_str:
                            dist = float(dist_str.split("-")[0])
                        else:
                            dist = float(dist_str)
                        cell.value = max(0, 10 - dist)
                    except:
                        cell.value = 0

                elif "Property Type Match" in feature_name:
                    prop_type = hotel.get("Property Type", "")
                    cell.value = scoring_map.get(prop_type, 0)

                elif "Type of Hotel - MIXED RSA" in feature_name:
                    hotel_type = hotel.get("Type of Hotel", "")
                    cell.value = scoring_map.get(hotel_type, 0)

                elif "Number of Room Types" in feature_name:
                    try:
                        num_types = int(hotel.get("Number of Room Types", 0))
                        cell.value = min(10, num_types)
                    except:
                        cell.value = 0

                elif "Total Room Count Match" in feature_name:
                    try:
                        keys = int(hotel.get("Total Number of Keys", 0))
                        # Grand Millennium has 339 rooms, score based on proximity
                        diff = abs(keys - 339)
                        if diff <= 50:
                            cell.value = 10
                        elif diff <= 100:
                            cell.value = 7
                        elif diff <= 200:
                            cell.value = 5
                        elif diff <= 400:
                            cell.value = 3
                        else:
                            cell.value = 0
                    except:
                        cell.value = 0

                elif "Number of Restaurants" in feature_name:
                    try:
                        count = int(hotel.get("Restaurants Count", 0))
                        cell.value = min(5, count)
                    except:
                        cell.value = 0

                elif "Number of Bars" in feature_name:
                    try:
                        count = int(hotel.get("Bars Count", 0))
                        cell.value = min(5, count * 1.5)
                    except:
                        cell.value = 0

                elif "Licensed for Alcohol" in feature_name:
                    val = hotel.get("Licensed for Alcohol (Y/N)", "N")
                    cell.value = scoring_map.get(val, 0)

                elif "Licensed for Shisha" in feature_name:
                    val = hotel.get("Licensed for Shisha (Y/N)", "N")
                    cell.value = scoring_map.get(val, 0)

                elif "Has Ballroom" in feature_name:
                    val = hotel.get("Number of Ballroom", "0")
                    cell.value = scoring_map.get(str(val) if int(val) > 0 else "0", 0)

                elif "Number of Meeting Rooms" in feature_name:
                    try:
                        count = int(hotel.get("Number of Meeting rooms", 0))
                        cell.value = min(5, count * 0.5)
                    except:
                        cell.value = 0

                elif "Health Club/Fitness" in feature_name:
                    val = hotel.get("Health Club / Fitness (Y/N)", "N")
                    cell.value = scoring_map.get(val, 0)

                elif "Spa Treatment Rooms" in feature_name:
                    try:
                        val = hotel.get("Treatment rooms (#)", "N/A")
                        if val != "N/A" and val:
                            count = int(val)
                            cell.value = min(5, count * 0.5)
                        else:
                            cell.value = 0
                    except:
                        cell.value = 0

                elif "Kids Club" in feature_name:
                    val = hotel.get("Kids Club", "N")
                    cell.value = scoring_map.get(val, 0)

                elif "Swimming Pool" in feature_name and "Score" not in feature_name:
                    val = hotel.get("Swimming Pool", "N")
                    cell.value = scoring_map.get(val, 0)

                elif "Retail Shops" in feature_name:
                    val = hotel.get("Retail shops", "N")
                    cell.value = scoring_map.get(val, 0)

            # One-hot encoding
            elif feature_type == "onehot":
                if "Type: MIXED_RSA" in feature_name:
                    hotel_type = hotel.get("Type of Hotel", "")
                    cell.value = 1 if "Mixed" in hotel_type and ("Apartments" in hotel_type or "Serviced Apartments" in hotel_type) else 0

                elif "Type: ROOMS_ONLY" in feature_name:
                    hotel_type = hotel.get("Type of Hotel", "")
                    cell.value = 1 if hotel_type == "Rooms" else 0

                elif "Type: MIXED_RS" in feature_name:
                    hotel_type = hotel.get("Type of Hotel", "")
                    cell.value = 1 if hotel_type == "Mixed (Rooms, Suites)" else 0

                elif "Type: MIXED_RSS" in feature_name:
                    hotel_type = hotel.get("Type of Hotel", "")
                    cell.value = 1 if "Studios" in hotel_type and "Suites" in hotel_type else 0

                elif "Type: SUITES_ONLY" in feature_name:
                    hotel_type = hotel.get("Type of Hotel", "")
                    cell.value = 1 if "All Suites" in hotel_type else 0

                elif "Star: 5-Star" in feature_name:
                    cell.value = 1 if hotel.get("Star Rating") == "5" else 0

                elif "Star: 4-Star" in feature_name:
                    cell.value = 1 if hotel.get("Star Rating") == "4" else 0

                elif "Property: City" in feature_name and "Lifestyle" not in feature_name:
                    cell.value = 1 if hotel.get("Property Type") == "City" else 0

                elif "Property: City/Lifestyle" in feature_name:
                    cell.value = 1 if "Lifestyle" in hotel.get("Property Type", "") else 0

                elif "Has_Alcohol_License" in feature_name:
                    cell.value = 1 if hotel.get("Licensed for Alcohol (Y/N)") == "Y" else 0

                elif "Has_Shisha_License" in feature_name:
                    cell.value = 1 if hotel.get("Licensed for Shisha (Y/N)") == "Y" else 0

                elif "Has_Ballroom" in feature_name:
                    try:
                        cell.value = 1 if int(hotel.get("Number of Ballroom", 0)) > 0 else 0
                    except:
                        cell.value = 0

                elif "Has_Kids_Club" in feature_name:
                    cell.value = 1 if hotel.get("Kids Club") == "Y" else 0

                elif "Has_Spa" in feature_name:
                    cell.value = 1 if hotel.get("Spa Name", "N/A") != "N/A" and hotel.get("Spa Name") else 0

                elif "Has_Retail" in feature_name:
                    cell.value = 1 if hotel.get("Retail shops") == "Y" else 0

    elif feature_type == "total":
        cell.fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
        cell.font = Font(bold=True, size=11)
        # Calculate totals (sum of scores above)
        for col_idx in range(2, len(hotel_names) + 2):
            cell = ws_ml.cell(row=row_idx, column=col_idx)
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
            cell.font = Font(bold=True, size=11)
            # Sum all score cells above (excluding headers and blanks)
            # We'll calculate this manually
            # For now, placeholder
            cell.value = "=SUM({}2:{}{})".format(get_column_letter(col_idx), get_column_letter(col_idx), row_idx - 1)

    elif feature_type == "percentage":
        cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        cell.font = Font(bold=True, size=11)
        for col_idx in range(2, len(hotel_names) + 2):
            cell = ws_ml.cell(row=row_idx, column=col_idx)
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
            cell.font = Font(bold=True, size=11)
            cell.value = f"={get_column_letter(col_idx)}{row_idx - 1}/100"
            cell.number_format = '0.0%'

    elif feature_type == "rank":
        cell.fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
        cell.font = Font(bold=True, size=12)
        for col_idx in range(2, len(hotel_names) + 2):
            cell = ws_ml.cell(row=row_idx, column=col_idx)
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
            cell.font = Font(bold=True, size=12)
            # Rank formula
            total_row = row_idx - 2
            cell.value = f"=RANK({get_column_letter(col_idx)}{total_row},$B${total_row}:$J${total_row},0)"

    row_idx += 1

# Freeze panes on ML sheet
ws_ml.freeze_panes = 'B2'

# Save workbook
output_file = r"C:\Users\reservations\Desktop\Compset Tool\Compset_Analysis_With_ML_Features.xlsx"
wb.save(output_file)
print(f"Comprehensive compset analysis with ML features created: {output_file}")
print("\n=== ML MODEL RECOMMENDATIONS ===")
print("For compset matching and similarity analysis, I recommend:")
print("\n1. K-Nearest Neighbors (KNN) - BEST FOR COMPSET MATCHING")
print("   - Ideal for finding similar hotels based on feature proximity")
print("   - Easy to interpret: identifies N most similar properties")
print("   - Works well with mixed numeric and categorical data")
print("   - Can use different distance metrics (Euclidean, Manhattan, etc.)")
print("\n2. Random Forest Classifier")
print("   - If you want to classify hotels into compset tiers")
print("   - Handles mixed data types well")
print("   - Provides feature importance rankings")
print("\n3. Hierarchical Clustering")
print("   - Creates dendrograms showing hotel similarity groupings")
print("   - Good for visualizing compset relationships")
print("   - No need to specify number of clusters upfront")
print("\n4. Cosine Similarity (Simple & Effective)")
print("   - Calculate similarity scores between hotels")
print("   - Fast and interpretable")
print("   - Works great with one-hot encoded features")
print("\nRECOMMENDATION: Start with KNN (k=3 to 5) using the normalized features")
print("The scoring system in the Excel already provides a weighted similarity measure!")
