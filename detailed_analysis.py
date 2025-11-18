#!/usr/bin/env python3
"""
Detailed analysis of each sheet to identify hotels and missing data
"""

import pandas as pd
from openpyxl import load_workbook

excel_path = '/home/gee_devops254/Downloads/Compset Tool/NEW - STR Competitor Set Analysis.xlsx'
wb = load_workbook(excel_path, data_only=False)

# Analyze Primary Compset comparison sheet
print("="*80)
print("PRIMARY COMPSET COMPARISON SHEET")
print("="*80)
ws = wb['Primary Compset comparison']

# Find where the data table starts by looking for column headers
print("\nSearching for data table...")
for i in range(1, min(40, ws.max_row)):
    cell_value = ws.cell(row=i, column=1).value
    if cell_value and ('competitor' in str(cell_value).lower() or 'hotel name' in str(cell_value).lower()):
        print(f"\nFound header at row {i}")
        # Print the header row
        header_row = []
        for j in range(1, min(15, ws.max_column + 1)):
            header_row.append(ws.cell(row=i, column=j).value)
        print(f"Headers: {header_row}")

        # Print next 20 rows
        print(f"\nData rows (first 20):")
        for k in range(i+1, min(i+21, ws.max_row + 1)):
            row_data = []
            for j in range(1, min(15, ws.max_column + 1)):
                row_data.append(ws.cell(row=k, column=j).value)
            print(f"Row {k}: {row_data}")
        break

# Analyze Business Mix sheet
print("\n" + "="*80)
print("BUSINESS MIX & OVERLAP ANALYSIS SHEET")
print("="*80)
ws = wb['Business Mix & Overalp Analysis']

for i in range(1, min(40, ws.max_row)):
    cell_value = ws.cell(row=i, column=1).value
    if cell_value and ('competitor' in str(cell_value).lower() or 'hotel name' in str(cell_value).lower()):
        print(f"\nFound header at row {i}")
        header_row = []
        for j in range(1, min(15, ws.max_column + 1)):
            header_row.append(ws.cell(row=i, column=j).value)
        print(f"Headers: {header_row}")

        # Print next 15 rows
        print(f"\nData rows (first 15):")
        for k in range(i+1, min(i+16, ws.max_row + 1)):
            row_data = []
            for j in range(1, min(9, ws.max_column + 1)):
                row_data.append(ws.cell(row=k, column=j).value)
            print(f"Row {k}: {row_data}")
        break

# Analyze Value Proposition sheet
print("\n" + "="*80)
print("VALUE PROPOSITION SHEET")
print("="*80)
ws = wb['Value Proposition']

for i in range(1, min(40, ws.max_row)):
    cell_value = ws.cell(row=i, column=1).value
    if cell_value and ('competitor' in str(cell_value).lower() or 'hotel name' in str(cell_value).lower()):
        print(f"\nFound header at row {i}")
        header_row = []
        for j in range(1, min(10, ws.max_column + 1)):
            header_row.append(ws.cell(row=i, column=j).value)
        print(f"Headers: {header_row}")

        # Print next 10 rows
        print(f"\nData rows (first 10):")
        for k in range(i+1, min(i+11, ws.max_row + 1)):
            row_data = []
            for j in range(1, min(10, ws.max_column + 1)):
                val = ws.cell(row=k, column=j).value
                row_data.append(val)
            print(f"Row {k}: {row_data}")
        break

print("\n" + "="*80)
