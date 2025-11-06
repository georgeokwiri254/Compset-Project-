import json
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl import load_workbook

# Additional hotel research data for remaining 15 hotels
additional_hotel_data = {
    "Millennium Place Barsha Heights Hotel Apartments": {
        "total_rooms": 447,
        "has_apartments": "Yes",
        "apartment_count": 447,
        "executive_lounge": "Yes",
        "meeting_space_sqm": 200,
        "ballroom_capacity": 170,
        "meeting_rooms_count": 4,
        "pool": "Yes (2 pools)",
        "spa": "Yes",
        "gym": "Yes",
        "restaurants_count": 2,
        "unique_features": "Full kitchenettes, Children's pool, Fully furnished apartments, Same building as hotel",
        "tripadvisor_rating": 4.0,
        "tripadvisor_reviews": 3500,
        "booking_rating": 8.4,
        "booking_reviews": 4200,
        "business_mix": 50,
        "leisure_mix": 30,
        "mice_mix": 20,
        "brand_affiliation": "Millennium Hotels & Resorts",
        "loyalty_program": "Yes - My Millennium",
        "technology_level": "Modern"
    },
    "Signature Hotel Al Barsha": {
        "total_rooms": 90,
        "has_apartments": "No",
        "apartment_count": 0,
        "executive_lounge": "No",
        "meeting_space_sqm": 100,
        "ballroom_capacity": 50,
        "meeting_rooms_count": 2,
        "pool": "Yes",
        "spa": "Yes",
        "gym": "Yes",
        "restaurants_count": 3,
        "unique_features": "Sauna, Steam room, Hot tub, Near Mall of Emirates",
        "tripadvisor_rating": 4.0,
        "tripadvisor_reviews": 800,
        "booking_rating": 7.9,
        "booking_reviews": 2500,
        "business_mix": 55,
        "leisure_mix": 40,
        "mice_mix": 5,
        "brand_affiliation": "Signature Hotels Group",
        "loyalty_program": "No",
        "technology_level": "Modern"
    },
    "Holiday Inn Express Dubai Internet City": {
        "total_rooms": 244,
        "has_apartments": "No",
        "apartment_count": 0,
        "executive_lounge": "No",
        "meeting_space_sqm": 150,
        "ballroom_capacity": 110,
        "meeting_rooms_count": 3,
        "pool": "No",
        "spa": "No",
        "gym": "Yes",
        "restaurants_count": 1,
        "unique_features": "5 floors, Free breakfast, Goose & Gander pub, Shuttle to Mall of Emirates",
        "tripadvisor_rating": 4.0,
        "tripadvisor_reviews": 1200,
        "booking_rating": 8.2,
        "booking_reviews": 3500,
        "business_mix": 65,
        "leisure_mix": 30,
        "mice_mix": 5,
        "brand_affiliation": "IHG Hotels & Resorts",
        "loyalty_program": "Yes - IHG One Rewards",
        "technology_level": "Modern"
    },
    "Class Hotel Apartments": {
        "total_rooms": 132,
        "has_apartments": "Yes",
        "apartment_count": 132,
        "executive_lounge": "No",
        "meeting_space_sqm": 60,
        "ballroom_capacity": 30,
        "meeting_rooms_count": 1,
        "pool": "Yes",
        "spa": "No",
        "gym": "Yes",
        "restaurants_count": 1,
        "unique_features": "15 floors, Rooftop pool, Full kitchens, Near Metro, Opened 2018",
        "tripadvisor_rating": 4.0,
        "tripadvisor_reviews": 600,
        "booking_rating": 8.1,
        "booking_reviews": 2000,
        "business_mix": 50,
        "leisure_mix": 45,
        "mice_mix": 5,
        "brand_affiliation": "Independent",
        "loyalty_program": "No",
        "technology_level": "Modern"
    },
    "Arjaan by Rotana - Dubai Media City": {
        "total_rooms": 242,
        "has_apartments": "Yes",
        "apartment_count": 242,
        "executive_lounge": "No",
        "meeting_space_sqm": 100,
        "ballroom_capacity": 50,
        "meeting_rooms_count": 2,
        "pool": "Yes (2 pools)",
        "spa": "No",
        "gym": "Yes",
        "restaurants_count": 2,
        "unique_features": "Children's pool, Full kitchens, Beach access at Palm, Cilantro Restaurant",
        "tripadvisor_rating": 4.0,
        "tripadvisor_reviews": 1500,
        "booking_rating": 8.1,
        "booking_reviews": 4000,
        "business_mix": 50,
        "leisure_mix": 40,
        "mice_mix": 10,
        "brand_affiliation": "Rotana Hotels",
        "loyalty_program": "Yes - Rotana Rewards",
        "technology_level": "Modern"
    },
    "Avani Plus Palm View Dubai Hotel & Suites": {
        "total_rooms": 247,
        "has_apartments": "Yes",
        "apartment_count": 150,
        "executive_lounge": "No",
        "meeting_space_sqm": 120,
        "ballroom_capacity": 60,
        "meeting_rooms_count": 2,
        "pool": "Yes (2 pools)",
        "spa": "Yes",
        "gym": "Yes",
        "restaurants_count": 2,
        "unique_features": "Children's pool, Full kitchens, Arabian Gulf views, Free bicycles",
        "tripadvisor_rating": 4.0,
        "tripadvisor_reviews": 1000,
        "booking_rating": 8.2,
        "booking_reviews": 3000,
        "business_mix": 45,
        "leisure_mix": 50,
        "mice_mix": 5,
        "brand_affiliation": "Minor Hotels",
        "loyalty_program": "Yes - DISCOVERY Loyalty",
        "technology_level": "Modern"
    },
    "Armada Avenue Hotel": {
        "total_rooms": 350,
        "has_apartments": "No",
        "apartment_count": 0,
        "executive_lounge": "No",
        "meeting_space_sqm": 150,
        "ballroom_capacity": 80,
        "meeting_rooms_count": 3,
        "pool": "Yes",
        "spa": "Yes",
        "gym": "Yes",
        "restaurants_count": 4,
        "unique_features": "6 floors, LEVEL1 Restaurant, Mythos Greek food, Pitchers Sports Bar",
        "tripadvisor_rating": 4.0,
        "tripadvisor_reviews": 1200,
        "booking_rating": 8.0,
        "booking_reviews": 3500,
        "business_mix": 55,
        "leisure_mix": 40,
        "mice_mix": 5,
        "brand_affiliation": "Independent",
        "loyalty_program": "No",
        "technology_level": "Modern"
    },
    "Oaks Liwa Heights": {
        "total_rooms": 168,
        "has_apartments": "Yes",
        "apartment_count": 168,
        "executive_lounge": "No",
        "meeting_space_sqm": 80,
        "ballroom_capacity": 40,
        "meeting_rooms_count": 2,
        "pool": "Yes",
        "spa": "No",
        "gym": "Yes",
        "restaurants_count": 1,
        "unique_features": "41 floors, 1 & 2-BR apartments, JLT location, Mixed-use building",
        "tripadvisor_rating": 3.5,
        "tripadvisor_reviews": 700,
        "booking_rating": 7.8,
        "booking_reviews": 2000,
        "business_mix": 45,
        "leisure_mix": 50,
        "mice_mix": 5,
        "brand_affiliation": "Oaks Hotels & Resorts",
        "loyalty_program": "Yes - Oaks Plus",
        "technology_level": "Standard"
    },
    "Marina Byblos Hotel": {
        "total_rooms": 184,
        "has_apartments": "No",
        "apartment_count": 0,
        "executive_lounge": "No",
        "meeting_space_sqm": 150,
        "ballroom_capacity": 80,
        "meeting_rooms_count": 3,
        "pool": "Yes",
        "spa": "Yes",
        "gym": "Yes",
        "restaurants_count": 6,
        "unique_features": "11 floors, Rooftop pool, Nightclub, 5 bars, Near JBR beach",
        "tripadvisor_rating": 4.0,
        "tripadvisor_reviews": 1000,
        "booking_rating": 7.9,
        "booking_reviews": 3000,
        "business_mix": 40,
        "leisure_mix": 55,
        "mice_mix": 5,
        "brand_affiliation": "Byblos Hospitality Group",
        "loyalty_program": "No",
        "technology_level": "Standard"
    },
    "The First Collection Dubai Marina": {
        "total_rooms": 493,
        "has_apartments": "No",
        "apartment_count": 0,
        "executive_lounge": "Yes",
        "meeting_space_sqm": 200,
        "ballroom_capacity": 100,
        "meeting_rooms_count": 4,
        "pool": "Yes",
        "spa": "Yes",
        "gym": "Yes",
        "restaurants_count": 4,
        "unique_features": "Blacksmith restaurant, Alloro Italian, Beach club access, Shuttle service",
        "tripadvisor_rating": 4.0,
        "tripadvisor_reviews": 1800,
        "booking_rating": 8.2,
        "booking_reviews": 5000,
        "business_mix": 50,
        "leisure_mix": 45,
        "mice_mix": 5,
        "brand_affiliation": "Marriott Tribute Portfolio",
        "loyalty_program": "Yes - Marriott Bonvoy",
        "technology_level": "Modern"
    },
    "Dubai Marriott Harbour Hotel & Suites": {
        "total_rooms": 261,
        "has_apartments": "Yes",
        "apartment_count": 261,
        "executive_lounge": "Yes",
        "meeting_space_sqm": 300,
        "ballroom_capacity": 150,
        "meeting_rooms_count": 5,
        "pool": "Yes",
        "spa": "Yes",
        "gym": "Yes",
        "restaurants_count": 3,
        "unique_features": "59 floors, Full kitchens, Marina & Palm views, Penthouses available",
        "tripadvisor_rating": 4.5,
        "tripadvisor_reviews": 2000,
        "booking_rating": 8.5,
        "booking_reviews": 6000,
        "business_mix": 55,
        "leisure_mix": 40,
        "mice_mix": 5,
        "brand_affiliation": "Marriott International",
        "loyalty_program": "Yes - Marriott Bonvoy",
        "technology_level": "High"
    },
    "Vida Dubai Marina & Yacht Club": {
        "total_rooms": 158,
        "has_apartments": "No",
        "apartment_count": 0,
        "executive_lounge": "No",
        "meeting_space_sqm": 100,
        "ballroom_capacity": 50,
        "meeting_rooms_count": 2,
        "pool": "Yes",
        "spa": "No",
        "gym": "Yes",
        "restaurants_count": 3,
        "unique_features": "57 floors, Infinity pool, Co-working space, Vida bikes, Pet-friendly",
        "tripadvisor_rating": 4.0,
        "tripadvisor_reviews": 1200,
        "booking_rating": 8.3,
        "booking_reviews": 4000,
        "business_mix": 50,
        "leisure_mix": 45,
        "mice_mix": 5,
        "brand_affiliation": "Emaar Hospitality Group",
        "loyalty_program": "No",
        "technology_level": "High - Co-working integration"
    },
    "Novotel Suites Mall Avenue Dubai": {
        "total_rooms": 180,
        "has_apartments": "No",
        "apartment_count": 0,
        "executive_lounge": "No",
        "meeting_space_sqm": 80,
        "ballroom_capacity": 40,
        "meeting_rooms_count": 2,
        "pool": "Yes",
        "spa": "No",
        "gym": "Yes",
        "restaurants_count": 1,
        "unique_features": "6 floors, Kids area, Library, Near Mall of Emirates, Opened 2009",
        "tripadvisor_rating": 4.0,
        "tripadvisor_reviews": 1000,
        "booking_rating": 8.0,
        "booking_reviews": 3000,
        "business_mix": 50,
        "leisure_mix": 45,
        "mice_mix": 5,
        "brand_affiliation": "Accor Hotels",
        "loyalty_program": "Yes - ALL Accor",
        "technology_level": "Standard"
    },
    "Grand Cosmopolitan Hotel": {
        "total_rooms": 228,
        "has_apartments": "No",
        "apartment_count": 0,
        "executive_lounge": "No",
        "meeting_space_sqm": 300,
        "ballroom_capacity": 200,
        "meeting_rooms_count": 4,
        "pool": "Yes (2 pools)",
        "spa": "Yes",
        "gym": "Yes",
        "restaurants_count": 3,
        "unique_features": "45 suites, Rooftop infinity pool, Children's pool, Near Metro, Starbucks",
        "tripadvisor_rating": 4.0,
        "tripadvisor_reviews": 1200,
        "booking_rating": 8.1,
        "booking_reviews": 3500,
        "business_mix": 55,
        "leisure_mix": 40,
        "mice_mix": 5,
        "brand_affiliation": "Independent",
        "loyalty_program": "No",
        "technology_level": "Modern"
    },
    "Flora Al Barsha Hotel at the Mall": {
        "total_rooms": 186,
        "has_apartments": "No",
        "apartment_count": 0,
        "executive_lounge": "No",
        "meeting_space_sqm": 100,
        "ballroom_capacity": 50,
        "meeting_rooms_count": 2,
        "pool": "Yes",
        "spa": "Yes",
        "gym": "Yes",
        "restaurants_count": 2,
        "unique_features": "Near Mall of Emirates, Connecting rooms, Modern interiors",
        "tripadvisor_rating": 4.0,
        "tripadvisor_reviews": 900,
        "booking_rating": 8.0,
        "booking_reviews": 2800,
        "business_mix": 50,
        "leisure_mix": 45,
        "mice_mix": 5,
        "brand_affiliation": "Flora Hospitality",
        "loyalty_program": "No",
        "technology_level": "Standard"
    }
}

print(f"Updating Excel file with research data for {len(additional_hotel_data)} more hotels...")
print("="*80)

# Load the Excel file
excel_file = r'C:\Users\reservations\Desktop\Compset Tool\Grand_Millennium_Dubai_CompSet_Analysis_Updated.xlsx'
output_file = r'C:\Users\reservations\Desktop\Compset Tool\Grand_Millennium_Dubai_CompSet_Analysis_Final.xlsx'
wb = load_workbook(excel_file)
ws = wb['Hotels']

# Update hotels with researched data
updated_count = 0
for row in range(2, ws.max_row + 1):
    hotel_name = ws.cell(row=row, column=2).value

    if hotel_name in additional_hotel_data:
        data = additional_hotel_data[hotel_name]
        updated_count += 1

        print(f"\n{updated_count}. Updating: {hotel_name}")
        print(f"   Rooms: {data['total_rooms']} | Apartments: {data['apartment_count']}")
        print(f"   Meeting Space: {data['meeting_space_sqm']} sqm | Rating: {data['booking_rating']}/10")

        # Populate all columns
        ws.cell(row=row, column=7, value=data["total_rooms"])
        ws.cell(row=row, column=8, value="Various room types & suites")
        ws.cell(row=row, column=9, value=data["has_apartments"])
        ws.cell(row=row, column=10, value=data["apartment_count"])
        ws.cell(row=row, column=11, value=data["executive_lounge"])
        ws.cell(row=row, column=12, value=data["meeting_space_sqm"])
        ws.cell(row=row, column=13, value=data["ballroom_capacity"])
        ws.cell(row=row, column=14, value=data["meeting_rooms_count"])
        ws.cell(row=row, column=15, value=data["pool"])
        ws.cell(row=row, column=16, value=data["spa"])
        ws.cell(row=row, column=17, value=data["gym"])
        ws.cell(row=row, column=18, value=data["restaurants_count"])
        ws.cell(row=row, column=19, value=data["unique_features"])
        ws.cell(row=row, column=20, value=data["tripadvisor_rating"])
        ws.cell(row=row, column=21, value=data["tripadvisor_reviews"])
        ws.cell(row=row, column=22, value=data["booking_rating"])
        ws.cell(row=row, column=23, value=data["booking_reviews"])
        ws.cell(row=row, column=26, value=data["business_mix"])
        ws.cell(row=row, column=27, value=data["leisure_mix"])
        ws.cell(row=row, column=28, value=data["mice_mix"])
        ws.cell(row=row, column=30, value=data["brand_affiliation"])
        ws.cell(row=row, column=31, value=data["loyalty_program"])
        ws.cell(row=row, column=33, value=data["technology_level"])

# Save the workbook
wb.save(output_file)

print("\n" + "="*80)
print(f"SUCCESS! Updated {updated_count} additional hotels with complete research data")
print(f"Excel file saved: {output_file}")
print(f"\nTotal hotels researched: 20 (previous batch) + {updated_count} (this batch) = {20 + updated_count}")
print("="*80)

# Save the additional hotel database to JSON for reference
with open(r'C:\Users\reservations\Desktop\Compset Tool\additional_hotel_research.json', 'w', encoding='utf-8') as f:
    json.dump(additional_hotel_data, f, indent=2, ensure_ascii=False)

print("\nAdditional hotel research database saved")
