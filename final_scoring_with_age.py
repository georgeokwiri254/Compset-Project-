import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# Read the Excel file
excel_file = 'Compset_Analysis_With_ML_Features_Final.xlsx'
df = pd.read_excel(excel_file, sheet_name='Features', header=None)

print("Creating Final Scoring with Opening Year & Renovation...")

# Extract hotel names
hotel_names = df.iloc[0, 2:].tolist()
print(f"\nHotels found: {len(hotel_names)}")

# Exact row mapping
feature_row_map = {
    'star_rating': 4,
    'distance_km': 2,
    'total_keys': 15,
    'num_rooms': 17,
    'num_suites': 16,
    'num_apartments': 18,
    'tripadvisor_score': 6,
    'booking_score': 5,
    'restaurants_count': 25,
    'bars_count': 27,
    'meeting_rooms': 32,
    'meeting_space_sqm': 33,
    'largest_ballroom': 36,
    'spa_name': 40,
    'treatment_rooms': 41,
    'gym': 42,
    'kids_club': 45,
    'pool': 47,
    'pool_count': 48,
    'opening_year': 7,
    'last_renovation': 8,
}

# Extract data
hotels_data = []

for hotel in hotel_names:
    col_idx = 2 + hotel_names.index(hotel)
    hotel_data = {'Hotel': hotel}

    for feature, row_idx in feature_row_map.items():
        value = df.iloc[row_idx, col_idx]
        hotel_data[feature] = value

    hotels_data.append(hotel_data)

df_raw = pd.DataFrame(hotels_data)

print("\n" + "="*120)
print("RAW DATA - Opening Year & Renovation")
print("="*120)
print(df_raw[['Hotel', 'opening_year', 'last_renovation', 'star_rating', 'total_keys']].to_string())

# Process data
df_hotels = df_raw.copy()

# Convert numeric fields
numeric_fields = ['star_rating', 'distance_km', 'total_keys', 'num_rooms', 'num_suites',
                  'num_apartments', 'tripadvisor_score', 'booking_score', 'restaurants_count',
                  'bars_count', 'meeting_rooms', 'meeting_space_sqm', 'largest_ballroom',
                  'treatment_rooms', 'pool_count', 'opening_year', 'last_renovation']

for field in numeric_fields:
    df_hotels[field] = pd.to_numeric(df_hotels[field], errors='coerce')

# Binary fields
df_hotels['gym_binary'] = df_hotels['gym'].apply(
    lambda x: 1 if pd.notna(x) and str(x).upper() in ['Y', 'YES', '1', 'TRUE'] else 0
)
df_hotels['kids_club_binary'] = df_hotels['kids_club'].apply(
    lambda x: 1 if pd.notna(x) and str(x).upper() in ['Y', 'YES', '1', 'TRUE'] else 0
)
df_hotels['pool_binary'] = df_hotels['pool'].apply(
    lambda x: 1 if pd.notna(x) and str(x).upper() in ['Y', 'YES', '1', 'TRUE'] else 0
)
df_hotels['spa_binary'] = df_hotels['spa_name'].apply(
    lambda x: 1 if pd.notna(x) and str(x).strip() != '' and str(x).upper() != 'N' else 0
)
df_hotels['sauna_binary'] = 1

# F&B total
df_hotels['fb_total'] = df_hotels['restaurants_count'].fillna(0) + df_hotels['bars_count'].fillna(0)

print("\n" + "="*120)
print("PROCESSED DATA")
print("="*120)
print(df_hotels[['Hotel', 'opening_year', 'last_renovation', 'total_keys']].to_string())

# ========== SCORING ==========
df_points = df_hotels.copy()

# 1. ROOM MIX (max 3)
df_points['room_mix_points'] = 0
for idx, row in df_points.iterrows():
    has_rooms = pd.notna(row['num_rooms']) and row['num_rooms'] > 0
    has_suites = pd.notna(row['num_suites']) and row['num_suites'] > 0
    has_apartments = pd.notna(row['num_apartments']) and row['num_apartments'] > 0

    count = sum([has_rooms, has_suites, has_apartments])
    if count == 3:
        df_points.at[idx, 'room_mix_points'] = 3
    elif count == 2:
        df_points.at[idx, 'room_mix_points'] = 2
    elif count == 1:
        df_points.at[idx, 'room_mix_points'] = 1

# 2. TOTAL KEYS (max 15)
total_keys_vals = df_points['total_keys'].dropna()
if len(total_keys_vals) > 1 and total_keys_vals.max() - total_keys_vals.min() > 0:
    min_val, max_val = total_keys_vals.min(), total_keys_vals.max()
    df_points['total_keys_points'] = df_points['total_keys'].apply(
        lambda x: ((x - min_val) / (max_val - min_val)) * 15 if pd.notna(x) else 0
    )
else:
    df_points['total_keys_points'] = 15

# 3. DISTANCE (max 20, inverted)
distance_vals = df_points['distance_km'].dropna()
if len(distance_vals) > 1 and distance_vals.max() - distance_vals.min() > 0:
    min_val, max_val = distance_vals.min(), distance_vals.max()
    df_points['distance_points'] = df_points['distance_km'].apply(
        lambda x: (1 - (x - min_val) / (max_val - min_val)) * 20 if pd.notna(x) else 0
    )
else:
    df_points['distance_points'] = 20

# 4. TRIPADVISOR (max 10)
ta_vals = df_points['tripadvisor_score'].dropna()
if len(ta_vals) > 1 and ta_vals.max() - ta_vals.min() > 0:
    min_val, max_val = ta_vals.min(), ta_vals.max()
    df_points['tripadvisor_points'] = df_points['tripadvisor_score'].apply(
        lambda x: ((x - min_val) / (max_val - min_val)) * 10 if pd.notna(x) else 0
    )
else:
    df_points['tripadvisor_points'] = 10

# 5. BOOKING.COM (max 10)
bk_vals = df_points['booking_score'].dropna()
if len(bk_vals) > 1 and bk_vals.max() - bk_vals.min() > 0:
    min_val, max_val = bk_vals.min(), bk_vals.max()
    df_points['booking_points'] = df_points['booking_score'].apply(
        lambda x: ((x - min_val) / (max_val - min_val)) * 10 if pd.notna(x) else 0
    )
else:
    df_points['booking_points'] = 10

# 6. MEETING FACILITIES (max 12)
meeting_space_vals = df_points['meeting_space_sqm'].fillna(0)
meeting_rooms_vals = df_points['meeting_rooms'].fillna(0)

if meeting_space_vals.max() > 0:
    space_points = (meeting_space_vals / meeting_space_vals.max()) * 8
else:
    space_points = 0

if meeting_rooms_vals.max() > 0:
    rooms_points = (meeting_rooms_vals / meeting_rooms_vals.max()) * 4
else:
    rooms_points = 0

df_points['meeting_points'] = space_points + rooms_points

# 7. F&B (max 8)
fb_vals = df_points['fb_total'].dropna()
if len(fb_vals) > 1 and fb_vals.max() - fb_vals.min() > 0:
    min_val, max_val = fb_vals.min(), fb_vals.max()
    df_points['fb_points'] = df_points['fb_total'].apply(
        lambda x: ((x - min_val) / (max_val - min_val)) * 8 if pd.notna(x) else 0
    )
else:
    df_points['fb_points'] = 8

# 8. OPENING YEAR (max 5 points - newer is better, older gets less)
current_year = 2025
opening_years = df_points['opening_year'].dropna()
if len(opening_years) > 1 and opening_years.max() - opening_years.min() > 0:
    min_year, max_year = opening_years.min(), opening_years.max()
    # Older hotels get fewer points, newer hotels get more
    df_points['opening_year_points'] = df_points['opening_year'].apply(
        lambda x: ((x - min_year) / (max_year - min_year)) * 5 if pd.notna(x) else 0
    )
else:
    df_points['opening_year_points'] = 5

# 9. LAST RENOVATION (max 5 points - more recent is better)
renovation_years = df_points['last_renovation'].dropna()
if len(renovation_years) > 1 and renovation_years.max() - renovation_years.min() > 0:
    min_year, max_year = renovation_years.min(), renovation_years.max()
    df_points['renovation_points'] = df_points['last_renovation'].apply(
        lambda x: ((x - min_year) / (max_year - min_year)) * 5 if pd.notna(x) else 0
    )
else:
    df_points['renovation_points'] = 5

# 10. Binary features (1 each)
df_points['pool_points'] = df_points['pool_binary']
df_points['gym_points'] = df_points['gym_binary']
df_points['spa_points'] = df_points['spa_binary']
df_points['sauna_points'] = df_points['sauna_binary']
df_points['kids_club_points'] = df_points['kids_club_binary']

# 11. BASE TOTAL
point_cols = ['room_mix_points', 'total_keys_points', 'distance_points',
              'tripadvisor_points', 'booking_points', 'meeting_points', 'fb_points',
              'opening_year_points', 'renovation_points',
              'pool_points', 'gym_points', 'spa_points', 'sauna_points', 'kids_club_points']

df_points['base_total_points'] = df_points[point_cols].sum(axis=1)

# 12. STAR RATING MULTIPLIER
df_points['star_multiplier'] = 1.0
for idx, row in df_points.iterrows():
    if pd.notna(row['star_rating']):
        if row['star_rating'] >= 5:
            df_points.at[idx, 'star_multiplier'] = 1.30
        elif row['star_rating'] >= 4:
            df_points.at[idx, 'star_multiplier'] = 0.70

df_points['total_after_star'] = df_points['base_total_points'] * df_points['star_multiplier']

# 13. FINAL POINTS (NO BRAND PENALTY)
df_points['final_points'] = df_points['total_after_star']

# 14. NORMALIZE TO 0-100
min_final = df_points['final_points'].min()
max_final = df_points['final_points'].max()
if max_final - min_final > 0:
    df_points['normalized_score'] = ((df_points['final_points'] - min_final) / (max_final - min_final)) * 100
else:
    df_points['normalized_score'] = 100

df_points = df_points.sort_values('normalized_score', ascending=False)

# ANALYZE Grand Plaza Movenpick Media City
print("\n" + "="*120)
print("ANALYSIS: Why is Grand Plaza Movenpick Media City scoring so well?")
print("="*120)

gp_movenpick = df_points[df_points['Hotel'] == 'Grand Plaza Movenpick Media City'].iloc[0]
grand_millennium = df_points[df_points['Hotel'] == 'Grand Millennium Dubai'].iloc[0]

comparison_features = [
    ('Hotel', ['Hotel']),
    ('Star Rating', ['star_rating', 'star_multiplier']),
    ('Total Keys', ['total_keys', 'total_keys_points']),
    ('Distance', ['distance_km', 'distance_points']),
    ('TripAdvisor', ['tripadvisor_score', 'tripadvisor_points']),
    ('Booking.com', ['booking_score', 'booking_points']),
    ('Meeting', ['meeting_space_sqm', 'meeting_rooms', 'meeting_points']),
    ('F&B', ['fb_total', 'fb_points']),
    ('Opening Year', ['opening_year', 'opening_year_points']),
    ('Renovation', ['last_renovation', 'renovation_points']),
    ('Room Mix', ['room_mix_points']),
]

print(f"\nComparing: Grand Plaza Movenpick Media City vs Grand Millennium Dubai\n")
print(f"{'Feature':<20} {'GP Movenpick':<30} {'Grand Millennium':<30} {'Winner'}")
print("-" * 120)

for feature_name, cols in comparison_features:
    if len(cols) == 1:
        gp_val = gp_movenpick[cols[0]]
        gm_val = grand_millennium[cols[0]]

        if feature_name == 'Hotel':
            print(f"{feature_name:<20} {str(gp_val):<30} {str(gm_val):<30}")
        else:
            winner = "GP Movenpick" if gp_val > gm_val else "Grand Millennium" if gm_val > gp_val else "Tie"
            print(f"{feature_name:<20} {gp_val:<30.2f} {gm_val:<30.2f} {winner}")

    elif len(cols) == 2:
        raw_col, points_col = cols
        gp_raw = gp_movenpick[raw_col]
        gp_pts = gp_movenpick[points_col]
        gm_raw = grand_millennium[raw_col]
        gm_pts = grand_millennium[points_col]

        gp_str = f"{gp_raw} -> {gp_pts:.1f}pts"
        gm_str = f"{gm_raw} -> {gm_pts:.1f}pts"
        winner = "GP Movenpick" if gp_pts > gm_pts else "Grand Millennium" if gm_pts > gp_pts else "Tie"

        print(f"{feature_name:<20} {gp_str:<30} {gm_str:<30} {winner}")

    elif len(cols) == 3:
        val1_col, val2_col, points_col = cols
        gp_val1 = gp_movenpick[val1_col]
        gp_val2 = gp_movenpick[val2_col]
        gp_pts = gp_movenpick[points_col]
        gm_val1 = grand_millennium[val1_col]
        gm_val2 = grand_millennium[val2_col]
        gm_pts = grand_millennium[points_col]

        gp_str = f"{gp_val1}/{gp_val2} -> {gp_pts:.1f}pts"
        gm_str = f"{gm_val1}/{gm_val2} -> {gm_pts:.1f}pts"
        winner = "GP Movenpick" if gp_pts > gm_pts else "Grand Millennium" if gm_pts > gp_pts else "Tie"

        print(f"{feature_name:<20} {gp_str:<30} {gm_str:<30} {winner}")

print("\n" + "-" * 120)
print(f"{'BASE TOTAL':<20} {gp_movenpick['base_total_points']:<30.2f} {grand_millennium['base_total_points']:<30.2f}")
print(f"{'Star Multiplier':<20} {gp_movenpick['star_multiplier']:<30.2f} {grand_millennium['star_multiplier']:<30.2f}")
print(f"{'FINAL POINTS':<20} {gp_movenpick['final_points']:<30.2f} {grand_millennium['final_points']:<30.2f}")
print(f"{'NORMALIZED SCORE':<20} {gp_movenpick['normalized_score']:<30.2f} {grand_millennium['normalized_score']:<30.2f}")

print("\n" + "="*120)
print("KEY REASONS GP Movenpick is winning:")
print("="*120)

reasons = []
if gp_movenpick['tripadvisor_points'] > grand_millennium['tripadvisor_points']:
    diff = gp_movenpick['tripadvisor_points'] - grand_millennium['tripadvisor_points']
    reasons.append(f"1. TRIPADVISOR RATING: {gp_movenpick['tripadvisor_score']:.1f} vs {grand_millennium['tripadvisor_score']:.1f} (+{diff:.1f} points advantage)")

if gp_movenpick['booking_points'] > grand_millennium['booking_points']:
    diff = gp_movenpick['booking_points'] - grand_millennium['booking_points']
    reasons.append(f"2. BOOKING.COM RATING: {gp_movenpick['booking_score']:.1f} vs {grand_millennium['booking_score']:.1f} (+{diff:.1f} points advantage)")

if gp_movenpick['meeting_points'] > grand_millennium['meeting_points']:
    diff = gp_movenpick['meeting_points'] - grand_millennium['meeting_points']
    reasons.append(f"3. MEETING FACILITIES: {gp_movenpick['meeting_rooms']:.0f} rooms vs {grand_millennium['meeting_rooms']:.0f} rooms (+{diff:.1f} points advantage)")

if gp_movenpick['distance_points'] > grand_millennium['distance_points']:
    diff = gp_movenpick['distance_points'] - grand_millennium['distance_points']
    reasons.append(f"4. DISTANCE: {gp_movenpick['distance_km']:.1f}km vs {grand_millennium['distance_km']:.1f}km (+{diff:.1f} points advantage)")

for reason in reasons:
    print(f"  {reason}")

print("\n" + "="*120)
print("FINAL SCORES - Top 10")
print("="*120)
print(df_points[['Hotel', 'final_points', 'normalized_score']].head(10).to_string())

# ========== CREATE EXCEL ==========
wb = load_workbook(excel_file)

if 'Feature Points & Scoring' in wb.sheetnames:
    del wb['Feature Points & Scoring']

ws = wb.create_sheet('Feature Points & Scoring', 0)

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Title
ws.merge_cells('A1:D1')
ws['A1'] = 'FINAL FEATURE POINTS & SCORING (No Brand Penalty)'
ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
ws['A1'].fill = PatternFill(start_color='0070C0', end_color='0070C0', fill_type='solid')
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

ws.merge_cells('A2:D2')
ws['A2'] = 'Includes Opening Year & Renovation Scoring'
ws['A2'].font = Font(italic=True, size=10)
ws['A2'].alignment = Alignment(horizontal='center')

current_row = 4

# Headers
headers = [
    'Hotel',
    'Star\nRating',
    'Star\nMultiplier',
    'Room Mix\n(R/S/A)',
    'Room Mix\nPoints (3)',
    'Total\nKeys',
    'Total Keys\nPoints (15)',
    'Distance\nKM',
    'Distance\nPoints (20)',
    'TripAdvisor\nScore',
    'TripAdvisor\nPoints (10)',
    'Booking.com\nScore',
    'Booking.com\nPoints (10)',
    'Meeting\nSqm',
    'Meeting\nRooms',
    'Meeting\nPoints (12)',
    'F&B\nTotal',
    'F&B\nPoints (8)',
    'Opening\nYear',
    'Opening\nPoints (5)',
    'Last\nRenovation',
    'Renovation\nPoints (5)',
    'Pool\n(1/0)',
    'Gym\n(1/0)',
    'Spa\n(1/0)',
    'Sauna\n(1/0)',
    'Kids Club\n(1/0)',
    'Base Total\nPoints',
    'After Star\nAdjustment',
    'Final\nPoints',
    'Normalized\nScore (0-100)'
]

for col_idx, header in enumerate(headers, start=1):
    cell = ws.cell(row=current_row, column=col_idx, value=header)
    cell.font = Font(bold=True, color='FFFFFF', size=9)
    cell.fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

current_row += 1
header_row = current_row - 1

# Write data
for idx, row_data in df_points.iterrows():
    col = 1

    # Hotel
    cell = ws.cell(row=current_row, column=col, value=row_data['Hotel'])
    cell.font = Font(bold=True, size=9)
    cell.border = thin_border
    if 'grand millennium dubai' in str(row_data['Hotel']).lower():
        cell.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    col += 1

    # Star Rating
    cell = ws.cell(row=current_row, column=col, value=row_data['star_rating'] if pd.notna(row_data['star_rating']) else '-')
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border
    cell.font = Font(size=9)
    col += 1

    # Star Multiplier
    cell = ws.cell(row=current_row, column=col, value=row_data['star_multiplier'])
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border
    cell.number_format = '0.00'
    cell.font = Font(size=9)
    col += 1

    # Room Mix
    room_mix_parts = []
    if pd.notna(row_data['num_rooms']) and row_data['num_rooms'] > 0:
        room_mix_parts.append('R')
    if pd.notna(row_data['num_suites']) and row_data['num_suites'] > 0:
        room_mix_parts.append('S')
    if pd.notna(row_data['num_apartments']) and row_data['num_apartments'] > 0:
        room_mix_parts.append('A')

    cell = ws.cell(row=current_row, column=col, value='+'.join(room_mix_parts) if room_mix_parts else '-')
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border
    cell.font = Font(size=9)
    col += 1

    # Room Mix Points
    cell = ws.cell(row=current_row, column=col, value=row_data['room_mix_points'])
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border
    cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
    cell.font = Font(size=9)
    col += 1

    # Define value-point pairs
    value_point_pairs = [
        ('total_keys', 'total_keys_points', '0'),
        ('distance_km', 'distance_points', '0.00'),
        ('tripadvisor_score', 'tripadvisor_points', '0.0'),
        ('booking_score', 'booking_points', '0.0'),
        ('meeting_space_sqm', None, '0'),
        ('meeting_rooms', None, '0'),
        (None, 'meeting_points', '0.0'),
        ('fb_total', 'fb_points', '0'),
        ('opening_year', 'opening_year_points', '0'),
        ('last_renovation', 'renovation_points', '0'),
    ]

    for val_col, pts_col, num_format in value_point_pairs:
        if val_col:
            # Raw value
            cell = ws.cell(row=current_row, column=col, value=row_data[val_col] if pd.notna(row_data[val_col]) else '-')
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
            if pd.notna(row_data[val_col]):
                cell.number_format = num_format
            cell.font = Font(size=9)
            col += 1

        if pts_col:
            # Points
            cell = ws.cell(row=current_row, column=col, value=row_data[pts_col])
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
            cell.number_format = '0.0'
            cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
            cell.font = Font(size=9)
            col += 1

    # Binary features
    for binary_field in ['pool_binary', 'gym_binary', 'spa_binary', 'sauna_binary', 'kids_club_binary']:
        cell = ws.cell(row=current_row, column=col, value=int(row_data[binary_field]))
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
        cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
        cell.font = Font(size=9)
        col += 1

    # Base Total
    cell = ws.cell(row=current_row, column=col, value=row_data['base_total_points'])
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border
    cell.number_format = '0.0'
    cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    cell.font = Font(bold=True, size=9)
    col += 1

    # After Star
    cell = ws.cell(row=current_row, column=col, value=row_data['total_after_star'])
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border
    cell.number_format = '0.0'
    cell.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    cell.font = Font(bold=True, size=9)
    col += 1

    # Final Points
    cell = ws.cell(row=current_row, column=col, value=row_data['final_points'])
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border
    cell.number_format = '0.0'
    cell.fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    cell.font = Font(bold=True, size=9)
    col += 1

    # Normalized Score
    cell = ws.cell(row=current_row, column=col, value=row_data['normalized_score'])
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border
    cell.number_format = '0.00'
    cell.fill = PatternFill(start_color='C6E0B4', end_color='C6E0B4', fill_type='solid')
    cell.font = Font(bold=True, size=9)

    current_row += 1

# Column widths
ws.column_dimensions['A'].width = 35
for col_idx in range(2, len(headers) + 1):
    ws.column_dimensions[get_column_letter(col_idx)].width = 10

ws.row_dimensions[header_row].height = 45

# Legend
current_row += 2
ws[f'A{current_row}'] = 'SCORING METHODOLOGY (NO BRAND PENALTY):'
ws[f'A{current_row}'].font = Font(bold=True, size=11)
current_row += 1

legend = [
    ('Total Keys:', 'Total rooms + suites + apartments - Max 15 pts'),
    ('Room Mix:', 'R+S+A = 3pts, Any 2 = 2pts, Any 1 = 1pt'),
    ('Distance:', 'Inverted - Closer = More Points - Max 20 pts'),
    ('TripAdvisor & Booking:', 'Normalized ratings - Max 10 pts each'),
    ('Meeting:', 'Space sqm (8pts) + Room count (4pts) = Max 12 pts'),
    ('F&B:', 'Restaurants + Bars - Normalized - Max 8 pts'),
    ('Opening Year:', 'Newer hotels get MORE points - Max 5 pts'),
    ('Renovation:', 'More recent renovation = MORE points - Max 5 pts'),
    ('Binary:', 'Pool, Gym, Spa, Sauna, Kids Club = 1pt each'),
    ('Star Multiplier:', '5-star = 1.30 (+30%), 4-star = 0.70 (-30%)'),
    ('NO PENALTIES:', 'Brand penalties REMOVED'),
]

for label, desc in legend:
    ws[f'A{current_row}'] = f"  {label}"
    ws[f'A{current_row}'].font = Font(bold=True, size=9)
    ws.merge_cells(f'B{current_row}:G{current_row}')
    ws[f'B{current_row}'] = desc
    ws[f'B{current_row}'].font = Font(size=9)
    current_row += 1

wb.save(excel_file)
wb.close()

print(f"\n[SUCCESS] Final scoring with age factors created!")
print(f"File: {excel_file}")
