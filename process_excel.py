#!/usr/bin/env python3
"""
Script to process and fill in the STR Competitor Set Analysis Excel file
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
import json
from pathlib import Path

# Load JSON data
def load_json_data():
    """Load all JSON data files"""
    base_path = Path('/home/gee_devops254/Downloads/Compset Tool')

    # Load hotel research data
    with open(base_path / 'additional_hotel_research.json', 'r') as f:
        additional_research = json.load(f)

    with open(base_path / 'hotel_research_database.json', 'r') as f:
        research_db = json.load(f)

    with open(base_path / 'filtered_hotels.json', 'r') as f:
        filtered_hotels = json.load(f)

    with open(base_path / 'Compset by distance.json', 'r') as f:
        distance_data_raw = f.read()
        # Extract JSON from the markdown file
        start = distance_data_raw.find('```json')
        end = distance_data_raw.find('```', start + 7)
        json_str = distance_data_raw[start+7:end].strip()
        distance_data = json.loads(json_str)

    # Merge all hotel data
    hotel_data = {}
    hotel_data.update(additional_research)
    hotel_data.update(research_db)

    return {
        'hotel_data': hotel_data,
        'filtered_hotels': filtered_hotels,
        'distance_data': distance_data
    }

# Load Excel file
excel_path = '/home/gee_devops254/Downloads/Compset Tool/NEW - STR Competitor Set Analysis.xlsx'
print(f"Loading Excel file: {excel_path}")

try:
    wb = load_workbook(excel_path)
    print(f"✓ Excel file loaded successfully")
    print(f"✓ Sheet names: {wb.sheetnames}")

    # Load JSON data
    print("\nLoading JSON data...")
    data = load_json_data()
    print(f"✓ Loaded data for {len(data['hotel_data'])} hotels")
    print(f"✓ Filtered hotels list: {len(data['filtered_hotels'])} hotels")
    print(f"✓ Distance data: {data['distance_data']['total_hotels_found']} hotels")

    # Display first few hotel names from distance data
    print("\n5-Star Hotels within proximity (sorted by distance):")
    luxury_hotels = [h for h in data['distance_data']['hotels']
                     if h['star_rating'] == 5 and h['distance_km'] <= 3.0]
    for hotel in luxury_hotels[:15]:
        print(f"  - {hotel['name']} ({hotel['distance_km']} km)")

    print("\nAvailable data fields for hotels:")
    if data['hotel_data']:
        sample_hotel = list(data['hotel_data'].keys())[0]
        print(f"Sample hotel: {sample_hotel}")
        print(f"Fields: {list(data['hotel_data'][sample_hotel].keys())}")

except Exception as e:
    print(f"✗ Error: {e}")
    import traceback
    traceback.print_exc()
