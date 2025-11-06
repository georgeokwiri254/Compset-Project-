from openpyxl import load_workbook
from openpyxl.chart import BarChart, PieChart, LineChart, ScatterChart, Reference, Series
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.marker import Marker
from openpyxl.drawing.fill import SolidColorFillProperties, ColorChoice
from openpyxl.styles import Font, PatternFill, Alignment
import openpyxl

# Load the Excel file
excel_file = r'C:\Users\reservations\Desktop\Compset Tool\Grand_Millennium_Dubai_CompSet_Analysis.xlsx'
wb = load_workbook(excel_file)

# Create a new sheet for charts if it doesn't exist
if 'Charts & Diagrams' in wb.sheetnames:
    del wb['Charts & Diagrams']

chart_sheet = wb.create_sheet('Charts & Diagrams', 1)

# Add title
chart_sheet['A1'] = 'GRAND MILLENNIUM DUBAI - VISUAL ANALYSIS DASHBOARD'
chart_sheet['A1'].font = Font(bold=True, size=16, color='FFFFFF')
chart_sheet['A1'].fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
chart_sheet.merge_cells('A1:P1')
chart_sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
chart_sheet.row_dimensions[1].height = 30

# Get the Hotels sheet
hotels_ws = wb['Hotels']

# Prepare data for charts - get top 15 hotels with scores
chart_data = []
for row in range(2, min(hotels_ws.max_row + 1, 100)):
    score = hotels_ws.cell(row=row, column=34).value
    if isinstance(score, (int, float)) and score > 0:
        name = hotels_ws.cell(row=row, column=2).value
        distance = hotels_ws.cell(row=row, column=5).value
        rooms = hotels_ws.cell(row=row, column=7).value
        meeting = hotels_ws.cell(row=row, column=12).value
        star = hotels_ws.cell(row=row, column=3).value
        trip_rating = hotels_ws.cell(row=row, column=20).value
        booking_rating = hotels_ws.cell(row=row, column=22).value
        recommendation = hotels_ws.cell(row=row, column=35).value

        chart_data.append({
            'name': name if name and len(name) <= 30 else (name[:27] + "..." if name else "Unknown"),
            'score': score,
            'distance': distance if distance else 0,
            'rooms': rooms if rooms else 0,
            'meeting': meeting if meeting else 0,
            'star': star if star else 0,
            'trip_rating': trip_rating if trip_rating else 0,
            'booking_rating': booking_rating if booking_rating else 0,
            'recommendation': recommendation if recommendation else "TBD"
        })

# Sort by score
chart_data.sort(key=lambda x: x['score'], reverse=True)
top_15 = chart_data[:15]

# Create data table on the chart sheet for chart references
start_row = 3
chart_sheet['A3'] = 'Hotel Name'
chart_sheet['B3'] = 'Score'
chart_sheet['C3'] = 'Distance'
chart_sheet['D3'] = 'Rooms'
chart_sheet['E3'] = 'Meeting Sqm'
chart_sheet['F3'] = 'Star Rating'
chart_sheet['G3'] = 'TripAdvisor'
chart_sheet['H3'] = 'Booking.com'
chart_sheet['I3'] = 'Recommendation'

# Header formatting
for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
    chart_sheet[f'{col}3'].font = Font(bold=True, color='FFFFFF')
    chart_sheet[f'{col}3'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')

# Populate data
for idx, hotel in enumerate(top_15, start=4):
    chart_sheet[f'A{idx}'] = hotel['name']
    chart_sheet[f'B{idx}'] = hotel['score']
    chart_sheet[f'C{idx}'] = hotel['distance']
    chart_sheet[f'D{idx}'] = hotel['rooms']
    chart_sheet[f'E{idx}'] = hotel['meeting']
    chart_sheet[f'F{idx}'] = hotel['star']
    chart_sheet[f'G{idx}'] = hotel['trip_rating']
    chart_sheet[f'H{idx}'] = hotel['booking_rating']
    chart_sheet[f'I{idx}'] = hotel['recommendation']

# Adjust column widths
chart_sheet.column_dimensions['A'].width = 35
chart_sheet.column_dimensions['B'].width = 10
chart_sheet.column_dimensions['C'].width = 10
chart_sheet.column_dimensions['D'].width = 10
chart_sheet.column_dimensions['E'].width = 12
chart_sheet.column_dimensions['F'].width = 12
chart_sheet.column_dimensions['G'].width = 12
chart_sheet.column_dimensions['H'].width = 14
chart_sheet.column_dimensions['I'].width = 18

print("Data table created successfully")

# CHART 1: Bar Chart - Top 10 Hotels by Overall Score
chart1 = BarChart()
chart1.type = "col"
chart1.style = 11
chart1.title = "Top 10 Hotels by Overall Compset Score"
chart1.y_axis.title = 'Compset Fitness Score (out of 100)'
chart1.x_axis.title = 'Hotels'

# Add data
data = Reference(chart_sheet, min_col=2, min_row=3, max_row=13)
cats = Reference(chart_sheet, min_col=1, min_row=4, max_row=13)
chart1.add_data(data, titles_from_data=True)
chart1.set_categories(cats)

# Customize appearance
chart1.height = 15
chart1.width = 26
chart1.legend = None

# Add to sheet
chart_sheet.add_chart(chart1, "K3")
print("Chart 1 (Bar Chart - Top 10 Scores) created")

# CHART 2: Horizontal Bar Chart - Review Ratings Comparison
chart2 = BarChart()
chart2.type = "bar"
chart2.style = 12
chart2.title = "Guest Review Ratings Comparison (Top 10)"
chart2.x_axis.title = 'Rating Score'
chart2.y_axis.title = 'Hotels'

# Add TripAdvisor data
data_trip = Reference(chart_sheet, min_col=7, min_row=3, max_row=13)
data_booking = Reference(chart_sheet, min_col=8, min_row=3, max_row=13)
cats = Reference(chart_sheet, min_col=1, min_row=4, max_row=13)

chart2.add_data(data_trip, titles_from_data=True)
chart2.add_data(data_booking, titles_from_data=True)
chart2.set_categories(cats)

# Customize
chart2.height = 15
chart2.width = 26
chart2.grouping = "stacked"
chart2.overlap = 100

# Add to sheet
chart_sheet.add_chart(chart2, "K28")
print("Chart 2 (Review Ratings Comparison) created")

# CHART 3: Scatter Chart - Distance vs Score
chart3 = ScatterChart()
chart3.title = "Distance vs. Compset Score Analysis"
chart3.style = 13
chart3.x_axis.title = 'Distance from Grand Millennium (km)'
chart3.y_axis.title = 'Compset Fitness Score'

# Add data
xvalues = Reference(chart_sheet, min_col=3, min_row=4, max_row=18)
yvalues = Reference(chart_sheet, min_col=2, min_row=4, max_row=18)
series = Series(values=yvalues, xvalues=xvalues, title="Hotels")
chart3.series.append(series)

# Customize markers
s1 = chart3.series[0]
s1.marker.symbol = "circle"
s1.marker.size = 10
s1.graphicalProperties.line.solidFill = "FF0000"
s1.graphicalProperties.solidFill = "FF0000"

chart3.height = 15
chart3.width = 26

# Add to sheet
chart_sheet.add_chart(chart3, "K53")
print("Chart 3 (Scatter - Distance vs Score) created")

# CHART 4: Pie Chart - Recommendation Distribution
# Count recommendations
rec_counts = {}
for hotel in chart_data:
    rec = hotel['recommendation']
    rec_counts[rec] = rec_counts.get(rec, 0) + 1

# Create data for pie chart in sheet
pie_start_row = 20
chart_sheet['K20'] = 'Recommendation Category'
chart_sheet['L20'] = 'Count'
chart_sheet['K20'].font = Font(bold=True)
chart_sheet['L20'].font = Font(bold=True)

row = 21
for rec, count in rec_counts.items():
    chart_sheet[f'K{row}'] = rec
    chart_sheet[f'L{row}'] = count
    row += 1

# Create pie chart
chart4 = PieChart()
chart4.title = "Compset Recommendation Distribution"
chart4.style = 10

# Add data
data = Reference(chart_sheet, min_col=12, min_row=20, max_row=20+len(rec_counts))
cats = Reference(chart_sheet, min_col=11, min_row=21, max_row=20+len(rec_counts))
chart4.add_data(data, titles_from_data=True)
chart4.set_categories(cats)

chart4.height = 12
chart4.width = 16

# Add to sheet
chart_sheet.add_chart(chart4, "W3")
print("Chart 4 (Pie Chart - Recommendations) created")

# CHART 5: Bar Chart - Facilities Comparison (Top 5)
chart5 = BarChart()
chart5.type = "col"
chart5.style = 10
chart5.title = "Meeting Space Comparison (Top 10)"
chart5.y_axis.title = 'Meeting Space (sqm)'
chart5.x_axis.title = 'Hotels'

data = Reference(chart_sheet, min_col=5, min_row=3, max_row=13)
cats = Reference(chart_sheet, min_col=1, min_row=4, max_row=13)
chart5.add_data(data, titles_from_data=True)
chart5.set_categories(cats)

chart5.height = 15
chart5.width = 26
chart5.legend = None

# Add to sheet
chart_sheet.add_chart(chart5, "W20")
print("Chart 5 (Meeting Space Comparison) created")

# CHART 6: Bar Chart - Room Count Comparison
chart6 = BarChart()
chart6.type = "col"
chart6.style = 11
chart6.title = "Total Room Inventory Comparison (Top 10)"
chart6.y_axis.title = 'Number of Rooms'
chart6.x_axis.title = 'Hotels'

data = Reference(chart_sheet, min_col=4, min_row=3, max_row=13)
cats = Reference(chart_sheet, min_col=1, min_row=4, max_row=13)
chart6.add_data(data, titles_from_data=True)
chart6.set_categories(cats)

chart6.height = 15
chart6.width = 26
chart6.legend = None

# Add to sheet
chart_sheet.add_chart(chart6, "W45")
print("Chart 6 (Room Count Comparison) created")

# Add summary statistics box
stats_row = 30
chart_sheet[f'A{stats_row}'] = 'ANALYSIS SUMMARY STATISTICS'
chart_sheet[f'A{stats_row}'].font = Font(bold=True, size=12, color='FFFFFF')
chart_sheet[f'A{stats_row}'].fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
chart_sheet.merge_cells(f'A{stats_row}:I{stats_row}')

stats_row += 2
chart_sheet[f'A{stats_row}'] = 'Total Hotels Analyzed:'
chart_sheet[f'B{stats_row}'] = len(chart_data)
stats_row += 1
chart_sheet[f'A{stats_row}'] = 'Average Compset Score:'
avg_score = sum(h['score'] for h in chart_data) / len(chart_data) if chart_data else 0
chart_sheet[f'B{stats_row}'] = f'{avg_score:.1f}/100'
stats_row += 1
chart_sheet[f'A{stats_row}'] = 'Highest Score:'
chart_sheet[f'B{stats_row}'] = f"{chart_data[0]['score']:.1f} ({chart_data[0]['name']})"
stats_row += 1
chart_sheet[f'A{stats_row}'] = 'Average Distance:'
avg_dist = sum(h['distance'] for h in chart_data) / len(chart_data) if chart_data else 0
chart_sheet[f'B{stats_row}'] = f'{avg_dist:.2f} km'
stats_row += 1
chart_sheet[f'A{stats_row}'] = 'Primary Compset Candidates:'
primary_count = sum(1 for h in chart_data if h['score'] >= 90)
chart_sheet[f'B{stats_row}'] = primary_count
stats_row += 1
chart_sheet[f'A{stats_row}'] = 'Secondary Compset Candidates:'
secondary_count = sum(1 for h in chart_data if 75 <= h['score'] < 90)
chart_sheet[f'B{stats_row}'] = secondary_count

# Format statistics
for r in range(32, stats_row + 1):
    chart_sheet[f'A{r}'].font = Font(bold=True)
    chart_sheet[f'B{r}'].font = Font(size=11)

# Save the workbook
wb.save(excel_file)

print("\n" + "="*80)
print("EXCEL CHARTS CREATED SUCCESSFULLY!")
print("="*80)
print(f"\nFile: {excel_file}")
print(f"Sheet: 'Charts & Diagrams'")
print(f"\nCharts created:")
print("1. Top 10 Hotels by Overall Compset Score (Bar Chart)")
print("2. Guest Review Ratings Comparison (Horizontal Bar Chart)")
print("3. Distance vs. Compset Score Analysis (Scatter Chart)")
print("4. Compset Recommendation Distribution (Pie Chart)")
print("5. Meeting Space Comparison (Bar Chart)")
print("6. Total Room Inventory Comparison (Bar Chart)")
print("\nAll charts are fully interactive Excel charts embedded in the workbook.")
print("="*80)
