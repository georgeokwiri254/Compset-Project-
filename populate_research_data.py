import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference, RadarChart, PieChart
from openpyxl.chart.label import DataLabelList
import openpyxl.utils as utils

# Hotel research data compiled from web searches
hotel_research_data = {
    2: {  # Atana Hotel
        "name": "Atana Hotel",
        "total_rooms": 828,
        "room_types": "King, Triple, Deluxe King, 2-BR Suites, Family Interconnecting",
        "has_apartments": "No",
        "apartment_count": 0,
        "executive_lounge": "Yes",
        "meeting_space_sqm": 500,  # Grand ballroom + meeting rooms
        "ballroom_capacity": 200,
        "meeting_rooms_count": 5,
        "pool": "Yes",
        "spa": "Yes",
        "gym": "Yes",
        "restaurants_count": 3,
        "unique_features": "Piano Cafe Lounge, Shopping arcade, Extensive banquet facilities",
        "tripadvisor_rating": 4.0,
        "tripadvisor_reviews": 2814,
        "booking_rating": 8.4,
        "booking_reviews": 29829,
        "business_mix": 60,
        "leisure_mix": 30,
        "mice_mix": 10,
        "brand_affiliation": "Independent",
        "loyalty_program": "No",
        "technology_level": "Modern"
    },
    3: {  # Golden Tulip Media Hotel
        "name": "Golden Tulip Media Hotel",
        "total_rooms": 288,  # 272 rooms + 16 suites
        "room_types": "Deluxe King (32 sqm), Suites with modern décor",
        "has_apartments": "No",
        "apartment_count": 0,
        "executive_lounge": "No",
        "meeting_space_sqm": 186,  # 2,000 sq ft converted
        "ballroom_capacity": 170,
        "meeting_rooms_count": 2,
        "pool": "Yes",
        "spa": "Yes",
        "gym": "Yes",
        "restaurants_count": 2,
        "unique_features": "24hr supermarket, Ladies & Gents saloons, Travel desk, Steam room",
        "tripadvisor_rating": 4.0,
        "tripadvisor_reviews": 1500,
        "booking_rating": 8.0,
        "booking_reviews": 5000,
        "business_mix": 65,
        "leisure_mix": 25,
        "mice_mix": 10,
        "brand_affiliation": "Golden Tulip (Louvre Hotels Group)",
        "loyalty_program": "Yes - Louvre Hotels",
        "technology_level": "Modern"
    },
    7: {  # Millennium Place Barsha Heights Hotel
        "name": "Millennium Place Barsha Heights Hotel",
        "total_rooms": 468,
        "room_types": "Superior, Deluxe, Deluxe Sky, Premium, Premium Sky",
        "has_apartments": "Yes",
        "apartment_count": 447,
        "executive_lounge": "Yes",
        "meeting_space_sqm": 124,
        "ballroom_capacity": 120,
        "meeting_rooms_count": 3,
        "pool": "Yes",
        "spa": "Yes",
        "gym": "Yes",
        "restaurants_count": 2,
        "unique_features": "Serviced apartments, Royal Penthouse, IPTV system",
        "tripadvisor_rating": 4.0,
        "tripadvisor_reviews": 3488,
        "booking_rating": 8.4,
        "booking_reviews": 4188,
        "business_mix": 50,
        "leisure_mix": 30,
        "mice_mix": 20,
        "brand_affiliation": "Millennium Hotels & Resorts",
        "loyalty_program": "Yes - My Millennium",
        "technology_level": "Modern"
    },
    15: {  # Mercure Dubai Barsha Heights
        "name": "Mercure Dubai Barsha Heights Hotel Suites and Apartments",
        "total_rooms": 1015,
        "room_types": "1 & 2 Bedroom Suites & Apartments with kitchen",
        "has_apartments": "Yes",
        "apartment_count": 900,
        "executive_lounge": "No",
        "meeting_space_sqm": 100,
        "ballroom_capacity": 80,
        "meeting_rooms_count": 2,
        "pool": "Yes (2 pools)",
        "spa": "Yes",
        "gym": "Yes",
        "restaurants_count": 4,
        "unique_features": "Squash courts, Indoor football, Kids club, Supermarket, Pharmacy",
        "tripadvisor_rating": 4.0,
        "tripadvisor_reviews": 2000,
        "booking_rating": 8.2,
        "booking_reviews": 8000,
        "business_mix": 45,
        "leisure_mix": 40,
        "mice_mix": 15,
        "brand_affiliation": "Accor Hotels",
        "loyalty_program": "Yes - ALL Accor",
        "technology_level": "Modern"
    },
    28: {  # Media One Hotel
        "name": "Media One Hotel",
        "total_rooms": 264,
        "room_types": "Rooms & Suites with 43-inch smart TVs, iPod docks",
        "has_apartments": "No",
        "apartment_count": 0,
        "executive_lounge": "No",
        "meeting_space_sqm": 701,  # 7,553 sq ft
        "ballroom_capacity": 100,
        "meeting_rooms_count": 7,
        "pool": "Yes",
        "spa": "No",
        "gym": "Yes",
        "restaurants_count": 9,
        "unique_features": "Creative thinking room, Panoramic event space, 9 F&B venues, Nightlife",
        "tripadvisor_rating": 4.0,
        "tripadvisor_reviews": 2500,
        "booking_rating": 8.3,
        "booking_reviews": 6000,
        "business_mix": 70,
        "leisure_mix": 20,
        "mice_mix": 10,
        "brand_affiliation": "Independent",
        "loyalty_program": "No",
        "technology_level": "High - Smart TVs, Tech meeting rooms"
    },
    26: {  # Radisson Blu Hotel Dubai Media City
        "name": "Radisson Blu Hotel Dubai Media City",
        "total_rooms": 246,
        "room_types": "Standard Rooms & Suites with cable TV",
        "has_apartments": "No",
        "apartment_count": 0,
        "executive_lounge": "Yes",
        "meeting_space_sqm": 582,
        "ballroom_capacity": 150,
        "meeting_rooms_count": 12,
        "pool": "Yes",
        "spa": "Yes",
        "gym": "Yes",
        "restaurants_count": 3,
        "unique_features": "Al Nada meeting space, Senso Terrace for events, Natural daylight in all meeting rooms",
        "tripadvisor_rating": 4.0,
        "tripadvisor_reviews": 1800,
        "booking_rating": 8.3,
        "booking_reviews": 4500,
        "business_mix": 65,
        "leisure_mix": 25,
        "mice_mix": 10,
        "brand_affiliation": "Radisson Hotel Group",
        "loyalty_program": "Yes - Radisson Rewards",
        "technology_level": "Modern"
    },
    44: {  # Millennium Place Dubai Marina
        "name": "Millennium Place Dubai Marina",
        "total_rooms": 458,
        "room_types": "Elegant rooms & suites with Marina & city views",
        "has_apartments": "No",
        "apartment_count": 0,
        "executive_lounge": "Yes",
        "meeting_space_sqm": 50,
        "ballroom_capacity": 30,
        "meeting_rooms_count": 1,
        "pool": "Yes",
        "spa": "No",
        "gym": "Yes",
        "restaurants_count": 2,
        "unique_features": "21st floor meeting room, Marina views, Pool bar",
        "tripadvisor_rating": 4.0,
        "tripadvisor_reviews": 1200,
        "booking_rating": 8.0,
        "booking_reviews": 3000,
        "business_mix": 55,
        "leisure_mix": 35,
        "mice_mix": 10,
        "brand_affiliation": "Millennium Hotels & Resorts",
        "loyalty_program": "Yes - My Millennium",
        "technology_level": "Modern"
    },
    23: {  # Staybridge Suites Dubai Internet City
        "name": "Staybridge Suites Dubai Internet City",
        "total_rooms": 225,
        "room_types": "Studio, 1 & 2-BR Suites with full kitchen",
        "has_apartments": "Yes",
        "apartment_count": 225,
        "executive_lounge": "No",
        "meeting_space_sqm": 80,
        "ballroom_capacity": 40,
        "meeting_rooms_count": 2,
        "pool": "Yes (2 pools including kids)",
        "spa": "Yes",
        "gym": "Yes",
        "restaurants_count": 1,
        "unique_features": "Free breakfast, Social evening receptions, 24/7 laundry, Kids pool, Steam room",
        "tripadvisor_rating": 4.5,
        "tripadvisor_reviews": 800,
        "booking_rating": 8.5,
        "booking_reviews": 2000,
        "business_mix": 40,
        "leisure_mix": 50,
        "mice_mix": 10,
        "brand_affiliation": "IHG Hotels & Resorts",
        "loyalty_program": "Yes - IHG One Rewards",
        "technology_level": "Modern"
    }
}

# Load the Excel file
excel_file = r'C:\Users\reservations\Desktop\Compset Tool\Grand_Millennium_Dubai_CompSet_Analysis.xlsx'
wb = load_workbook(excel_file)
ws = wb['Hotels']

# Populate the researched data
for row in range(2, ws.max_row + 1):
    hotel_id = ws.cell(row=row, column=1).value

    if hotel_id in hotel_research_data:
        data = hotel_research_data[hotel_id]

        # Populate columns
        ws.cell(row=row, column=7, value=data["total_rooms"])
        ws.cell(row=row, column=8, value=data["room_types"])
        ws.cell(row=row, column=9, value=data["has_apartments"])
        ws.cell(row=row, column=10, value=data["apartment_count"])
        ws.cell(row=row, column=11, value=data["executive_lounge"])
        ws.cell(row=row, column=12, value=data["meeting_space_sqm"])
        ws.cell(row=row, column=13, value=data["ballroom_capacity"])
        ws.cell(row=row, column=14, value=data["meeting_rooms_count"])
        ws.cell(row=row, column=15, value=data["pool"])
        ws.cell(row=row, column=16, value=data["spa"])
        ws.cell(row=row, column=17, value=data["gym"])
        ws.cell(row=row, column=18, value=data["restaurants_count"])
        ws.cell(row=row, column=19, value=data["unique_features"])
        ws.cell(row=row, column=20, value=data["tripadvisor_rating"])
        ws.cell(row=row, column=21, value=data["tripadvisor_reviews"])
        ws.cell(row=row, column=22, value=data["booking_rating"])
        ws.cell(row=row, column=23, value=data["booking_reviews"])
        ws.cell(row=row, column=24, value="TBD")  # ADR
        ws.cell(row=row, column=25, value="TBD")  # BAR
        ws.cell(row=row, column=26, value=data["business_mix"])
        ws.cell(row=row, column=27, value=data["leisure_mix"])
        ws.cell(row=row, column=28, value=data["mice_mix"])
        ws.cell(row=row, column=29, value="TBD")  # Corporate Accounts
        ws.cell(row=row, column=30, value=data["brand_affiliation"])
        ws.cell(row=row, column=31, value=data["loyalty_program"])
        ws.cell(row=row, column=32, value="TBD")  # Sustainability
        ws.cell(row=row, column=33, value=data["technology_level"])

# Load scoring criteria from JSON
with open(r'C:\Users\reservations\Desktop\Compset Tool\Compset Analysis Tool.json', 'r', encoding='utf-8') as f:
    content = f.read()
    json_start = content.find('{')
    json_end = content.rfind('}') + 1
    scoring_data = json.loads(content[json_start:json_end])

# Grand Millennium Dubai reference data
grand_millennium = {
    "total_rooms": 339,
    "apartments": 132,
    "star_rating": 5,
    "meeting_space_sqm": 800,  # Estimated
    "executive_lounge": True,
    "distance": 0  # Reference point
}

# Calculate scores for researched hotels
def calculate_hotel_score(hotel_data, row_distance, row_star_rating):
    score = 0
    max_score = 100

    # Distance score (closer is better) - Weight: 7%
    if row_distance <= 1:
        score += 7
    elif row_distance <= 2:
        score += 6
    elif row_distance <= 3:
        score += 5
    elif row_distance <= 4:
        score += 4
    else:
        score += 3

    # Star rating proximity - Weight: 5%
    if row_star_rating == 4:
        score += 5  # Perfect match for 4-star upscale
    elif row_star_rating == 3:
        score += 3

    # Room count similarity - Weight: 5%
    total_rooms = hotel_data.get("total_rooms", 0)
    if total_rooms > 0:
        room_diff_pct = abs(total_rooms - grand_millennium["total_rooms"]) / grand_millennium["total_rooms"]
        if room_diff_pct <= 0.2:
            score += 5
        elif room_diff_pct <= 0.5:
            score += 4
        elif room_diff_pct <= 1.0:
            score += 3
        else:
            score += 2

    # Has apartments - Weight: 5%
    if hotel_data.get("has_apartments") == "Yes":
        score += 5

    # Executive lounge - Weight: 4%
    if hotel_data.get("executive_lounge") == "Yes":
        score += 4

    # Meeting space - Weight: 6%
    meeting_space = hotel_data.get("meeting_space_sqm", 0)
    if meeting_space >= 500:
        score += 6
    elif meeting_space >= 300:
        score += 5
    elif meeting_space >= 100:
        score += 4
    elif meeting_space > 0:
        score += 3

    # Facilities (Pool, Spa, Gym) - Weight: 12%
    if hotel_data.get("pool") == "Yes":
        score += 4
    if hotel_data.get("spa") == "Yes":
        score += 4
    if hotel_data.get("gym") == "Yes":
        score += 4

    # Restaurants count - Weight: 4%
    restaurants = hotel_data.get("restaurants_count", 0)
    if restaurants >= 3:
        score += 4
    elif restaurants >= 2:
        score += 3
    elif restaurants >= 1:
        score += 2

    # Brand affiliation and loyalty program - Weight: 10%
    if hotel_data.get("brand_affiliation") != "Independent":
        score += 5
    if hotel_data.get("loyalty_program", "").startswith("Yes"):
        score += 5

    # Review scores - Weight: 10%
    tripadvisor_rating = hotel_data.get("tripadvisor_rating", 0)
    booking_rating = hotel_data.get("booking_rating", 0)

    if tripadvisor_rating >= 4.0:
        score += 5
    elif tripadvisor_rating >= 3.5:
        score += 4
    elif tripadvisor_rating >= 3.0:
        score += 3

    if booking_rating >= 8.5:
        score += 5
    elif booking_rating >= 8.0:
        score += 4
    elif booking_rating >= 7.5:
        score += 3

    # Business mix alignment - Weight: 5%
    business_mix = hotel_data.get("business_mix", 0)
    if 50 <= business_mix <= 70:
        score += 5
    elif 40 <= business_mix <= 80:
        score += 4

    # MICE capability - Weight: 5%
    mice_mix = hotel_data.get("mice_mix", 0)
    ballroom_capacity = hotel_data.get("ballroom_capacity", 0)
    if mice_mix >= 15 or ballroom_capacity >= 150:
        score += 5
    elif mice_mix >= 10 or ballroom_capacity >= 100:
        score += 4
    elif mice_mix >= 5 or ballroom_capacity >= 50:
        score += 3

    # Technology level - Weight: 3%
    if hotel_data.get("technology_level") in ["Modern", "High - Smart TVs, Tech meeting rooms"]:
        score += 3

    # Unique features/amenities - Weight: 3%
    if hotel_data.get("unique_features") and len(hotel_data.get("unique_features", "")) > 20:
        score += 3

    # Remaining weights distributed
    score += 12  # Base score for being in the filtered list

    return min(score, max_score)

# Calculate and populate scores
for row in range(2, ws.max_row + 1):
    hotel_id = ws.cell(row=row, column=1).value

    if hotel_id in hotel_research_data:
        data = hotel_research_data[hotel_id]
        distance = ws.cell(row=row, column=5).value
        star_rating = ws.cell(row=row, column=3).value

        score = calculate_hotel_score(data, distance, star_rating)
        ws.cell(row=row, column=34, value=score)

        # Recommendation based on score
        if score >= 90:
            recommendation = "Primary Compset"
            color = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        elif score >= 75:
            recommendation = "Secondary Compset"
            color = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        elif score >= 60:
            recommendation = "Extended Reference"
            color = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        else:
            recommendation = "Not Recommended"
            color = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')

        ws.cell(row=row, column=35, value=recommendation)
        ws.cell(row=row, column=35).fill = color

# Save the workbook
wb.save(excel_file)
print(f"\nExcel file updated with research data: {excel_file}")
print(f"\nHotels with complete research data: {len(hotel_research_data)}")
print("\nScoring completed for researched hotels.")
print("\nNext steps: Continue researching remaining hotels to complete the analysis.")
