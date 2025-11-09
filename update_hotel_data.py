import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from copy import copy

# Read the main file
main_file = r"C:\Users\reservations\Desktop\Compset Tool\Compset_Analysis_With_ML_Features_Final_Formatted.xlsx"
output_file = r"C:\Users\reservations\Desktop\Compset Tool\Compset_Analysis_With_ML_Features_Updated.xlsx"

# Load workbook
wb = openpyxl.load_workbook(main_file)
ws = wb.active

# Create a mapping of field names to row indices
field_to_row = {}
for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1), start=1):
    cell_value = row[0].value
    if cell_value:
        field_to_row[str(cell_value)] = idx

# Create a mapping of hotel names to column indices
hotel_to_col = {}
for idx, cell in enumerate(ws[1], start=1):
    if cell.value and idx > 2:  # Skip first 2 columns
        hotel_to_col[str(cell.value)] = idx

print("Fields found:", list(field_to_row.keys())[:20])
print("\nHotels found:", list(hotel_to_col.keys()))

# Hotel data to fill in based on web research
hotel_updates = {
    "Grand Millennium Dubai": {
        "Opening Year": "2007",
        "Last Renovation": "2018",
        "Total Function Space (sqm)": "315",
        "Largest Room/Ballroom (sqm)": "280",
        "Maximum capacity (theater style)": "315"
    },
    "Zabeel House The Greens": {
        "Last Renovation": "2022",
        "Treatment rooms (#)": "N/A",
        "Spa total area (sqm)": "N/A"
    },
    "voco Bonnington Dubai": {
        "Opening Year": "2022",  # Rebranding year
        "Percent Suites": "N/A",
        "Total Function Space (sqm)": "640",
        "Largest Room/Ballroom (sqm)": "640",
        "Treatment rooms (#)": "5",
        "Spa total area (sqm)": "N/A",
        "Kid Amenities": "N"
    },
    "Radisson Blu Media City": {
        "Licensed for Shisha (Y/N)": "N",
        "Largest Room/Ballroom (sqm)": "N/A",  # No dedicated ballroom
        "Spa total area (sqm)": "N/A"
    },
    "Media One Hotel": {
        "Last Renovation": "N/A",
        "Largest Room/Ballroom (sqm)": "N/A",
        "Treatment rooms (#)": "N/A",
        "Spa total area (sqm)": "N/A"
    },
    "Pullman JLT": {
        "Last Renovation": "2025",  # Lobby renovation
        "Largest Room/Ballroom (sqm)": "N/A",
        "Maximum capacity (theater style)": "N/A",
        "Spa total area (sqm)": "N/A"
    },
    "Millennium Place Barsha Heights": {
        "Opening Year": "2019",
        "Last Renovation": "N/A",
        "Treatment rooms (#)": "Multiple",
        "Spa total area (sqm)": "N/A"
    },
    "TRYP by Wyndham Dubai": {
        "Last Renovation": "N/A",
        "Treatment rooms (#)": "7",
        "Spa total area (sqm)": "N/A"
    },
    "Mercure Dubai Barsha Heights": {
        "Opening Year": "N/A",
        "Last Renovation": "N/A",
        "Total Function Space (sqm)": "446",
        "Largest Room/Ballroom (sqm)": "N/A",
        "Treatment rooms (#)": "N/A",
        "Spa total area (sqm)": "N/A"
    },
    "Novotel Dubai Al Barsha": {
        "Last Renovation": "N/A",
        "Total Number of Function Rooms": "10",
        "Total Function Space (sqm)": "1800",
        "Number of Ballroom": "N/A",
        "Number of Meeting rooms": "10",
        "Largest Room/Ballroom (sqm)": "N/A",
        "Maximum capacity (theater style)": "250",
        "Spa total area (sqm)": "N/A"
    },
    "Two Seasons": {
        "Opening Year": "N/A",
        "Restaurants Count": "6-8",
        "Restaurants": "Le Grand Cafe, Noodle House, Qutoof Restaurant",
        "Bars Count": "0",
        "Bars": "None (No alcohol)",
        "Total Number of Function Rooms": "N/A",
        "Total Function Space (sqm)": "N/A",
        "Number of Ballroom": "N/A",
        "Number of Meeting rooms": "N/A",
        "Largest Room/Ballroom (sqm)": "N/A",
        "Maximum capacity (theater style)": "N/A",
        "Spa Name": "N/A",
        "Treatment rooms (#)": "N/A",
        "Spa total area (sqm)": "N/A"
    },
    "Staybridge Suites Dubai Internet City": {
        "Total Number of Function Rooms": "Multiple",
        "Total Function Space (sqm)": "227",
        "Number of Meeting rooms": "Multiple",
        "Largest Room/Ballroom (sqm)": "N/A",
        "Maximum capacity (theater style)": "N/A",
        "Spa Name": "N/A",
        "Spa total area (sqm)": "N/A"
    },
    "Arjaan by Rotana - Dubai Media City": {
        "Last Renovation": "N/A",
        "Total Number of Function Rooms": "19",
        "Total Function Space (sqm)": "N/A",
        "Number of Meeting rooms": "19",
        "Largest Room/Ballroom (sqm)": "N/A",
        "Maximum capacity (theater style)": "N/A",
        "Treatment rooms (#)": "N/A",
        "Spa total area (sqm)": "N/A"
    },
    "Mövenpick JLT": {
        "Last Renovation": "2011",
        "Number of Suites": "Multiple",
        "Number of Rooms": "168",
        "Number of Apartments": "0",
        "Percent Suites": "N/A",
        "Total Function Space (sqm)": "N/A",
        "Number of Ballroom": "N/A",
        "Largest Room/Ballroom (sqm)": "N/A",
        "Maximum capacity (theater style)": "N/A",
        "Spa Name": "Mövenpick Spa",
        "Treatment rooms (#)": "7",
        "Spa total area (sqm)": "N/A",
        "Kid Amenities": "Y",
        "Additional Facilities": "Hair salon, nail bar, spa with 7 treatment rooms, outdoor pool, gym"
    },
    "Grand Plaza Movenpick Media City": {
        "Last Renovation": "N/A",
        "Number of Rooms": "235",
        "Number of Apartments": "0",
        "Spa total area (sqm)": "N/A",
        "Kids Club": "N",
        "Kid Amenities": "N",
        "Swimming Pool Count": "1",
        "Additional Facilities": "Convention Centre, ballroom for 800 guests, 9 meeting rooms, M Spa, 24h gym, 5 restaurants/lounges/bars"
    },
    "Atana Hotel": {
        "TripAdvisor Score": "4.0",
        "Last Renovation": "N/A",
        "Brand": "Independent",
        "Owning Company/Management": "N/A",
        "Number of Suites": "44",
        "Number of Rooms": "828",
        "Number of Apartments": "0",
        "Percent Suites": "5.3%",
        "Total Function Space (sqm)": "N/A",
        "Largest Room/Ballroom (sqm)": "N/A",
        "Treatment rooms (#)": "N/A",
        "Spa total area (sqm)": "N/A",
        "Kid Amenities": "N",
        "Additional Facilities": "Grand ballroom, conference facilities, fitness center, outdoor pool, 3 dining outlets"
    },
    "Taj JLT": {
        "Booking.com Score": "9.0",
        "TripAdvisor Score": "4.0",
        "Last Renovation": "N/A",
        "Number of Rooms": "200",
        "Number of Apartments": "0",
        "Spa total area (sqm)": "N/A (No Spa)",
        "Kids Club": "N",
        "Kid Amenities": "N",
        "Additional Facilities": "4 restaurants, rooftop bar, fitness center, pool, free breakfast, free parking"
    }
}

# Fill in the data
updates_made = 0
for hotel_name, updates in hotel_updates.items():
    if hotel_name in hotel_to_col:
        hotel_col = hotel_to_col[hotel_name]
        print(f"\nUpdating {hotel_name} (Column {hotel_col}):")

        for field_name, value in updates.items():
            if field_name in field_to_row:
                field_row = field_to_row[field_name]
                cell = ws.cell(row=field_row, column=hotel_col)

                # Only update if cell is empty
                if cell.value is None or str(cell.value).strip() == '':
                    cell.value = value
                    updates_made += 1
                    print(f"  [+] {field_name}: {value}")
                else:
                    print(f"  [-] {field_name}: Already has value '{cell.value}'")
            else:
                print(f"  [X] Field '{field_name}' not found in spreadsheet")
    else:
        print(f"\n[X] Hotel '{hotel_name}' not found in spreadsheet")

print(f"\n\nTotal updates made: {updates_made}")
print(f"Saving to: {output_file}")

# Save the workbook
wb.save(output_file)
print("[OK] File saved successfully!")
