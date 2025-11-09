import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from copy import copy

# Load both files
updated_file = r"C:\Users\reservations\Desktop\Compset Tool\Compset_Analysis_With_ML_Features_Updated.xlsx"
reference_file = r"C:\Users\reservations\Desktop\Compset Tool\Compset_Analysis_With_ML_Features_Final_Formatted - Copy - Copy.xlsx"
output_file = r"C:\Users\reservations\Desktop\Compset Tool\Compset_Analysis_With_ML_Features_Final.xlsx"

# Load workbooks
wb_updated = openpyxl.load_workbook(updated_file)
wb_reference = openpyxl.load_workbook(reference_file)

ws_updated = wb_updated.active
ws_reference = wb_reference.active

print("Analyzing reference file formatting...")

# Copy formatting from reference to updated file
# We'll copy column widths, row heights, fonts, fills, borders, and alignment

# Copy column widths
for col_letter in ws_reference.column_dimensions:
    if col_letter in ws_updated.column_dimensions:
        ws_updated.column_dimensions[col_letter].width = ws_reference.column_dimensions[col_letter].width

# Copy row heights
for row_num in ws_reference.row_dimensions:
    if row_num in ws_updated.row_dimensions:
        ws_updated.row_dimensions[row_num].height = ws_reference.row_dimensions[row_num].height

print("Copying cell formatting...")

# Copy cell formatting for all cells in the reference file
for row_idx, row in enumerate(ws_reference.iter_rows(), start=1):
    for col_idx, ref_cell in enumerate(row, start=1):
        # Get corresponding cell in updated file
        upd_cell = ws_updated.cell(row=row_idx, column=col_idx)

        # Copy font
        if ref_cell.font:
            upd_cell.font = copy(ref_cell.font)

        # Copy fill
        if ref_cell.fill:
            upd_cell.fill = copy(ref_cell.fill)

        # Copy alignment
        if ref_cell.alignment:
            upd_cell.alignment = copy(ref_cell.alignment)

        # Copy border
        if ref_cell.border:
            upd_cell.border = copy(ref_cell.border)

        # Copy number format
        if ref_cell.number_format:
            upd_cell.number_format = ref_cell.number_format

    if row_idx % 10 == 0:
        print(f"  Processed {row_idx} rows...")

# Also apply formatting to the new hotels (columns 16-19) if they exist
# We'll use the formatting from column 15 (last hotel in reference) as a template
if ws_updated.max_column > 15:
    print("\nApplying formatting to additional hotel columns...")
    template_col = 15  # Use Arjaan by Rotana as template

    for new_col in range(16, ws_updated.max_column + 1):
        for row_idx in range(1, ws_updated.max_row + 1):
            template_cell = ws_reference.cell(row=row_idx, column=template_col)
            new_cell = ws_updated.cell(row=row_idx, column=new_col)

            # Copy formatting
            if template_cell.font:
                new_cell.font = copy(template_cell.font)
            if template_cell.fill:
                new_cell.fill = copy(template_cell.fill)
            if template_cell.alignment:
                new_cell.alignment = copy(template_cell.alignment)
            if template_cell.border:
                new_cell.border = copy(template_cell.border)
            if template_cell.number_format:
                new_cell.number_format = template_cell.number_format

# Set some nice default formatting for better readability
print("\nApplying additional formatting...")

# Header row (row 1) - bold and centered
for cell in ws_updated[1]:
    if cell.value:
        cell.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Section headers (rows with text in column A like "ROOMS INFORMATION", "FOOD & BEVERAGE", etc.)
section_headers = ['ROOMS INFORMATION', 'FOOD & BEVERAGE', 'MEETINGS & EVENTS', 'FITNESS & SPA', 'OTHER FACILITIES']
for row in ws_updated.iter_rows(min_row=1, max_row=ws_updated.max_row):
    if row[0].value in section_headers:
        for cell in row:
            cell.font = Font(name='Calibri', size=10, bold=True)
            cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
            cell.alignment = Alignment(horizontal='left', vertical='center')

# First two columns - labels
for row_idx in range(1, ws_updated.max_row + 1):
    for col_idx in [1, 2]:
        cell = ws_updated.cell(row=row_idx, column=col_idx)
        if cell.value and cell.value not in ['GENERAL INFORMATION', 'NORMALIZATION/CATEGORIZATION']:
            cell.font = Font(name='Calibri', size=10)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

# Data cells - center align
for row_idx in range(2, ws_updated.max_row + 1):
    for col_idx in range(3, ws_updated.max_column + 1):
        cell = ws_updated.cell(row=row_idx, column=col_idx)
        if cell.value:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Set column widths
ws_updated.column_dimensions['A'].width = 30
ws_updated.column_dimensions['B'].width = 35

for col_idx in range(3, ws_updated.max_column + 1):
    col_letter = openpyxl.utils.get_column_letter(col_idx)
    ws_updated.column_dimensions[col_letter].width = 25

# Freeze panes - freeze first 2 columns and first row
ws_updated.freeze_panes = 'C2'

print(f"\nSaving formatted file to: {output_file}")
wb_updated.save(output_file)
print("[OK] Formatting complete!")

print(f"\nSummary:")
print(f"  - Copied formatting from reference file")
print(f"  - Applied consistent styling to all columns")
print(f"  - Set column widths for readability")
print(f"  - Froze first row and first 2 columns")
print(f"  - Total columns: {ws_updated.max_column}")
print(f"  - Total rows: {ws_updated.max_row}")
