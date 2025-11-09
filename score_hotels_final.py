import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from sklearn.preprocessing import MinMaxScaler

# Read the Excel file - Features sheet
excel_file = 'Compset_Analysis_With_ML_Features_Final 22.xlsx'
df = pd.read_excel(excel_file, sheet_name='Features', header=None)

print("Reading Features sheet...")
print(f"Shape: {df.shape}")

# First row contains hotel names (starting from column 2)
hotel_names = df.iloc[0, 2:].tolist()
print(f"\nHotels found ({len(hotel_names)}):")
for i, hotel in enumerate(hotel_names, 1):
    print(f"  {i}. {hotel}")

# Define exact row indices for each feature based on the actual Excel structure
FEATURE_ROWS = {
    'star_rating': 4,
    'booking_rating': 5,
    'tripadvisor_rating': 6,
    'distance_km': 2,
    'total_rooms': 15,  # Total Number of Keys
    'meeting_space_sqm': 33,  # Total Function Space (sqm)
    'f_b_outlets': 25,  # Restaurants Count
    'pool_count': 48,  # Swimming Pool Count
    'gym': 42,  # Health Club / Fitness (Y/N)
    'spa': 40,  # Spa Name (if exists, then Y)
}

# Create a dictionary to store hotel data
hotels_data = {}

# Initialize data structure for each hotel
for hotel in hotel_names:
    hotels_data[hotel] = {
        'star_rating': None,
        'distance_km': None,
        'total_rooms': None,
        'tripadvisor_rating': None,
        'google_rating': None,
        'booking_rating': None,
        'meeting_space_sqm': None,
        'f_b_outlets': None,
        'pool_count': None,
        'gym': None,
        'spa': None,
    }

# Extract data for each hotel
for col_idx, hotel in enumerate(hotel_names, start=2):  # Start at column 2

    # Star Rating (Row 4)
    value = df.iloc[FEATURE_ROWS['star_rating'], col_idx]
    if pd.notna(value):
        try:
            hotels_data[hotel]['star_rating'] = float(value)
        except:
            pass

    # Distance (Row 2)
    value = df.iloc[FEATURE_ROWS['distance_km'], col_idx]
    if pd.notna(value):
        try:
            hotels_data[hotel]['distance_km'] = float(value)
        except:
            pass

    # Total Rooms (Row 15 - Total Number of Keys)
    value = df.iloc[FEATURE_ROWS['total_rooms'], col_idx]
    if pd.notna(value):
        try:
            hotels_data[hotel]['total_rooms'] = float(value)
        except:
            pass

    # TripAdvisor Rating (Row 6)
    value = df.iloc[FEATURE_ROWS['tripadvisor_rating'], col_idx]
    if pd.notna(value):
        try:
            hotels_data[hotel]['tripadvisor_rating'] = float(value)
        except:
            pass

    # Booking.com Rating (Row 5)
    value = df.iloc[FEATURE_ROWS['booking_rating'], col_idx]
    if pd.notna(value):
        try:
            hotels_data[hotel]['booking_rating'] = float(value)
        except:
            pass

    # Meeting Space (Row 33)
    value = df.iloc[FEATURE_ROWS['meeting_space_sqm'], col_idx]
    if pd.notna(value):
        try:
            hotels_data[hotel]['meeting_space_sqm'] = float(value)
        except:
            pass

    # F&B Outlets (Row 25 - Restaurants Count)
    value = df.iloc[FEATURE_ROWS['f_b_outlets'], col_idx]
    if pd.notna(value):
        try:
            # Handle special cases like "6-8"
            if isinstance(value, str) and '-' in value:
                # Take the average
                parts = value.split('-')
                hotels_data[hotel]['f_b_outlets'] = (float(parts[0]) + float(parts[1])) / 2
            else:
                hotels_data[hotel]['f_b_outlets'] = float(value)
        except:
            pass

    # Pool (Row 48)
    value = df.iloc[FEATURE_ROWS['pool_count'], col_idx]
    if pd.notna(value):
        try:
            hotels_data[hotel]['pool_count'] = float(value)
        except:
            pass

    # Gym (Row 42 - Health Club / Fitness Y/N)
    value = df.iloc[FEATURE_ROWS['gym'], col_idx]
    if pd.notna(value):
        if str(value).upper() in ['Y', 'YES', 'TRUE', '1']:
            hotels_data[hotel]['gym'] = 1.0
        elif str(value).upper() in ['N', 'NO', 'FALSE', '0']:
            hotels_data[hotel]['gym'] = 0.0

    # Spa (Row 40 - Spa Name, if exists then Y)
    value = df.iloc[FEATURE_ROWS['spa'], col_idx]
    if pd.notna(value) and str(value).strip() and str(value).lower() not in ['no', 'n', 'none', '0', 'nan']:
        hotels_data[hotel]['spa'] = 1.0
    else:
        hotels_data[hotel]['spa'] = 0.0

# Convert to DataFrame
df_hotels = pd.DataFrame.from_dict(hotels_data, orient='index')
df_hotels['hotel_name'] = df_hotels.index
df_hotels = df_hotels.reset_index(drop=True)

print("\n" + "="*120)
print("EXTRACTED HOTEL DATA")
print("="*120)
print(df_hotels.to_string())

# Define scoring weights (total = 1.0)
scoring_features = {
    'star_rating': {'weight': 0.15, 'higher_better': True},
    'total_rooms': {'weight': 0.08, 'higher_better': True},
    'distance_km': {'weight': 0.20, 'higher_better': False},  # Lower distance is better (PROXIMITY BONUS)
    'tripadvisor_rating': {'weight': 0.12, 'higher_better': True},
    'booking_rating': {'weight': 0.24, 'higher_better': True},  # Combined weight for online ratings (no Google data)
    'meeting_space_sqm': {'weight': 0.08, 'higher_better': True},
    'f_b_outlets': {'weight': 0.05, 'higher_better': True},
    'pool_count': {'weight': 0.04, 'higher_better': True},
    'gym': {'weight': 0.02, 'higher_better': True},
    'spa': {'weight': 0.02, 'higher_better': True},
}

# Brand penalties (applied AFTER raw scoring)
brand_penalties = {
    'atana': -0.15,  # 15% penalty
    'two seasons': -0.15,  # 15% penalty
}

# Initialize score columns
df_hotels['raw_score'] = 0.0
df_hotels['brand_penalty'] = 0.0
df_hotels['final_score'] = 0.0
df_hotels['normalized_score'] = 0.0

# Compute raw scores for each hotel
for idx, row in df_hotels.iterrows():
    raw_score = 0.0

    for feature, config in scoring_features.items():
        value = row[feature]

        # Skip if value is NaN
        if pd.isna(value):
            continue

        # Get all non-NaN values for this feature
        all_values = df_hotels[feature].dropna()

        if len(all_values) == 0 or len(all_values) == 1:
            continue

        # Normalize to 0-1 scale using min-max
        min_val = all_values.min()
        max_val = all_values.max()

        if max_val - min_val == 0:
            normalized_value = 1.0
        else:
            normalized_value = (value - min_val) / (max_val - min_val)

        # Invert if lower is better (e.g., distance - closer is better)
        if not config['higher_better']:
            normalized_value = 1 - normalized_value

        # Apply weight
        weighted_score = normalized_value * config['weight']
        raw_score += weighted_score

    df_hotels.at[idx, 'raw_score'] = raw_score

# Apply brand penalties
print("\n" + "="*120)
print("APPLYING BRAND PENALTIES")
print("="*120)
for idx, row in df_hotels.iterrows():
    hotel_name = str(row['hotel_name']).lower()
    penalty = 0.0

    for brand, brand_penalty in brand_penalties.items():
        if brand in hotel_name:
            penalty = brand_penalty
            print(f"  > {row['hotel_name']}: {brand_penalty*100}% penalty applied")
            break

    df_hotels.at[idx, 'brand_penalty'] = penalty
    df_hotels.at[idx, 'final_score'] = row['raw_score'] + penalty

# Normalize final scores to 0-1 scale
final_scores = df_hotels['final_score'].values.reshape(-1, 1)
scaler = MinMaxScaler()
normalized_scores = scaler.fit_transform(final_scores)
df_hotels['normalized_score'] = normalized_scores.flatten()

# Sort by normalized score (descending)
df_hotels_sorted = df_hotels.sort_values('normalized_score', ascending=False)

# Display results
print("\n" + "="*120)
print("COMPETITIVE SCORING RESULTS (Ranked)")
print("="*120)
print(df_hotels_sorted[['hotel_name', 'raw_score', 'brand_penalty', 'final_score', 'normalized_score']].to_string())

# Write scores to Excel - REPLACE the existing scores section
wb = load_workbook(excel_file)
ws = wb['Features']

# Find where COMPETITIVE SCORES section starts (should be around row 55)
score_start_row = None
for row in range(50, 80):
    cell_value = ws.cell(row=row, column=1).value
    if cell_value and 'COMPETITIVE SCORES' in str(cell_value):
        score_start_row = row
        break

# If not found, add it after the existing data
if score_start_row is None:
    # Find the last row with data in column 1
    last_row = ws.max_row
    score_start_row = last_row + 3

# Clear existing score rows (if they exist)
for row in range(score_start_row, score_start_row + 10):
    for col in range(1, 20):
        ws.cell(row=row, column=col).value = None

# Section header
ws.cell(row=score_start_row, column=1, value="COMPETITIVE SCORES")
ws.cell(row=score_start_row, column=1).font = Font(bold=True, size=12, color="FFFFFF")
ws.cell(row=score_start_row, column=1).fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")

ws.cell(row=score_start_row, column=2, value="Normalized 0-1 Scale")
ws.cell(row=score_start_row, column=2).font = Font(italic=True, size=10)

# Score labels
score_labels = [
    ('Raw Score', score_start_row + 2),
    ('Brand Penalty', score_start_row + 3),
    ('Final Score', score_start_row + 4),
    ('Normalized Score (0-1)', score_start_row + 5),
]

for label, row_num in score_labels:
    ws.cell(row=row_num, column=1, value=label)
    ws.cell(row=row_num, column=1).font = Font(bold=True)
    ws.cell(row=row_num, column=1).fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

# Write scores for each hotel
for col_idx, hotel_name in enumerate(hotel_names, start=2):  # Start at column 2
    hotel_data = df_hotels[df_hotels['hotel_name'] == hotel_name]

    if not hotel_data.empty:
        hotel_info = hotel_data.iloc[0]

        # Write scores
        ws.cell(row=score_labels[0][1], column=col_idx, value=hotel_info['raw_score'])
        ws.cell(row=score_labels[1][1], column=col_idx, value=hotel_info['brand_penalty'])
        ws.cell(row=score_labels[2][1], column=col_idx, value=hotel_info['final_score'])
        ws.cell(row=score_labels[3][1], column=col_idx, value=hotel_info['normalized_score'])

        # Format cells
        for _, row_num in score_labels:
            cell = ws.cell(row=row_num, column=col_idx)
            cell.number_format = '0.0000'
            cell.alignment = Alignment(horizontal='center')

            # Highlight normalized score
            if row_num == score_labels[3][1]:
                cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

# Save workbook
wb.save(excel_file)
print(f"\n[SUCCESS] Scores successfully updated in '{excel_file}' in 'Features' sheet")
print(f"  Starting at row {score_start_row}")

wb.close()

# Print methodology
print("\n" + "="*120)
print("SCORING METHODOLOGY")
print("="*120)
print("\nFeature Weights (Total = 100%):")
for feature, config in scoring_features.items():
    direction = "higher better" if config['higher_better'] else "CLOSER BETTER (proximity bonus)"
    print(f"  - {feature:25s}: {config['weight']*100:5.1f}%  {direction}")

print("\nBrand Penalties (applied after raw scoring):")
for brand, penalty in brand_penalties.items():
    print(f"  - {brand:25s}: {penalty*100:5.1f}% penalty")

print("\nScoring Process:")
print("  1. Min-max normalization (0-1) for each feature across all hotels")
print("  2. Distance scoring INVERTED (closer = higher score)")
print("  3. Weighted sum of normalized features = Raw Score")
print("  4. Brand penalties applied to Atana and Two Seasons")
print("  5. Final scores min-max normalized to 0-1 scale")
print("  6. Higher score = more competitive/similar to Grand Millennium Dubai")
print("="*120)
