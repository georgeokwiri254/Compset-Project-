#!/usr/bin/env python3
"""
Find all existing hotels in the workbook and identify gaps
"""

import pandas as pd
from openpyxl import load_workbook
import json

excel_path = '/home/gee_devops254/Downloads/Compset Tool/NEW - STR Competitor Set Analysis.xlsx'
wb = load_workbook(excel_path, data_only=True)

# Check Business Mix sheet
print("="*80)
print("HOTELS IN BUSINESS MIX & OVERLAP ANALYSIS SHEET")
print("="*80)
ws = wb['Business Mix & Overalp Analysis']

# Find all hotel names (column A) after row 12
hotels_present = []
for i in range(12, ws.max_row + 1):
    hotel_name = ws.cell(row=i, column=1).value
    if hotel_name and hotel_name not in ['', None] and not hotel_name.startswith('Step'):
        hotels_present.append(hotel_name)

print(f"\nFound {len(hotels_present)} hotels:")
for idx, hotel in enumerate(hotels_present, 1):
    print(f"{idx}. {hotel}")

# Count how many more needed
target_count = 20  # Grand Millennium + 10 from sheet + 10 more
hotels_needed = max(0, target_count - len(hotels_present))
print(f"\n✓ Current count: {len(hotels_present)}")
print(f"✓ Target: {target_count} total hotels")
print(f"✓ Additional hotels needed: {hotels_needed}")

# Load JSON data to find suitable candidates
print("\n" + "="*80)
print("FINDING ADDITIONAL HOTEL CANDIDATES")
print("="*80)

with open('/home/gee_devops254/Downloads/Compset Tool/filtered_hotels.json', 'r') as f:
    filtered_hotels = json.load(f)

# Get hotels not already in the list
available_hotels = []
for hotel in filtered_hotels:
    if hotel['name'] not in hotels_present:
        available_hotels.append(hotel)

# Sort by distance and star rating
available_hotels.sort(key=lambda x: (x['distance_km'], -x['star_rating']))

print(f"\nTop 15 candidate hotels (by proximity and star rating):")
for idx, hotel in enumerate(available_hotels[:15], 1):
    print(f"{idx}. {hotel['name']}")
    print(f"   - Star Rating: {hotel['star_rating']}, Distance: {hotel['distance_km']} km")
    print(f"   - Category: {hotel['category']}, Area: {hotel['area']}")
    print()

print("="*80)
