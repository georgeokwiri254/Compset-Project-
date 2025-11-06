from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

# Load the Excel file
excel_file = r'C:\Users\reservations\Desktop\Compset Tool\Grand_Millennium_Dubai_CompSet_Analysis.xlsx'
wb = load_workbook(excel_file)

# Create or get the Executive Summary sheet
if 'Executive Summary Report' in wb.sheetnames:
    del wb['Executive Summary Report']

summary_sheet = wb.create_sheet('Executive Summary Report', 0)

# Title and Date
summary_sheet['A1'] = 'GRAND MILLENNIUM DUBAI'
summary_sheet['A1'].font = Font(bold=True, size=18, color='FFFFFF')
summary_sheet['A1'].fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
summary_sheet.merge_cells('A1:H1')
summary_sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')

summary_sheet['A2'] = 'Competitive Set Analysis Report'
summary_sheet['A2'].font = Font(bold=True, size=14, color='FFFFFF')
summary_sheet['A2'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
summary_sheet.merge_cells('A2:H2')
summary_sheet['A2'].alignment = Alignment(horizontal='center', vertical='center')

summary_sheet['A3'] = f'Generated: {datetime.now().strftime("%B %d, %Y")}'
summary_sheet.merge_cells('A3:H3')
summary_sheet['A3'].alignment = Alignment(horizontal='center')

# Executive Summary Section
summary_sheet['A5'] = 'EXECUTIVE SUMMARY'
summary_sheet['A5'].font = Font(bold=True, size=14, color='1F4E78')
summary_sheet.merge_cells('A5:H5')

summary_sheet['A7'] = 'Objective:'
summary_sheet['A7'].font = Font(bold=True)
summary_sheet['B7'] = 'Identify optimal competitive set for Grand Millennium Dubai based on proximity, facilities, room mix, business mix, star rating, and market positioning.'
summary_sheet.merge_cells('B7:H7')
summary_sheet['B7'].alignment = Alignment(wrap_text=True)

summary_sheet['A9'] = 'Analysis Scope:'
summary_sheet['A9'].font = Font(bold=True)
summary_sheet['B9'] = 'Non-luxury hotels within 5 km radius'
summary_sheet['B10'] = '56 hotels identified and filtered'
summary_sheet['B11'] = '8 hotels researched in detail with complete data'
summary_sheet['B12'] = 'Scoring based on 15+ criteria weighted by importance'

# Key Findings
summary_sheet['A15'] = 'KEY FINDINGS'
summary_sheet['A15'].font = Font(bold=True, size=14, color='1F4E78')
summary_sheet.merge_cells('A15:H15')

summary_sheet['A17'] = '1. PRIMARY COMPSET RECOMMENDATION (Score ≥ 90):'
summary_sheet['A17'].font = Font(bold=True, size=11)
summary_sheet['A17'].fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
summary_sheet.merge_cells('A17:H17')

summary_sheet['B18'] = '• Millennium Place Barsha Heights Hotel'
summary_sheet['C18'] = 'Score: 91.0'
summary_sheet['D18'] = 'Distance: 1.0 km'
summary_sheet['E18'] = 'Rooms: 468 + 447 apartments'
summary_sheet['F18'] = 'Meeting Space: 124 sqm'
summary_sheet['G18'] = 'Rating: 8.4/10 (Booking.com)'

summary_sheet['B19'] = 'Rationale:'
summary_sheet['B19'].font = Font(bold=True, size=9)
summary_sheet['C19'] = 'Same brand family, similar room mix with apartments, strong business & MICE capability, excellent location, comparable facilities'
summary_sheet.merge_cells('C19:H19')
summary_sheet['C19'].alignment = Alignment(wrap_text=True)

summary_sheet['A21'] = '2. SECONDARY COMPSET RECOMMENDATIONS (Score 75-89):'
summary_sheet['A21'].font = Font(bold=True, size=11)
summary_sheet['A21'].fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
summary_sheet.merge_cells('A21:H21')

secondary_hotels = [
    ('Radisson Blu Hotel Dubai Media City', 88.0, 2.0, 246, 582, 8.3),
    ('Golden Tulip Media Hotel', 83.0, 0.18, 288, 186, 8.0),
    ('Mercure Dubai Barsha Heights', 81.0, 0.6, 1015, 100, 8.2),
    ('Staybridge Suites Dubai Internet City', 79.0, 1.8, 225, 80, 8.5),
    ('Atana Hotel', 77.0, 0.14, 828, 500, 8.4),
    ('Millennium Place Dubai Marina', 76.0, 4.8, 458, 50, 8.0)
]

row = 22
for hotel, score, dist, rooms, meeting, rating in secondary_hotels:
    summary_sheet[f'B{row}'] = f'• {hotel}'
    summary_sheet[f'C{row}'] = f'{score:.1f}'
    summary_sheet[f'D{row}'] = f'{dist} km'
    summary_sheet[f'E{row}'] = f'{rooms} rooms'
    summary_sheet[f'F{row}'] = f'{meeting} sqm'
    summary_sheet[f'G{row}'] = f'{rating}/10'
    row += 1

# Competitive Analysis
summary_sheet['A30'] = 'COMPETITIVE POSITIONING ANALYSIS'
summary_sheet['A30'].font = Font(bold=True, size=14, color='1F4E78')
summary_sheet.merge_cells('A30:H30')

summary_sheet['A32'] = 'Market Segmentation:'
summary_sheet['A32'].font = Font(bold=True)
summary_sheet['B32'] = '• Business/Corporate: 40-70% across compset'
summary_sheet['B33'] = '• Leisure: 20-50% across compset'
summary_sheet['B34'] = '• MICE: 10-20% across compset'

summary_sheet['A36'] = 'Distance Distribution:'
summary_sheet['A36'].font = Font(bold=True)
summary_sheet['B36'] = '• Immediate Zone (0-2 km): 22 hotels - High competition'
summary_sheet['B37'] = '• Near Zone (2-4 km): 15 hotels - Moderate competition'
summary_sheet['B38'] = '• Extended Zone (4-5 km): 19 hotels - Secondary competition'

summary_sheet['A40'] = 'Facility Benchmarks:'
summary_sheet['A40'].font = Font(bold=True)
summary_sheet['B40'] = '• Average Rooms: 400-500 for primary compset'
summary_sheet['B41'] = '• Meeting Space: 100-600 sqm range'
summary_sheet['B42'] = '• All compset hotels have: Pool, Gym, Multiple F&B outlets'
summary_sheet['B43'] = '• 75% have Spa facilities'
summary_sheet['B44'] = '• 50% offer serviced apartments'

# Strategic Recommendations
summary_sheet['A47'] = 'STRATEGIC RECOMMENDATIONS'
summary_sheet['A47'].font = Font(bold=True, size=14, color='1F4E78')
summary_sheet.merge_cells('A47:H47')

recommendations = [
    ('1.', 'PRIMARY COMPSET', 'Include Millennium Place Barsha Heights as core competitor for direct benchmarking'),
    ('2.', 'SECONDARY COMPSET', 'Monitor Radisson Blu Media City, Golden Tulip, Mercure Barsha Heights, and Staybridge Suites'),
    ('3.', 'RATE STRATEGY', 'Position ADR within 10-15% of primary compset average based on seasonal demand'),
    ('4.', 'MEETING & EVENTS', 'Leverage meeting space advantage (if larger) or enhance technology offerings'),
    ('5.', 'UNIQUE DIFFERENTIATORS', 'Emphasize Grand Millennium brand heritage, apartment offerings, and executive lounge'),
    ('6.', 'DISTRIBUTION', 'Ensure strong presence on OTAs where compset hotels have high visibility'),
    ('7.', 'GUEST EXPERIENCE', 'Target review scores of 8.5+ to compete with top-rated properties like Staybridge Suites'),
    ('8.', 'BUSINESS DEVELOPMENT', 'Target corporate accounts similar to Media One and Radisson Blu in Media/Internet City'),
]

row = 49
for num, title, desc in recommendations:
    summary_sheet[f'A{row}'] = num
    summary_sheet[f'A{row}'].font = Font(bold=True)
    summary_sheet[f'B{row}'] = title
    summary_sheet[f'B{row}'].font = Font(bold=True)
    summary_sheet[f'C{row}'] = desc
    summary_sheet.merge_cells(f'C{row}:H{row}')
    summary_sheet[f'C{row}'].alignment = Alignment(wrap_text=True)
    row += 1

# Next Steps
summary_sheet['A59'] = 'NEXT STEPS'
summary_sheet['A59'].font = Font(bold=True, size=14, color='1F4E78')
summary_sheet.merge_cells('A59:H59')

next_steps = [
    '1. Complete detailed research on remaining 48 hotels to expand analysis',
    '2. Gather April 2025 rate data for ADR comparison across compset',
    '3. Conduct detailed MICE capability assessment with site visits',
    '4. Validate corporate account overlap with sales team',
    '5. Set up monthly compset monitoring using STR reports',
    '6. Review and adjust compset quarterly based on market changes'
]

row = 61
for step in next_steps:
    summary_sheet[f'B{row}'] = step
    summary_sheet.merge_cells(f'B{row}:H{row}')
    summary_sheet[f'B{row}'].alignment = Alignment(wrap_text=True)
    row += 1

# Footer
summary_sheet['A69'] = 'Report prepared using comprehensive market research, online reviews, and competitive intelligence'
summary_sheet.merge_cells('A69:H69')
summary_sheet['A69'].alignment = Alignment(horizontal='center')
summary_sheet['A69'].font = Font(italic=True, size=9)

# Adjust column widths
summary_sheet.column_dimensions['A'].width = 12
summary_sheet.column_dimensions['B'].width = 35
summary_sheet.column_dimensions['C'].width = 15
summary_sheet.column_dimensions['D'].width = 15
summary_sheet.column_dimensions['E'].width = 20
summary_sheet.column_dimensions['F'].width = 18
summary_sheet.column_dimensions['G'].width = 18
summary_sheet.column_dimensions['H'].width = 5

# Set row heights for better readability
for row in range(1, 70):
    summary_sheet.row_dimensions[row].height = 20

summary_sheet.row_dimensions[1].height = 30
summary_sheet.row_dimensions[2].height = 25

# Save the workbook
wb.save(excel_file)

print(f"\nExecutive Summary Report created successfully!")
print(f"Sheet 'Executive Summary Report' added to: {excel_file}")
print("\n" + "="*80)
print("SUMMARY OF ANALYSIS")
print("="*80)
print("\nPrimary Compset Recommendation:")
print("• Millennium Place Barsha Heights Hotel (Score: 91.0)")
print("\nTop 6 Secondary Compset Hotels:")
for hotel, score, dist, rooms, meeting, rating in secondary_hotels:
    print(f"• {hotel} (Score: {score:.1f}, Distance: {dist} km)")
