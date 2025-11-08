import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Create a new workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Room Types Comparison"

# Define styles
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
hotel_header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
hotel_header_font = Font(bold=True, color="FFFFFF", size=12)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
wrap_alignment = Alignment(wrap_text=True, vertical='top')

# Hotels data
hotels = {
    "Grand Millennium Dubai": {
        "Total Rooms": "339",
        "Room Types": [
            {"Name": "Standard Room", "Size": "N/A", "Features": "Air conditioning, minibar, private bathroom, bathrobe, slippers, toiletries"},
            {"Name": "Twin Room", "Size": "N/A", "Features": "2 beds, flat-screen TV, cable channels, soundproof walls, minibar, tea/coffee maker, sea views"},
            {"Name": "Deluxe Room", "Size": "N/A", "Features": "Views of Jumeirah coastline"},
            {"Name": "Suite", "Size": "N/A", "Features": "Large living room, panoramic city views, fully equipped kitchen"},
            {"Name": "Family Suite", "Size": "N/A", "Features": "King bed room + twin bed room, open kitchen, dining table, seating area"},
            {"Name": "Apartment", "Size": "N/A", "Features": "Separate living room, dining room, fully equipped kitchen"},
            {"Name": "Club Room", "Size": "N/A", "Features": "Club Lounge access, complimentary cocktails, one-way airport limousine"}
        ],
        "Key Amenities": "Free WiFi, rooftop pool, spa, sauna, steam room, Jacuzzi, health club, 24-hour room service, kitchenette, Netflix, streaming services",
        "Hotel Type": "Mixed (Rooms + Suites + Apartments)"
    },
    "Zabeel House The Greens": {
        "Total Rooms": "210",
        "Room Types": [
            {"Name": "Popular Room", "Size": "28 sqm", "Features": "Contemporary design, views over Dubai neighbourhood, organic toiletries, luxury bedding"},
            {"Name": "Popular Balcony Room", "Size": "28 sqm", "Features": "Balcony, New York Loft style, modern Middle Eastern flair"},
            {"Name": "Studio Room", "Size": "28 sqm", "Features": "Loft-inspired, bedroom + separate lounge"},
            {"Name": "The Apartment (Suite)", "Size": "28 sqm", "Features": "Cool design combining comfort, vibrant space"},
            {"Name": "Connected Rooms", "Size": "56 sqm", "Features": "Two connected Popular rooms for families"}
        ],
        "Key Amenities": "Organic toiletries, luxury bedding, minibar, air conditioning, flat screen TV, writing table, multi-channel TV, mini fridge, coffee/tea facilities",
        "Hotel Type": "Mixed (Rooms + Studios + Apartments)"
    },
    "voco Bonnington Dubai": {
        "Total Rooms": "208",
        "Room Types": [
            {"Name": "Classic Guest Room", "Size": "N/A", "Features": "Blackout blinds, rainfall shower, stand-alone bathtub, luxury amenities"},
            {"Name": "Executive Room", "Size": "N/A", "Features": "Premium features with enhanced amenities"}
        ],
        "Key Amenities": "Flat-screen smart TV, rain showers, 100% recycled bedding, Antipodes organic amenities, high-speed WiFi, executive desks, 24-hour room service",
        "Hotel Type": "Rooms"
    },
    "Radisson Blu Media City": {
        "Total Rooms": "246",
        "Room Types": [
            {"Name": "Standard Room", "Size": "N/A", "Features": "Bright, spacious, flat-screen TV, contemporary furnishings, free WiFi, tea/coffee maker"},
            {"Name": "Executive Room", "Size": "N/A", "Features": "Mini-bar, individual climate control, Executive Lounge access, complimentary breakfast"},
            {"Name": "One Bedroom Suite with Lounge Access", "Size": "N/A", "Features": "Spacious living area separated from bedroom, turndown service, lounge access"},
            {"Name": "Connected Rooms", "Size": "N/A", "Features": "Spacious surroundings, baby-sitting service for kids"}
        ],
        "Key Amenities": "Free WiFi, 24-hour room service, laptop-compatible safes, laptop-friendly workspaces, air conditioning, minibar",
        "Hotel Type": "Mixed (Rooms + Suites)"
    },
    "Media One Hotel": {
        "Total Rooms": "260-264",
        "Room Types": [
            {"Name": "Hip Room (Standard)", "Size": "30 sqm", "Features": "Queen bed, working desk, sleeps 2 adults"},
            {"Name": "Hip Urban/City Deluxe", "Size": "30 sqm", "Features": "Queen bed, city or golf views"},
            {"Name": "Business Room", "Size": "30 sqm", "Features": "Queen bed, Palm views"},
            {"Name": "Executive Room", "Size": "35 sqm", "Features": "King bed, city views"},
            {"Name": "Studio", "Size": "50 sqm", "Features": "King bed, Palm views, separate seating area"},
            {"Name": "Suite", "Size": "70 sqm", "Features": "King bed, Palm & Dubai skyline views, separate living room"}
        ],
        "Key Amenities": "Floor-to-ceiling windows, air-conditioning, minibar, docking stations, 43-inch smart TV, glass-walled bathrooms, WiFi, safe, 24-hour dining",
        "Hotel Type": "Mixed (Rooms + Studios + Suites)"
    },
    "Pullman JLT": {
        "Total Rooms": "354 (including 76 suites/apartments)",
        "Room Types": [
            {"Name": "Superior Room", "Size": "35 sqm", "Features": "King or twin bed, LED HD TV, WiFi, minibar, safe, tea/coffee facilities"},
            {"Name": "Deluxe Room", "Size": "40 sqm", "Features": "Enhanced space, walk-in shower and tub, modern bathroom"},
            {"Name": "Executive Room", "Size": "35 sqm", "Features": "Nespresso machine, Bose docking station, Executive Lounge access"},
            {"Name": "Deluxe Executive Room", "Size": "40 sqm", "Features": "Larger space + executive benefits"},
            {"Name": "Studio Suite", "Size": "N/A", "Features": "Equipped kitchen, living/dining area"},
            {"Name": "1-Bedroom Suite", "Size": "N/A", "Features": "Attached bathroom, powder room, living/dining, fitted kitchen, wardrobes, marble floors"},
            {"Name": "2-Bedroom Suite", "Size": "N/A", "Features": "Two attached + common bathroom, modern kitchen, wardrobes, marble floors, balcony"},
            {"Name": "3-Bedroom Suite", "Size": "N/A", "Features": "Three attached + guest bathroom, marble floors, fitted kitchen, living/dining, balcony"}
        ],
        "Key Amenities": "Free WiFi, LED HD TV, tea/coffee facilities, minibar, safe, iron/board, walk-in shower, Nespresso (Executive+), views of Lake/Marina/Golf",
        "Hotel Type": "Mixed (Rooms + Suites + Apartments)"
    },
    "Millennium Place Barsha Heights": {
        "Total Rooms": "468 rooms + apartments",
        "Room Types": [
            {"Name": "Superior Room", "Size": "N/A", "Features": "High-speed WiFi, USB charging, 43-inch IPTV system"},
            {"Name": "Deluxe Room", "Size": "N/A", "Features": "Enhanced amenities, modern facilities"},
            {"Name": "Deluxe Sky Room", "Size": "N/A", "Features": "Higher floor, enhanced views"},
            {"Name": "Premium Room", "Size": "N/A", "Features": "Premium amenities for business travelers"},
            {"Name": "Premium Sky Room", "Size": "N/A", "Features": "Higher floor premium features"},
            {"Name": "Deluxe Suite", "Size": "N/A", "Features": "Separate living space"},
            {"Name": "Two-Bedroom Suite", "Size": "N/A", "Features": "Family accommodation"},
            {"Name": "Studio Apartment", "Size": "N/A", "Features": "Kitchenette, refrigerator, stovetop, microwave, 43-inch LED TV"},
            {"Name": "One-Bedroom Apartment", "Size": "N/A", "Features": "Full kitchen facilities, separate bedroom"},
            {"Name": "Two-Bedroom Apartment", "Size": "N/A", "Features": "Multiple bedrooms, full kitchen"},
            {"Name": "Royal Penthouse", "Size": "N/A", "Features": "Luxury top-tier accommodation"}
        ],
        "Key Amenities": "Free WiFi, 43-inch IPTV, USB charging, laptop-friendly workspaces, air conditioning, safes, kitchenettes in apartments",
        "Hotel Type": "Mixed (Hotel Rooms + Apartments)"
    },
    "TRYP by Wyndham Dubai": {
        "Total Rooms": "650",
        "Room Types": [
            {"Name": "TRYP Room", "Size": "N/A", "Features": "Chic interiors, free WiFi, streaming facilities, tea/coffee, work area, some with balconies"},
            {"Name": "TRYP King Room", "Size": "N/A", "Features": "King bed, WiFi, streaming, safe, minibar, tea/coffee, work area, bathtub + shower"},
            {"Name": "Triple Room", "Size": "N/A", "Features": "Queen + extra bed, streaming, safe, minibar, tea/coffee, work area"},
            {"Name": "TRYP One Bedroom Suite", "Size": "441 sqft", "Features": "Accommodates 4 pax (2 adults + 2 children) or 3 adults, separate living area, sofa bed, king bed"},
            {"Name": "TRYP Premium Suite", "Size": "N/A", "Features": "Loft-type suite, king bed, convertible sofa bed, streaming facilities, spacious balcony"},
            {"Name": "Accessible Room", "Size": "20 sqm", "Features": "Single beds, streaming, safe, minibar, tea/coffee, work area, accessible bathroom"}
        ],
        "Key Amenities": "Free WiFi, flat-screen TV, desk, minibar, safe, coffee maker, hair dryer, streaming facilities, work areas",
        "Hotel Type": "Mixed (Rooms + Suites)"
    },
    "Mercure Dubai Barsha Heights": {
        "Total Rooms": "408 suites",
        "Room Types": [
            {"Name": "1-Bedroom Deluxe Suite", "Size": "60 sqm", "Features": "Walk-in shower, smart TVs in bedroom & living room, coffee/tea, private balcony, views of Sheikh Zayed Road & Burj Al Arab or Dubai Marina with sea & Palm views"},
            {"Name": "2-Bedroom Prestige Suite", "Size": "105 sqm", "Features": "Lounge area, private balconies, premium amenities, walk-in shower, suitable for families/groups"}
        ],
        "Key Amenities": "Smart TVs, walk-in showers, bathtubs, balconies with stunning views, coffee/tea facilities, dedicated living areas, suitable for long & short stays",
        "Hotel Type": "All Suites & Apartments"
    }
}

# Column mapping for each hotel
col_start = 1
row_current = 1

# Create header row
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=1)
ws.cell(row=1, column=1).value = "ROOM TYPE COMPARISON"
ws.cell(row=1, column=1).font = Font(bold=True, size=14, color="FFFFFF")
ws.cell(row=1, column=1).fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
ws.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')

row_current = 3

# Row labels
row_labels = [
    "Hotel Name",
    "Total Number of Rooms",
    "Hotel Type",
    "",
    "ROOM TYPES",
    "Room Type 1",
    "Size",
    "Key Features",
    "",
    "Room Type 2",
    "Size",
    "Key Features",
    "",
    "Room Type 3",
    "Size",
    "Key Features",
    "",
    "Room Type 4",
    "Size",
    "Key Features",
    "",
    "Room Type 5",
    "Size",
    "Key Features",
    "",
    "Room Type 6",
    "Size",
    "Key Features",
    "",
    "Room Type 7",
    "Size",
    "Key Features",
    "",
    "Room Type 8",
    "Size",
    "Key Features",
    "",
    "Room Type 9",
    "Size",
    "Key Features",
    "",
    "Room Type 10",
    "Size",
    "Key Features",
    "",
    "Room Type 11",
    "Size",
    "Key Features",
    "",
    "KEY AMENITIES & FACILITIES"
]

# Write row labels
for idx, label in enumerate(row_labels):
    cell = ws.cell(row=row_current + idx, column=1)
    cell.value = label
    cell.alignment = wrap_alignment
    cell.border = border
    if label in ["Hotel Name", "ROOM TYPES", "KEY AMENITIES & FACILITIES"]:
        cell.fill = header_fill
        cell.font = header_font
    elif label.startswith("Room Type"):
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        cell.font = Font(bold=True)

# Set column A width
ws.column_dimensions['A'].width = 25

# Write hotel data
col_offset = 2
for hotel_name, hotel_data in hotels.items():
    col = col_offset

    # Hotel name
    cell = ws.cell(row=row_current, column=col)
    cell.value = hotel_name
    cell.fill = hotel_header_fill
    cell.font = hotel_header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border

    # Total rooms
    cell = ws.cell(row=row_current + 1, column=col)
    cell.value = hotel_data["Total Rooms"]
    cell.alignment = wrap_alignment
    cell.border = border

    # Hotel type
    cell = ws.cell(row=row_current + 2, column=col)
    cell.value = hotel_data["Hotel Type"]
    cell.alignment = wrap_alignment
    cell.border = border

    # Empty row
    cell = ws.cell(row=row_current + 3, column=col)
    cell.border = border

    # ROOM TYPES header
    cell = ws.cell(row=row_current + 4, column=col)
    cell.border = border

    # Room types
    room_row_start = row_current + 5
    for idx, room in enumerate(hotel_data["Room Types"]):
        base_row = room_row_start + (idx * 4)

        # Room name
        cell = ws.cell(row=base_row, column=col)
        cell.value = room["Name"]
        cell.alignment = wrap_alignment
        cell.border = border
        cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

        # Size
        cell = ws.cell(row=base_row + 1, column=col)
        cell.value = room["Size"]
        cell.alignment = wrap_alignment
        cell.border = border

        # Features
        cell = ws.cell(row=base_row + 2, column=col)
        cell.value = room["Features"]
        cell.alignment = wrap_alignment
        cell.border = border

        # Empty row
        cell = ws.cell(row=base_row + 3, column=col)
        cell.border = border

    # Fill remaining room slots with empty cells
    for idx in range(len(hotel_data["Room Types"]), 11):
        base_row = room_row_start + (idx * 4)
        for offset in range(4):
            cell = ws.cell(row=base_row + offset, column=col)
            cell.border = border

    # Key amenities
    amenities_row = room_row_start + (11 * 4)
    cell = ws.cell(row=amenities_row, column=col)
    cell.value = hotel_data["Key Amenities"]
    cell.alignment = wrap_alignment
    cell.border = border

    # Set column width
    ws.column_dimensions[get_column_letter(col)].width = 35

    col_offset += 1

# Freeze panes
ws.freeze_panes = 'B4'

# Auto-fit row heights
for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
    for cell in row:
        if cell.value and len(str(cell.value)) > 50:
            ws.row_dimensions[cell.row].height = None

# Save the workbook
output_file = r"C:\Users\reservations\Desktop\Compset Tool\Hotel_Room_Types_Comparison.xlsx"
wb.save(output_file)
print(f"Excel file created successfully: {output_file}")
