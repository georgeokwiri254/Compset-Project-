import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import os

# Load filtered hotels
with open(r'C:\Users\reservations\Desktop\Compset Tool\filtered_hotels.json', 'r', encoding='utf-8') as f:
    filtered_hotels = json.load(f)

# Load the scoring metrics
with open(r'C:\Users\reservations\Desktop\Compset Tool\Compset Analysis Tool.json', 'r', encoding='utf-8') as f:
    content = f.read()
    json_start = content.find('{')
    json_end = content.rfind('}') + 1
    scoring_data = json.loads(content[json_start:json_end])

# Check if Excel file exists
excel_file = r'C:\Users\reservations\Desktop\Compset Tool\Grand_Millennium_Dubai_Updated_CompSet_Analysis_2025.xlsx'

try:
    # Try to load existing Excel file
    wb = load_workbook(excel_file)
    print(f"Loaded existing Excel file: {excel_file}")
    print(f"Existing sheets: {wb.sheetnames}")

    # Check if there's already a Hotels sheet
    if 'Hotels' in wb.sheetnames:
        ws = wb['Hotels']
        print(f"Found 'Hotels' sheet with {ws.max_row} rows")
    else:
        ws = wb.create_sheet('Hotels', 0)
        print("Created new 'Hotels' sheet")

except FileNotFoundError:
    print(f"Excel file not found. Will check other versions...")
    # Try the other version
    excel_file = r'C:\Users\reservations\Desktop\Compset Tool\Grand_Millennium_Dubai_Updated_CompSet_Analysis_2025_1.xlsx'
    try:
        wb = load_workbook(excel_file)
        print(f"Loaded existing Excel file: {excel_file}")
        print(f"Existing sheets: {wb.sheetnames}")
    except FileNotFoundError:
        print("No existing Excel file found. Creating new one...")
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = 'Hotels'
        excel_file = r'C:\Users\reservations\Desktop\Compset Tool\Grand_Millennium_Dubai_CompSet_Analysis.xlsx'

# Get or create the Hotels sheet
if 'Hotels' not in wb.sheetnames:
    ws = wb.create_sheet('Hotels', 0)
else:
    ws = wb['Hotels']

# Clear existing content (optional - comment out if you want to append)
ws.delete_rows(1, ws.max_row)

# Define column headers based on compset analysis requirements
headers = [
    'ID', 'Hotel Name', 'Star Rating', 'Category', 'Distance (km)', 'Area',
    'Total Rooms', 'Room Types', 'Has Apartments', 'Apartment Count',
    'Executive Lounge', 'Meeting Space (sqm)', 'Ballroom Capacity',
    'Meeting Rooms Count', 'Pool', 'Spa', 'Gym', 'Restaurants Count',
    'Unique Features', 'TripAdvisor Rating', 'TripAdvisor Reviews',
    'Booking.com Rating', 'Booking.com Reviews', 'ADR (AED)', 'BAR (AED)',
    'Business Mix %', 'Leisure Mix %', 'MICE Mix %',
    'Corporate Accounts', 'Brand Affiliation', 'Loyalty Program',
    'Sustainability Cert', 'Technology Level', 'Overall Score', 'Recommendation'
]

# Add headers with styling
header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)
header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment

# Add filtered hotels to the sheet
for idx, hotel in enumerate(filtered_hotels, 2):
    ws.cell(row=idx, column=1, value=hotel['id'])
    ws.cell(row=idx, column=2, value=hotel['name'])
    ws.cell(row=idx, column=3, value=hotel['star_rating'])
    ws.cell(row=idx, column=4, value=hotel['category'])
    ws.cell(row=idx, column=5, value=hotel['distance_km'])
    ws.cell(row=idx, column=6, value=hotel['area'])

    # Initialize other columns with placeholder values
    for col in range(7, len(headers) + 1):
        ws.cell(row=idx, column=col, value='TBD')

# Adjust column widths
column_widths = {
    'A': 8, 'B': 40, 'C': 12, 'D': 12, 'E': 14, 'F': 20,
    'G': 12, 'H': 30, 'I': 15, 'J': 15, 'K': 15, 'L': 18,
    'M': 18, 'N': 18, 'O': 10, 'P': 10, 'Q': 10, 'R': 16,
    'S': 30, 'T': 18, 'U': 18, 'V': 18, 'W': 18, 'X': 14,
    'Y': 14, 'Z': 14, 'AA': 14, 'AB': 14, 'AC': 20, 'AD': 18,
    'AE': 18, 'AF': 18, 'AG': 14, 'AH': 15, 'AI': 15
}

for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# Freeze first row
ws.freeze_panes = 'A2'

# Save the workbook
output_file = r'C:\Users\reservations\Desktop\Compset Tool\Grand_Millennium_Dubai_CompSet_Analysis.xlsx'
wb.save(output_file)
print(f"\nExcel file created successfully: {output_file}")
print(f"Total hotels added: {len(filtered_hotels)}")
print(f"Columns: {len(headers)}")
